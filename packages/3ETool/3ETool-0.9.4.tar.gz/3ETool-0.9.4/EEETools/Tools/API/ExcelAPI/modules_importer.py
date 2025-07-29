from EEETools.Tools.API.Tools.main_tools import get_result_data_frames, update_exergy_values, get_debug_data_frames
from EEETools.Tools.API.DatAPI.modules_importer import export_dat
from EEETools.MainModules.main_module import CalculationOptions
from openpyxl import Workbook, load_workbook, styles, utils
from EEETools.MainModules.main_module import ArrayHandler
from datetime import date, datetime
import math, pandas, os
from io import BytesIO
import warnings


def calculate_excel(excel_path, calculation_option=None, new_exergy_list=None, export_solution=True):

    array_handler = import_excel_input(excel_path, calculation_option)

    if new_exergy_list is not None:

        array_handler = update_exergy_values(array_handler, new_exergy_list)

    array_handler.calculate()

    if export_solution:

        export_solution_to_excel(excel_path, array_handler)

    return array_handler


def convert_excel_to_dat(excel_path: str):

    array_handler = import_excel_input(excel_path)

    if ".xlsm" in excel_path:

        dat_path = excel_path.replace(".xlsm", ".dat")

    elif ".xlsx" in excel_path:

        dat_path = excel_path.replace(".xlsx", ".dat")

    else:

        dat_path = excel_path.replace(".xls", ".dat")

    export_dat(dat_path, array_handler)


def import_excel_input(excel_path, calculation_option=None) -> ArrayHandler:

    with warnings.catch_warnings():

        warnings.simplefilter("ignore")

        __check_excel_version(excel_path)
        array_handler = ArrayHandler()

        if calculation_option is not None and type(calculation_option) == CalculationOptions:

            array_handler.options = calculation_option

        # import connections
        excel_connection_data = pandas.read_excel(excel_path, sheet_name="Stream")

        for line in excel_connection_data.values:

            line = line.tolist()
            if not math.isnan(line[0]):
                new_conn = array_handler.append_connection()

                new_conn.index = line[0]
                new_conn.name = str(line[1])
                new_conn.exergy_value = line[2]

        # import blocks
        excel_block_data = pandas.read_excel(excel_path, sheet_name="Componenti")

        for line in excel_block_data.values:

            line = line.tolist()

            if not (math.isnan(line[0]) or type(line[0]) is str):

                if line[0] > 0:

                    if "Heat Exchanger" in str(line[2]) or "Scambiatore" in str(line[2]):

                        new_block = array_handler.append_block("Heat Exchanger")
                        excel_connection_list = list()
                        excel_connection_list.append(str(line[2]))
                        excel_connection_list.extend(line[5:])

                    else:

                        new_block = array_handler.append_block(str(line[2]))
                        excel_connection_list = line[5:]

                    new_block.index = line[0]
                    new_block.name = str(line[1])
                    new_block.comp_cost = line[3]

                    new_block.initialize_connection_list(excel_connection_list)

                else:

                    __append_excel_costs_and_useful_output(array_handler, line[5:-1], line[0] == 0, line[3])

        return array_handler


def __append_excel_costs_and_useful_output(handler, input_list, add_useful_output, input_cost):

    for elem in input_list:

        new_conn = handler.find_connection_by_index(elem)

        if not new_conn is None:

            if add_useful_output:
                new_conn.is_useful_effect = True

            else:
                new_conn.rel_cost = input_cost


def export_solution_to_excel(excel_path, array_handler: ArrayHandler, return_bytes=False):

    result_df = get_result_data_frames(array_handler)
    return __write_excel_file(excel_path, result_df, return_bytes)


def export_debug_info_to_excel(excel_path, array_handler: ArrayHandler, return_bytes=False):

    result_df = get_debug_data_frames(array_handler)
    return __write_excel_file(excel_path, result_df, return_bytes)


def __write_excel_file(excel_path, result_df, return_bytes):

    today = date.today()
    now = datetime.now()
    today_str = today.strftime("%d %b")
    now_str = now.strftime("%H.%M")

    output = BytesIO()

    if (not os.path.isfile(str(excel_path))) or return_bytes:
        wb = Workbook()

    else:
        wb = load_workbook(excel_path)

    if "Sheet" in wb.sheetnames:
        std = wb["Sheet"]
        wb.remove(std)

    for idx, key in enumerate(result_df.keys()):

        __write_excel_sheet(
            wb,
            sheet_name=(key + " - " + today_str + " - " + now_str),
            data_frame=result_df[key]
        )

        if not return_bytes:
            wb.save(excel_path)

    if return_bytes:
        wb.save(output)
        output.seek(0)
        return output.getvalue()


def __write_excel_sheet(wb: Workbook, sheet_name, data_frame: dict):

    data_frame = __convert_result_data_frames(data_frame)

    if not sheet_name in wb.sheetnames:
        wb.create_sheet(sheet_name)

    sheet = wb[sheet_name]

    col = 2
    for key in data_frame.keys():

        row = 2
        sub_data_frame = data_frame[key]
        n_sub_element = len(sub_data_frame["unit"])

        if key == "Name":

            column_dimension = 35

        elif key == "Stream":

            column_dimension = 8

        else:

            column_dimension = 20

        if n_sub_element == 0:

            sheet.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col)
            n_sub_element = 1

        else:

            if n_sub_element > 1:
                sheet.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + n_sub_element - 1)

            for n in range(n_sub_element):
                row = 3
                cell = sheet.cell(row, col + n, value="[" + sub_data_frame["unit"][n] + "]")
                cell.alignment = styles.Alignment(horizontal="center", vertical="center")
                cell.font = styles.Font(italic=True, size=10)

        cell = sheet.cell(2, col, value=key)
        cell.alignment = styles.Alignment(horizontal="center", vertical="center")
        cell.font = styles.Font(bold=True)

        for n in range(n_sub_element):

            row = 5
            data_list = (sub_data_frame["values"])[n][0]
            sheet.column_dimensions[utils.get_column_letter(col)].width = column_dimension

            for data in data_list:
                sheet.cell(row, col, value=data)
                row += 1

            col += 1


def __convert_result_data_frames(data_frame: dict) -> dict:

    new_data_frame = dict()

    for key in data_frame.keys():

        sub_dict = __get_sub_dict(key)
        sub_dict.update({"values": list()})

        if not sub_dict["name"] in new_data_frame.keys():

            new_data_frame.update({sub_dict["name"]: sub_dict})

        else:

            (new_data_frame[sub_dict["name"]])["unit"].append(sub_dict["unit"][0])

        (new_data_frame[sub_dict["name"]])["values"].append([data_frame[key]])

    return new_data_frame


def __get_sub_dict(key):

    if "[" in key:

        parts = key.split("[")

        name = parts[0].replace("[", "")
        measure_unit = [parts[1].split("]")[0].replace("]", "")]

    else:

        name = key
        measure_unit = list()

    return {"name": name, "unit": measure_unit}


def __check_excel_version(excel_path):

    pass