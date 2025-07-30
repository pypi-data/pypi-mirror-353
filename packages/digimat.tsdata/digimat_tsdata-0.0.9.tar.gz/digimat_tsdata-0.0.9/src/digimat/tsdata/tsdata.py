import csv
import io
import time
import struct
import re
from datetime import datetime, timedelta
import logging
import logging.handlers

from importlib.metadata import version
import requests
from urllib3.exceptions import InsecureRequestWarning
requests.packages.urllib3.disable_warnings(category=InsecureRequestWarning)

import pandas as pd
from openpyxl import Workbook, load_workbook
# import matplotlib.pyplot as plt

from digimat.units import Units


class TSDataLogger():
    def __init__(self, title="TSDATA"):
        self._title=title

    def create(self):
        return logging.getLogger(self._title)

    def tcp(self, level=logging.DEBUG, host='localhost'):
        logger=self.create()
        logger.setLevel(level)
        handler = logging.handlers.SocketHandler(host, logging.handlers.DEFAULT_TCP_LOGGING_PORT)
        logger.addHandler(handler)
        return logger

    def null(self):
        logger=self.create()
        logger.setLevel(logging.ERROR)
        handler=logging.NullHandler()
        logger.addHandler(handler)
        return logger


class TSDataQuery():
    def __init__(self, tsdata, accept='*/*'):
        self._tsdata=tsdata
        self._accept=accept

    @property
    def logger(self):
        return self._tsdata.logger

    def _validate(self):
        return True

    def _build(self):
        raise NotImplementedError

    def _decode(self, response):
        return response.text

    def acceptCsv(self):
        self._accept='text/csv'

    def acceptBinary(self):
        self._accept='application/octet-stream'

    def query(self, accept=None):
        try:
            if self._tsdata:
                if self._validate():
                    request=self._build()
                    response=self._tsdata.query(request, accept=accept or self._accept)
                    if response is not None:
                        tstart=time.time()
                        data=self._decode(response)
                        self.logger.debug('decoding time %.1fs' % (time.time()-tstart))
                        return data
        except:
            self.logger.excpetion('query()')

        return None


class TSDataQueryAreas(TSDataQuery):
    def __init__(self, tsdata):
        super().__init__(tsdata)
        self._areas=None

    def _build(self):
        return 'LIST AREA'

    def areas(self):
        if not self._areas:
            self.query()
        return self._areas

    def _decode(self, response):
        try:
            # contentType=response.headers['content-type']
            self._areas=response.text.splitlines()
            return self._areas
        except:
            self.logger.exception(f'{__class__.__name__}._decode()')


class TSDataQueryBrowseRecords(TSDataQuery):
    def __init__(self, tsdata, area):
        super().__init__(tsdata)
        self._area=area
        self._keys=None

    def _build(self):
        return f'LIST TS AREA {self._area}'

    def keys(self):
        if not self._keys:
            self.query()
        return self._keys

    def _decode(self, response):
        try:
            # contentType=response.headers['content-type']
            self._keys=response.text.splitlines()
            return self._keys
        except:
            self.logger.exception(f'{__class__.__name__}._decode()')


class TSDataQueryCountRecords(TSDataQuery):
    def __init__(self, tsdata, key):
        super().__init__(tsdata)
        self._key=key

    def _build(self):
        return f'COUNT RECORD {self._key}'

    def _decode(self, response):
        try:
            # contentType=response.headers['content-type']
            count=0
            try:
                count=response.json()['value']
            finally:
                return count
        except:
            self.logger.exception(f'{__class__.__name__}._decode()')


class TSDataQueryRecords(TSDataQuery):
    def __init__(self, tsdata, key):
        super().__init__(tsdata)
        self._key=key
        self._df=None
        self.reset()

    def reset(self):
        """Reset (clear) the stored record (the query)"""
        self._df=None
        self._utc=False
        self._condition=None
        self._asc=False
        self._limit=0
        self._offset=None

    @property
    def key(self):
        """Return the key of the record"""
        return self._key

    def where(self, condition):
        """Specify the where condition of the query"""
        self.reset()
        self._condition=condition
        return self

    def timein(self, condition):
        """Specify the 'WHERE time IN' condition of the query"""
        self.reset()
        return self.where(f'time IN {condition}')

    def __getitem__(self, key):
        return self.timein(key)

    def yp(self, n=1):
        """Year Period condition (WHERE time in ypN, N=1 by default)"""
        if n>1:
            return self.timein(f'yp{n}')
        return self.timein('yp')

    def mp(self, n=1):
        """Month Period condition (WHERE time in mpN, N=1 by default)"""
        if n>1:
            return self.timein(f'mp{n}')
        return self.timein('mp')

    def wp(self, n=1):
        """Week Period condition (WHERE time in wpN, N=1 by default)"""
        if n>1:
            return self.timein(f'wp{n}')
        return self.timein('wp')

    def dp(self, n=1):
        """Day Period condition (WHERE time in dpN, N=1 by default) """
        if n>1:
            return self.timein(f'dp{n}')
        return self.timein('dp')

    def y(self, n=None):
        """Year Period condition"""
        if n is None:
            return self.timein('y')
        if isinstance(n, int):
            if n==0:
                return self.timein('y')
            else:
                return self.timein(f'y{n}')
        return self.timein(f'yd{n}')

    def m(self, n=None):
        """Month Period condition"""
        if n is None:
            return self.timein('m')
        if isinstance(n, int):
            if n==0:
                return self.timein('m')
            else:
                return self.timein(f'm{n}')
        return self.timein(f'm.{n}')

    def w(self, n=None):
        """Week Period condition"""
        if n is None:
            return self.timein('w')
        if isinstance(n, int):
            if n==0:
                return self.timein('w')
            else:
                return self.timein(f'w{n}')

    def d(self, n=None):
        """Day Period condition"""
        if n is None:
            return self.timein('d')
        if isinstance(n, int):
            if n<0:
                return self.timein(f'd{n}')
            elif n>0:
                return self.timein(f'dp{n}')
        return self.timein('d')

    def p(self, date):
        """Generic Period condition (WHERE time in pX)
        period:
           date
           date1-date2
        date:
           j.m
           j.m.a
           """
        return self.timein(f'p{date}')

    def f(self):
        """Full Period condition (WHERE time in f)"""
        return self.timein('f')

    def time(self, condition):
        """Specify the 'WHERE time' condition of the query"""
        return self.where(f'time {condition}')

    def fromto(self, stampFrom, stampTo):
        """Specify the 'WHERE time >= stampFrom [<= stampTo]' condition of the query"""
        return self.time(f'>= {stampFrom} <= {stampTo}')

    def utc(self):
        self.reset()
        self._utc=True
        return self

    def asc(self):
        """Specify the ASCending sort condition of the query"""
        self.reset()
        self._asc=True
        return self

    def desc(self):
        """Specify the DECCending sort condition of the query"""
        self.reset()
        self._asc=False
        return self

    def limit(self, count, offset=0):
        """Specify the LIMIT count [,offset] condition of the query"""
        self.reset()
        self._limit=count
        self._offset=offset
        return self

    def _validate(self):
        return True

    def _build(self):
        request=f'GET TS {self.key}'
        if self._utc:
            request += ' UTC'
        if self._condition:
            request += f' WHERE {self._condition}'
        if self._asc:
            request += ' ASC'
        else:
            request += ' DESC'
        if self._limit:
            request += ' LIMIT'
            if self._offset:
                request += f' {self._offset},'
            request += f' {self._limit}'

        return request

    def _loadData(self, stamps, values, units, strunits, flags, tz=None):
        data={'value': values, 'unit': units, 'strunit': strunits, 'flags': flags}
        index=pd.DatetimeIndex(stamps, dtype='datetime64[ns]', tz=tz)
        self._df=pd.DataFrame(data, index=index)
        self._df.index.name=self.key
        return self._df

    def __repr__(self):
        return f'{self.__class__.__name__}(key={self.key}, {self._build()}, {self.count()} records)'

    def df(self):
        """"Execute the query (if needed) and return the corresponding pandas dataframe (df)"""
        try:
            if self._df is None:
                self.query()
            return self._df
        except:
            pass

    def dataframe(self):
        """"Execute the query (if needed) and return the corresponding pandas dataframe (df)"""
        return self.df()

    def s(self, df=None):
        """Execute the query (if needed) and return the corresponding numpy serie (df.value)"""
        try:
            if df is None:
                df=self.df()
            return self.df()['value']
        except:
            pass

    def serie(self, df=None):
        """Execute the query (if needed) and return the corresponding numpy serie (df.value)"""
        return self.s(df)

    def values(self):
        """Execute the query (if needed) and return the corresponding numpy serie (df.value)"""
        return self.serie()['value']

    def indexes(self):
        try:
            return self.df().index
        except:
            pass

    def count(self):
        try:
            return self._df.count()[0]
        except:
            return 0

    def __len__(self):
        return self.count()

    def __iter__(self):
        return self.df().iterrows()

    def min(self):
        try:
            return self.df().min()[0]
        except:
            pass

    def max(self):
        try:
            return self.df().max()[0]
        except:
            pass

    def dtmin(self):
        try:
            return self.df().index.min()
        except:
            pass

    def dtmax(self):
        try:
            return self.df().index.min()
        except:
            pass

    def window(self):
        return (self.dtmin(), self.min(), self.dtmax(), self.max())


class TSDataQueryCsvRecords(TSDataQueryRecords):
    def __init__(self, tsdata, key):
        super().__init__(tsdata, key)
        self.acceptCsv()

    def _decode(self, response):
        # TODO: could be optimized using builtin DataFrame.read_csv()
        try:
            u=Units()

            self.reset()
            contentType=response.headers['content-type']
            if contentType=='text/csv':
                response.encoding='UTF-8'
                fcsv=io.StringIO(response.text)
                csvreader=csv.reader(fcsv, delimiter=',')
                next(csvreader)  # skip csv headers

                stamps=[]
                values=[]
                flags=[]
                units=[]
                strunits=[]

                for row in csvreader:
                    stamps.append(datetime.fromisoformat(row[0]))
                    values.append(float(row[1]))
                    units.append(row[2])
                    strunits.append(u.str(row[2]))
                    flags.append(row[3])

                self.logger.debug(f'{self.key}: retrieved {len(stamps)} records')
                self._df=self._loadData(stamps, values, units, strunits, flags)
                return self._df
        except:
            self.logger.exception(f'{__class__.__name__}._decode()')


class TSDataQueryBinaryRecords(TSDataQueryRecords):
    def __init__(self, tsdata, key):
        super().__init__(tsdata, key)
        self.acceptBinary()

    def _decode(self, response):
        try:
            u=Units()
            self.reset()
            contentType=response.headers['content-type']
            if contentType=='application/octet-stream':
                dtRef=datetime(1, 1, 1)
                rformat='<QdHBB'
                rsize=struct.calcsize(rformat)
                rcount=len(response.content)/rsize

                stamps=[]
                values=[]
                flags=[]
                units=[]
                strunits=[]

                pos=0
                n=0
                while n<rcount:
                    # Read RECORD (TTTTTTTTVVVVVVVVUUF), T=msec since reference
                    # The first bit (left most) of the T field indicate the time format : 1=LOCALTIME, 0=UTC
                    data=struct.unpack(rformat, response.content[pos:pos+rsize])
                    dt=data[0] & 0x7fffffffffffffff
                    date=dtRef+timedelta(milliseconds=dt)
                    stamps.append(date)
                    values.append(data[1])
                    units.append(data[2])
                    strunits.append(u.str(data[2]))
                    flags.append(data[3])
                    pos+=rsize
                    n+=1

                self.logger.debug(f'{self.key}: retrieved {len(stamps)} records')
                self._df=self._loadData(stamps, values, units, strunits, flags)
                return self._df
        except:
            self.logger.exception(f'{__class__.__name__}._decode()')


class TSData():
    def __init__(self, url=None, timeout=3.0, token=None, user=None, logger=None, debug=False):
        self._debug=debug
        if logger is None:
            logger=TSDataLogger().tcp()
        self._logger=logger

        self._url=None
        self._token=None
        self._user=None
        self.setApiUrl(url)
        self.setApiToken(token, user)
        self._timeout=timeout

        self.logger.info(f'using tsdata version {self.version}')
        self.logger.debug(f'using requests version {requests.__version__}')

    def _getUrl(self, service):
        if self._url:
            return f'{self._url}/{service}'

    def setApiUrl(self, url):
        if url:
            self.logger.debug(f'using api {url}')
            self._url=url

    def setApiToken(self, token, user=None):
        if token:
            self._token=token
            self._user=user

    def debug(self, state=True):
        self._debug=state

    def nodebug(self):
        self.debug(False)

    def isDebug(self):
        if self._debug:
            return True
        return False

    def getVersion(self):
        try:
            return version("digimat.tsdata")
        except:
            pass

    @property
    def version(self):
        return self.getVersion()

    @property
    def logger(self):
        return self._logger

    def __getitem__(self, key):
        return self.ts(key)

    def dump(self):
        pass

    def __repr__(self):
        return f'{self.__class__.__name__}({self._url}:{self._user})'

    def _get(self, service, params=None, accept='*/*'):
        url=self._getUrl(service)
        if url and self._token:
            headers={'x-api-key': self._token, 'Accept': accept}
            try:
                tstart=time.time()
                r=requests.get(url, headers=headers, params=params, verify=False)
                self.logger.debug('GET(%s) %.1fs %s:%dkb' % (r.request.url, time.time()-tstart, r.headers['content-type'], len(r.content)/1024))
                if r and r.status_code==requests.codes.ok:
                    return r
            except:
                self.logger.exception(f'GET({url})')
                # self.logger.error(f'GET({url})')
        else:
            self.logger.error('api url invalid')

    def query(self, query, accept='*/*'):
        """Execute the tsdata query"""
        try:
            return self._get('ts/query', params={'query': query}, accept=accept)
        except:
            pass

    def ts(self, key):
        """Return a TS (TimeSerie) query for the given key"""
        return TSDataQueryBinaryRecords(self, key)

    def s(self, key, timein=None):
        """Return a numpy serie for the given key and time filter"""
        try:
            ts=self.ts(key)
            if timein:
                ts.timein(timein)
            return ts.serie()
        except:
            pass

    def serie(self, key, timein=None):
        """Return a numpy serie for the given key and time filter"""
        return self.s(key, timein)

    def df(self, key, timein=None, sample=None):
        """Return a pandas dataframe for the given key and time filter"""
        try:
            ts=self.ts(key)
            if timein:
                ts.timein(timein)
            df=ts.df()
            if sample:
                df=self.sample(df, sample)
            return df
        except:
            pass

    def areas(self):
        """Return the available list of areas"""
        return TSDataQueryAreas(self).query()

    def browse(self, area):
        """Return the available ts for the given area"""
        return TSDataQueryBrowseRecords(self, area).query()

    def count(self, key):
        """Return the number of records for the given key"""
        return TSDataQueryCountRecords(self, key).query()

    def sample(self, df, period):
        # https://pandas.pydata.org/pandas-docs/stable/user_guide/timeseries.html#dateoffset-objects
        m=re.search(r'(\d)?([qhdwmy])', period)
        count=m.group(1)
        period=m.group(2)

        if period=='q':
            period='15T'
        elif period=='h':
            if count:
                period='%dh' % int(count)
            else:
                period='h'
        elif period=='d':
            if count:
                period='%dD' % int(count)
            else:
                period='D'
        elif period=='w':
            if count:
                period='%dW' % int(count)
            else:
                period='W'
        elif period=='m':
            if count:
                period='%dMS' % int(count)
            else:
                period='MS'
        elif period=='y':
            if count:
                period='%dAS' % int(count)
            else:
                period='AS'

        # https://dataindependent.com/pandas/pandas-resample-pd-df-resample/
        # return df.resample(period, convention='start').first()
        return df.resample(period).first()

    def conso(self, df, period):
        df=self.sample(df, period)
        # make a diff on first column (values)
        df.iloc[:, 0] = df.iloc[:, 0].diff()
        return df

    def xlsx(self, fpath,  df):
        if df is not None:
            if not isinstance(df, list):
                df=[df]

            with pd.ExcelWriter(fpath, mode='a', if_sheet_exists='replace', engine='openpyxl') as writer:
                for d in df:
                    d.to_excel(writer, sheet_name=d.index.name)

    def xlsxProcessData(self, fpath, period='w', differential=False):
        data=[]
        wb=load_workbook(fpath)
        if wb:
            ws=wb['digimat']
            if ws:
                row=1
                while True:
                    tag=ws.cell(row=row, column=1).value
                    name=ws.cell(row=row, column=2).value
                    key=ws.cell(row=row, column=5).value
                    if not key:
                        break

                    print(tag, key, name)
                    print('Retrieve data...')
                    df=self.ts(key).yp(3).query()
                    df=self.sample(df, period)
                    if differential:
                        df.diff()
                    print(df)
                    data.append(df)

                    row+=1
            wb.close()

        if data:
            print('Writing xls...')
            self.xlsx(fpath, data)


if __name__ == "__main__":
    pass
