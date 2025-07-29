import os
import re
import json
import time
import shutil
import string
import logging
import subprocess
import pandas as pd
from typing import Callable
from datetime import datetime, date, timedelta


def read_or_create_file_txt(
        text: str,
        name_file: str,
        path_to_save: str | os.PathLike,
        subs: bool = False
        ) -> str:
    os.makedirs(path_to_save, exist_ok=True)
    full_path_file = os.path.abspath(f'{path_to_save}/{name_file}.txt')

    if not os.path.exists(full_path_file) or subs:
        with open(full_path_file, 'w') as f:
            f.write(text)
    else:
        with open(full_path_file, 'r') as f:
            text = f.read()
    return text


def read_or_create_file_bin(
        content: str | bytes,
        name_file: str,
        path_to_save: str | os.PathLike,
        subs: bool = False,
        encoding='utf-8'
        ) -> bytes:
    os.makedirs(path_to_save, exist_ok=True)
    full_path_file = os.path.abspath(f'{path_to_save}/{name_file}.bin')

    if isinstance(content, bytes):
        content = content.decode(encoding)
    if isinstance(content, str):
        content = content.encode(encoding)

    if not os.path.exists(full_path_file) or subs:
        with open(full_path_file, 'wb') as f:
            f.write(content)
    else:
        with open(full_path_file, 'rb') as f:
            content = f.read()
    return content


def read_or_create_file_json(
        dict_file: dict,
        name_file: str,
        path_to_save: str | os.PathLike,
        subs: bool = False
        ) -> dict:
    os.makedirs(path_to_save, exist_ok=True)
    full_path_file = os.path.abspath(f'{path_to_save}{name_file}.json')

    if not os.path.exists(full_path_file) or subs:
        with open(full_path_file, 'w') as f:
            json.dump(dict_file, f)
    else:
        with open(full_path_file, "r") as readit:
            dict_file = json.load(readit)

    return dict_file


def get_dummy() -> int:
    timestamp = time.time()
    dummy = int(timestamp * 1000)
    return dummy


def move_folder(
        folder_copy: str | os.PathLike,
        folder_write: str | os.PathLike
        ) -> None:
    shutil.copytree(folder_copy, folder_write)


def delete_files_folder(folder: str | os.PathLike) -> None:
    if os.path.exists(folder):
        shutil.rmtree(folder)
        os.makedirs(folder, exist_ok=True)
    else:
        os.makedirs(folder, exist_ok=True)


def delete_folder(folder: str | os.PathLike) -> None:
    shutil.rmtree(folder)


def delete_file(path_file: str | os.PathLike) -> None:
    if os.path.isfile(path_file):
        os.remove(path_file)
    else:
        raise OSError(
            "The path provided is not a existing file path"
        )


def move_file(
        from_path_file: str | os.PathLike,
        to_path_file: str | os.PathLike
        ) -> None:
    shutil.copy2(from_path_file, to_path_file)


def rename_file(
        from_path_file: str | os.PathLike,
        to_path_file: str | os.PathLike
        ) -> None:
    from pathlib import Path
    Path(from_path_file).rename(to_path_file)


def _ignore_unnamed_columns(df: pd.DataFrame) -> pd.DataFrame:
    return df.loc[:, ~df.columns.str.lower().str.contains('^unnamed')]


def _read_excel(file_path: str | os.PathLike, **kwargs) -> pd.DataFrame:
    # getting unnamend_columns value and removing it from kwargs
    unnamed_columns = kwargs.pop('unnamed_columns')
    try:
        df = pd.read_excel(file_path, **kwargs)
        if not unnamed_columns:
            df = _ignore_unnamed_columns(df)
        if 'print' in kwargs:
            logging.info(df.head())
        return df
    except ValueError as e:
        logging.error(e)


def _read_csv(file_path: str | os.PathLike, **kwargs) -> pd.DataFrame:
    try:
        # getting unnamend_columns value and removing it from kwargs
        unnamed_columns = kwargs.pop('unnamed_columns')
        # Define default sep
        sep = ';' if 'sep' not in kwargs else kwargs.pop('sep')
        df = pd.read_csv(file_path, sep=sep, **kwargs)
        if not unnamed_columns:
            df = _ignore_unnamed_columns(df)
        if 'print' in kwargs:
            logging.info(df.head())
        return df
    except ValueError as e:
        logging.error(e)


def _read_html(file_path: str | os.PathLike, **kwargs) -> pd.DataFrame:
    try:
        # getting unnamend_columns and number file values and removing it form
        # kwargs
        unnamed_columns = kwargs.pop('unnamed_columns')
        number_file = kwargs.pop('number_file')
        # Define default thousands
        thousands = '' if 'thousands' not in kwargs else kwargs.pop('thousands')

        # Define default decimal
        decimal = ',' if 'decimal' not in kwargs else kwargs.pop('decimal')

        # Define default flavor
        flavor = 'lxml' if 'flavor' not in kwargs else kwargs.pop('flavor')
        df = pd.read_html(
            file_path,
            thousands=thousands,
            decimal=decimal,
            flavor=flavor,
            **kwargs
            )
        if 'print' in kwargs:
            {logging.info(d) for d in df}
        df = df[number_file]
        if not unnamed_columns:
            df = _ignore_unnamed_columns(df)
        return df
    except ValueError as e:
        logging.error(e)


def _read_json(file_path: str | os.PathLike) -> dict:
    with open(file_path, "r") as readit:
        dict_file = json.load(readit)
    return dict_file


def _read_bin(file_path: str | os.PathLike) -> str:
    with open(file_path, "rb") as f:
        bytes_f = f.read()
        valor = bytes_f.decode()
    return valor


def _read_text(file_path: str | os.PathLike) -> str:
    with open(file_path, "r") as f:
        read_file = f.read()
    return read_file


def get_file_number_on_folder(folder: str | os.PathLike, index_file) -> str:
    files = os.listdir(folder)
    file_n = files[index_file]
    return os.path.join(folder, file_n)


def import_file(
        file_path: str | os.PathLike,
        **kwargs) -> pd.DataFrame | dict | str:
    """
    Import data from a file in various formats.

    :param file_path: (str) Path to the file to be imported.
    :param number_file: (int, optional) Index of the file to read if file_path
    is a directory. Default is 0.
    :param encoding: (str, optional) File encoding. Default is 'ISO-8859-1'.
    :param sep: (str, optional) Delimiter used in CSV files. Default is ';'.
    :param thousands: (str, optional) Thousands separator used in XLS files.
    Default is ''.
    :param decimal: (str, optional) Decimal separator used in XLS files.
    Default is ','.
    :param flavor: (str, optional) HTML parser flavor used in XLS files.
    Default is 'lxml'.
    :param unnamed_columns: (bool, optional) Whether to ignore unnamed columns
    in the imported data. Default is True.
    :param kwargs: Additional keyword arguments for specific file types.
    :return: (DataFrame, dict, str) Imported data, as a Pandas DataFrame, a
    dictionary or a string.

    This method imports data from a file in various formats, including Excel,
    CSV, XLS, JSON and text files. If the
    provided file_path is a directory, it reads the file at the specified
    index in the directory (default is the
    first file). The supported file formats are determined by the file
    extension, which is used to select the
    appropriate import function. Additional arguments can be passed as keyword
    arguments to the import functions,
    such as the delimiter used in CSV files or the encoding of the input file.
    By default, unnamed columns are
    ignored in the imported data.

    Example usage:
        >>> df = Handler.import_file('example.xlsx', sheet_name='Sheet1')
        >>> logging.info(df.head())
    """
    # Index file to read on the list os files on the path folder if file_path
    # provide only the path folder
    number_file = (0 if 'number_file' not in kwargs
                   else kwargs.pop('number_file'))
    read_file = None

    # Verify if file_path provide the file with extention
    if len(file_path.split('.')) > 1:

        # configure unnamed_columns parameter if  keep or delete columns with
        # the name unnamed in the name column
        unnamed_columns = (False if 'unnamed_columns' not in kwargs
                           else kwargs.pop('unnamed_columns'))

        # Get the extention
        ext = file_path.split('.')[-1]

        # Define default encoding
        encoding = ('ISO-8859-1' if 'encoding' not in kwargs
                    else kwargs.pop('encoding'))

        # Read excel
        if 'xlsx' == ext.lower():
            read_file = _read_excel(
                file_path,
                unnamed_columns=unnamed_columns,
                **kwargs
                )

        # Read csv
        elif 'csv' == ext.lower() or 'sswweb' == ext.lower():
            read_file = _read_csv(
                file_path,
                unnamed_columns=unnamed_columns,
                encoding=encoding,
                **kwargs
                )

        # Read sheet html
        elif 'xls' == ext.lower():
            read_file = _read_html(
                file_path,
                unnamed_columns=unnamed_columns,
                encoding=encoding,
                number_file=number_file,
                **kwargs
                )

        # Read json
        elif 'json' == ext.lower():
            read_file = _read_json(file_path)

        # Read binary
        elif 'bin' == ext.lower():
            read_file = _read_bin(file_path)

        # Try to read as text
        else:
            read_file = _read_text(file_path)

        return read_file
    else:
        file_path = get_file_number_on_folder(file_path, number_file)
        return import_file(file_path, **kwargs)


def to_csv(
        df: pd.DataFrame,
        name_file: str,
        folder_to_salve: str | os.PathLike,
        **kwargs
        ) -> None:
    sep = ';' if 'sep' not in kwargs else kwargs.pop('sep')
    index = False if 'index' not in kwargs else kwargs.pop('index')
    os.makedirs(folder_to_salve, exist_ok=True)
    df.to_csv(
        f'{folder_to_salve}/{name_file}.csv',
        sep=sep,
        index=index,
        **kwargs
        )


def to_excel(
        df: pd.DataFrame,
        name_file: str,
        folder_to_salve: str | os.PathLike,
        **kwargs
        ) -> None:
    index = False if 'index' not in kwargs else kwargs.pop('index')
    sheet_name = (name_file if 'sheet_name' not in kwargs
                  else kwargs.pop('sheet_name'))
    os.makedirs(folder_to_salve, exist_ok=True)
    df.to_excel(
        f'{folder_to_salve}/{name_file}.xlsx',
        sheet_name=sheet_name,
        index=index,
        **kwargs
        )


def float_converter(
        table: pd.DataFrame,
        column: str
        ) -> pd.DataFrame:
    table[column] = table[column].astype('str')
    table[column] = list(
        map(
            (lambda x: x.replace(' ', '').replace('.', '')
             .replace(',', '.').strip()),
            table[column].tolist())
            )
    table[column] = list(
        map(lambda x: str(format(float(x), ".2f")), table[column].tolist()))
    return table[column]


def adjust_encode_df_html(
        df: pd.DataFrame,
        encoding: str = 'latin-1',
        deconding: str = 'utf-8'
        ) -> pd.DataFrame:
    adjusted_columns = [
        col.encode(encoding).decode(deconding) for col in df.columns
        ]
    df.columns = adjusted_columns

    for column in df.columns:
        df[column] = list(
            map(lambda x: str(x).encode(encoding).decode(deconding),
                df[column])
                )
    return df


@staticmethod
def convert_thousands_decimal_formats(table, column):
    """
    Convert values in a specific column of a DataFrame from thousands and decimal format to a standard format.
    """
    def adjust_decimal(value):
        try:
            # Replace '.' (thousands separator) and ',' (decimal) to a standard format
            if ',' in value and '.' in value and value.index(',') > value.index('.'):
                value = value.replace('.', '').replace(',', '.')
            elif ',' in value:
                value = value.replace(',', '.')
            return value
        except Exception as e:
            raise e
            return value  # Return original value if errors occur

    table[column] = table[column].apply(lambda x: adjust_decimal(str(x)))
    return table[column]



def int_float_converter(table, column):
    """
    Convert values in a specific column of a DataFrame to either integers or floats, handling different formats.
    """
    try:
        table[column] = table[column].astype(str)
        table[column] = replace_nan_from_table(table, column, 0)
        table[column] = table[column].str.replace(' ', '').str.replace('#', '0').str.replace('*', '0').str.strip()
        table[column] = table[column].apply(lambda x: 0 if x == '' else x)
        table[column] = convert_thousands_decimal_formats(table, column)

        try:
            table[column] = table[column].astype('int')
        except ValueError:
            try:
                table[column] = table[column].astype('float')
            except ValueError:
                pass

        # Ensure consistent formatting for floats
        if table[column].dtype == 'float64':
            max_decimal_house = max(
                [len(str(val).split('.')[1]) for val in table[column] if '.' in str(val)] or [0]
            )
            table[column] = table[column].apply(lambda x: round(x, max_decimal_house))

        return table[column]

    except Exception as e:
        logging.error("Error converting column to int/float: %s", e)
        raise e



def clear_invalid_characters_from_list(
        lista: list,
        replace: str = ' '
        ) -> list:
    lista = [
        re.sub(r'\W+', replace, valor).strip()
        if len(re.sub(r'\W+', replace, valor)) > 0 else valor.strip()
        for valor in lista
        ]
    return lista


def clear_invalid_characters_from_str(
        string_to_clear: str,
        replace: str = ' '
        ) -> str:
    return re.sub(r'\W+', replace, string_to_clear).strip()


@staticmethod
def replace_nan_from_table(table, column, replacement=''):
    """
    Remove NaN values from a specific column of a DataFrame.
    """
    table[column] = table[column].apply(lambda x: replacement if pd.isna(x) else x)
    return table[column]


def replace_none_for_zero_on_column(table: pd.DataFrame, column: str) -> pd.Series:
    """
    Replaces 'None' strings in a column with '0'.

    Parameters:
    - table: DataFrame containing the data.
    - column: Name of the column to process.

    Returns:
    - Updated column as a pandas Series.
    """
    return table[column].apply(lambda x: '0' if str(x).lower() == 'none' else x)


def clear_table(table: pd.DataFrame, clear_char_in_list: list | None = None) -> pd.DataFrame:
    """
    Cleans the DataFrame by performing the following steps:
    - Fills NaN values with an empty string.
    - Converts all data to strings.
    - Replaces 'None' values with '0'.
    - Removes double spaces and trims strings.
    - Replaces 'NaT' with an empty string.
    - Optionally removes specific characters provided in `clear_char_in_list`.

    Parameters:
    - table: DataFrame to clean.
    - clear_char_in_list: List of characters to remove from columns (optional).

    Returns:
    - Cleaned DataFrame.
    """
    # Fill NaN values and convert to string
    table = table.fillna('').astype(str)

    for column in table.columns:
        # Replace 'None' with '0'
        table[column] = replace_none_for_zero_on_column(table, column)

        # Remove double spaces and trim
        table[column] = table[column].str.replace('  ', '').str.strip()

        # Replace 'NaT' with an empty string
        table[column] = table[column].replace('NaT', '')

        # Remove specified characters if provided
        if clear_char_in_list:
            for char in clear_char_in_list:
                table[column] = table[column].replace(char, '')

    return table


def match_all_white_spaces(string: str) -> str:
    pattern = r"\s"
    matches = re.findall(pattern, string)
    for m in matches:
        string = string.replace(m, '')
    return string


def replace_all_white_spaces(table: pd.DataFrame, column) -> pd.DataFrame:
    table[column] = list(
        map(lambda x: match_all_white_spaces(str(x)),
            table[column])
        )  # .tolist()
    return table[column]


def identify_numeric_columns(table: pd.DataFrame) -> list:
    table = table.copy()
    lista_columns_numericas = []
    for column in table.columns:

        table[column] = table[column].astype('str')
        table[column] = replace_all_white_spaces(table, column)
        table[column] = list(
            map(
                lambda x: x.lower().replace(' ', '').replace('.', '')
                .replace(',', '')
                .replace('#', '0').strip().replace('*', '0').strip()
                if len(str(x).strip()) > 0 else 0,
                table[column]))  # .tolist()
        lista_cel = table[column].astype('str').str.isnumeric()
        lista_cel = list(
            map(lambda x: str(x), lista_cel))
        if 'False' not in lista_cel:
            lista_columns_numericas.append(column)

    return lista_columns_numericas


def extract_file_from_rar(
        path_rar_file: str | os.PathLike,
        path_to_save: str | os.PathLike,
        patool_path: str | os.PathLike
        ) -> None:
    from pyunpack import Archive
    # Extract the rar file
    Archive(path_rar_file).extractall(path_to_save, patool_path=patool_path)


def extract_file_from_zip(
        path_zip_file: str | os.PathLike,
        folder_to_save: str | os.PathLike
        ) -> None:
    import zipfile
    # Open the zip file in read mode
    with zipfile.ZipFile(path_zip_file, 'r') as zip_ref:
        # Extract all the files
        zip_ref.extractall(folder_to_save)


def zip_folder(
        folder_path: str | os.PathLike,
        output_path: str | os.PathLike
        ) -> None:
    import zipfile
    with zipfile.ZipFile(
            output_path, 'w',
            zipfile.ZIP_DEFLATED
            ) as zip_file:
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, folder_path)
                zip_file.write(file_path, arcname)


def force_kill_process_by_pid(pid: int) -> None:
    os.system(f"taskkill /F /PID {pid}")


def kill_process(key_word: str = 'main_bot', search_by: str = 'name') -> None:
    import psutil
    for proc in psutil.process_iter(['pid', 'name', 'exe']):  #
        logging.info([proc.info[key] for key in ['pid', 'name', 'exe']])
        try:
            if key_word.lower() in str(proc.info[search_by]).lower():
                app_pid = proc.info['pid']
                # Terminate the app process
                app_process = psutil.Process(app_pid)
                app_process.kill()
        except psutil.AccessDenied:
            logging.info(
                (f"Access denied for process with PID {proc.info['pid']}. "
                 "Continuing...")
                )
            force_kill_process_by_pid(app_pid)


def hash_string(s: str) -> int:
    import hashlib
    return int(hashlib.sha256(s.encode('utf-8')).hexdigest(), 16)


def sweep(
        path: str | os.PathLike,
        file_to_seach,
        ignore_folders: list | None = None
        ) -> str | None:
    if not ignore_folders:
        ignore_folders = []
    if os.path.exists(path):
        try:
            contents = os.listdir(path)
        except Exception:
            return None

        files = [file for file in contents if '.' in str(file)]
        folders = [
            folder for folder in contents if '.' not in str(folder)
            and folder not in ignore_folders
            ]
        if file_to_seach in ''.join(files):
            path_file = os.path.abspath(os.path.join(path, file_to_seach))
            return path_file
        else:
            for folder in folders:
                path_child = os.path.join(path, folder)
                resp = sweep(path_child, file_to_seach)

                if resp is not None:
                    return resp

            return None
    else:
        return None


def divide_table(
        df: pd.DataFrame,
        qtde: int,
        name_file: str,
        folder_to_salve: str | os.PathLike
        ) -> None:
    tamanho = len(df)
    partes = int(tamanho / qtde)
    for i in range(qtde):
        cut_ini = i * partes
        cut_fim = (i + 1) * partes - 1
        if i == qtde - 1:
            table_aux = df.loc[cut_ini:]
        else:
            table_aux = df.loc[cut_ini:cut_fim]
        filename = f'{name_file}_part_{i + 1}'
        to_csv(table_aux, filename, folder_to_salve)


def get_str_format_datetime_by_switch_case(i: int) -> str:
    cases = {
        1: '%d/%m/%y',
        2: '%d/%m/%Y',
        3: '%m/%d/%y',
        4: '%m/%d/%Y',
        5: '%d/%m/%y %H:%M:%S',
        6: '%d/%m/%Y %H:%M:%S',
        7: '%m/%d/%y %H:%M:%S',
        8: '%m/%d/%Y %H:%M:%S',
        9: '%Y/%m/%d %H:%M:%S',
        10: '%y/%m/%d %H:%M:%S',
        11: '%d-%m-%y %H:%M:%S',
        12: '%d-%m-%Y %H:%M:%S',
        13: '%m-%d-%y %H:%M:%S',
        14: '%m-%d-%Y %H:%M:%S',
        15: '%Y-%m-%d %H:%M:%S',
        16: '%y-%m-%d %H:%M:%S',
        17: '%d/%m/%y %H:%M',
        18: '%d/%m/%Y %H:%M',
        19: '%m/%d/%Y %H:%M',
        20: '%m/%d/%Y %H:%M',
        21: '%Y/%m/%d',
        22: '%y/%m/%d',
        23: '%Y-%m-%d',
        24: '%Y-%m-%d',
    }
    return cases.get(i, 'Default Case')


def _convert_column_to_datetime_with_strptime(
        table: pd.DataFrame,
        column: str,
        format_d: str = "%d/%m/%y"
        ) -> pd.DataFrame:
    for i in range(24):
        try:
            table[column] = table[column].apply(
                (lambda x: datetime.strptime(x, format_d)
                 if (x != '') else 'NULL')
                )
            table[column] = pd.to_datetime(table[column], format=format_d)
            table[column] = list(
                map(
                    lambda x: x if (x is not None)
                    and len(str(x)) >= 8 else 'NULL',
                    table[column].tolist())
                    )
            break
        except Exception as e:
            format_d = get_str_format_datetime_by_switch_case(i + 1)
    return table[column]


def _convert_column_to_datetime_with_datetime64(
        table: pd.DataFrame,
        column: str,
        format_d: str = "%d/%m/%y"
        ) -> pd.DataFrame:
    for i in range(24):
        try:
            table[column] = table[column].astype('datetime64[ns]')
            table[column] = pd.to_datetime(table[column], format=format_d)
            table[column] = list(
                map(
                    lambda x: x if (len(str(x)) >= 10) else 'NULL',
                    table[column].tolist()
                    )
                    )
            break
        except Exception as e:
            format_d = get_str_format_datetime_by_switch_case(i + 1)
    return table[column]


def convert_column_to_datetime(
        table: pd.DataFrame,
        column: str
        ) -> pd.DataFrame:
    table[column] = replace_nan_from_table(table, column)

    # Trying many ways to convert the column to datetime
    try:
        table[column] = _convert_column_to_datetime_with_strptime(
            table,
            column
            )
    except Exception:
        try:
            table[column] = pd.to_datetime(table[column], format='%d/%m/%y')
        except Exception:
            try:
                table[column] = _convert_column_to_datetime_with_datetime64(
                    table,
                    column
                    )
            except Exception as e:
                pass

    # replacing empty values with 'NULL' value to be set on the database
    table[column] = list(
        map(
            lambda x: x if len(str(x)) >= 8 else 'NULL',
            table[column].tolist()
            )
        )

    return table[column]


def get_str_format_time_by_switch_case(i: int) -> str:
    cases = {
        1: '%H:%M:%S',
        2: '%H:%M',
    }
    return cases.get(i, 'Default Case')


def convert_column_to_time_with_strptime(
        table: pd.DataFrame,
        column: str,
        format_t: str = '%H:%M:%S'
        ) -> pd.DataFrame:
    for i in range(2):
        try:
            table[column] = table[column].apply(
                (lambda x: datetime.strptime(str(x).strip(), format_t)
                 if (x != '') else None)
                )
            table[column] = pd.to_datetime(table[column], format=format_t)
            break
        except Exception as e:
            format_t = get_str_format_time_by_switch_case(i + 1)
    # replacing empty values with 'NULL' value to be set on the database
    table[column] = list(
        map(
            lambda x: x if len(str(x)) >= 4 else 'NULL',
            table[column].tolist()
            )
            )
    return table[column]


def _convert_column_to_time_with_datetime64(
        table: pd.DataFrame,
        column: str,
        format_t: str = '%H:%M:%S'
        ) -> pd.DataFrame:
    for i in range(2):
        try:
            table[column] = pd.to_datetime(table[column], format=format_t)
            table[column] = table[column].astype('datetime64[ns]')
            table[column] = pd.to_datetime(
                table[column],
                format=format_t
                ).dt.time
            break
        except Exception as e:
            format_t = get_str_format_time_by_switch_case(i + 1)

    # replacing empty values with 'NULL' value to be set on the database
    table[column] = list(
        map(
            lambda x: x if len(str(x)) >= 4 else 'NULL',
            table[column].tolist()
            )
        )
    return table[column]


def _convert_column_to_time(table: pd.DataFrame, column: str) -> pd.DataFrame:
    # try many ways to convert the column to time format
    try:
        # converting with strptime function to format H:M:S
        table[column] = convert_column_to_time_with_strptime(table, column)
    except Exception as e:
        try:
            # converting with datetime64 parameter to format H:M
            table[column] = _convert_column_to_time_with_datetime64(
                table,
                column
                )
        except Exception as e:
            pass

    return table[column]


def convert_table_columns_to_datetime(
        table: pd.DataFrame,
        ignore_type_columns: list | None = None,
        dtypes: dict | None = None
        ) -> pd.DataFrame:
    for column in table.columns:
        column = str(column)
        # logging.info(column)
        # verify if name column seem like a name column date
        if (ignore_type_columns is None
                or column.lower() not in ignore_type_columns):
            if True in [
                True for arg in dtypes['datetime']
                    if arg.lower() in column.lower()
                    ]:
                table[column] = convert_column_to_datetime(table, column)

        # verify if name column seem like a name column time
        if (ignore_type_columns is None
                or column.lower() not in ignore_type_columns):
            if True in [
                    True for arg in dtypes['time']
                    if arg.lower() in column.lower()
                    ]:
                table[column] = _convert_column_to_time(table, column)
        # logging.info(table[column].info())
    return table


def convert_table_types(
        table: pd.DataFrame,
        ignore_type_columns: list | None = None,
        dtypes: dict | None = None
        ) -> pd.DataFrame:
    if dtypes is None:
        dtypes = {}
    if ignore_type_columns is not None:
        ignore_type_columns = [col.lower() for col in ignore_type_columns]

    dtypes['datetime'] = (['data', 'dia'] if 'datetime' not in dtypes.keys()
                          else dtypes['datetime'])
    dtypes['time'] = (['hora'] if 'time' not in dtypes.keys()
                      else dtypes['time'])

    # try to convert columns to datetime
    table = convert_table_columns_to_datetime(
        table,
        ignore_type_columns,
        dtypes
        )

    # try to identify and convert numeric columns
    numeric_columns = identify_numeric_columns(table)
    for column_n in numeric_columns:
        if (ignore_type_columns is None
                or column_n.lower() not in ignore_type_columns):
            try:
                table[column_n] = int_float_converter(table, column_n)
            except Exception as e:
                logging.error(f'ERROR COLUMN: {column_n}')
                raise e
    return table


def get_date_back(days: int) -> datetime:
    date_ = datetime.now()
    return datetime(date_.year, date_.month, date_.day) - timedelta(days)


def get_date_forward(days: int) -> datetime:
    date_ = datetime.now()
    return datetime(date_.year, date_.month, date_.day) + timedelta(days)


def _verify_month_range(month: str | int) -> None:
    # verify month range
    if int(month) not in range(1, 13):
        raise ValueError('Expected month value in range 1-12')


def _verify_year_format(year: str | int) -> str:
    # Verrify year format
    if len(str(year)) == 2:
        date_format = '%d/%m/%y'
    elif len(str(year)) == 4:
        date_format = '%d/%m/%Y'
    else:
        raise ValueError("Expected year format 'yy' or 'yyyy'")
    return date_format


def _verify_numeric_month_year(month: str | int, year: str | int) -> None:
    # verifying if the params are numeric
    if not str(month).isnumeric() or not str(year).isnumeric():
        raise ValueError('Numeric values expected!')


def get_full_range_days_from_month(
        month: str | int = 0,
        year: str | int = 0,
        format_date: str = ''
        ) -> tuple[str | datetime, str | datetime]:
    _verify_numeric_month_year(month, year)

    # get the month before
    month_before = (str(date.today().month - 1)
                    if str(date.today().month - 1) != '0' else '12')

    # if month is zero define as month before
    month = str(month) if str(month) != '0' else month_before

    # adjust the month length
    month = '0' + str(month) if len(month) < 2 else month

    # verify month range
    _verify_month_range(month)

    # define next month
    month_prox = str(int(month) + 1) if str(int(month) + 1) != '13' else '01'
    month_prox = '0' + str(month_prox) if len(month_prox) < 2 else month_prox

    # if year is zero define current year
    year = str(year) if str(year) != '0' else str(date.today().year)

    # Verrify year format
    date_format = _verify_year_format(year)

    # get the firt day month in datetime
    first_day_month = datetime.strptime(f'01/{month}/{year}', date_format)

    # verify if month_prox will lead to next year or not
    year = year if month_prox != '01' else str(int(year) + 1)
    year = '0' + str(year) if len(year) < 2 else year

    # get the firt day from next month
    first_day_prox_month = datetime.strptime(
        f'01/{month_prox}/{year}',
        date_format
        )

    # get the last day month
    last_day_month = first_day_prox_month - timedelta(1)

    if len(str(format_date)) > 0:
        last_day_month = string_format_br_date(last_day_month, format_date)
        first_day_month = string_format_br_date(first_day_month, format_date)

    return first_day_month, last_day_month


def get_range_days_from_current_month(
        end_date_limit: int = 0,
        format_date: str = ''
        ) -> tuple[str | datetime, str | datetime]:
    current_date = get_date_back(end_date_limit)
    month = ('0' + str(current_date.month)
             if len(str(current_date.month)) < 2 else current_date.month)
    first_date = datetime.strptime(
        f'01/{month}/{current_date.year}',
        '%d/%m/%Y'
        )
    if len(str(format_date)) > 0:
        current_date = string_format_br_date(current_date, format_date)
        first_date = string_format_br_date(first_date, format_date)
    return first_date, current_date


def convert_string_date_to_datetime(
        date_string: str,
        date_format: str = "%d/%m/%Y",
        date_option: int = 0
        ) -> None | datetime:
    if (date_string is None
            or len(str(date_string).strip()) == 0
            or date_option > 30):
        return None
    # Convert the date string to a datetime object
    if date_format is not None:
        return datetime.strptime(date_string, date_format)
    else:
        try:
            format = get_str_format_datetime_by_switch_case(date_option + 1)
            return datetime.strptime(date_string, format)
        except Exception:
            return convert_string_date_to_datetime(
                date_string,
                None,
                date_option=date_option + 1)


def string_format_br_date(
        date: datetime,
        format_date: str = '6'
        ) -> str | datetime:
    """
    This function takes a datetime object and a format code as inputs and
    returns the date formatted in various Brazilian formats.

    :param date: The input date as a datetime object
    :type date: datetime

    :param format_date: Format code to specify the desired date format.
                        Available options are:
                            - '6' : ddmmyy
                            - '8a': ddmmyyyy
                            - '8b': dd/mm/yy
                            - '8c': dd-mm-yy
                            - '10a': dd/mm/yyyy
                            - '10b': dd-mm-yyyy
    :type format_date: str

    :return: The date formatted in the specified Brazilian date format
    :rtype: str

    :process:
        1. Extract the day and month as two-digit strings.
        2. Extract the year in two formats: 'yy' and 'yyyy'.
        3. Format the date string based on the format code provided.

    Example:
        >>> from datetime import datetime
        >>> string_format_br_date(datetime(2023, 1, 1), '8a')
        '01012023'
    """
    if date is None:
        return date
    # adjusting day and month length
    day = '0' + str(date.day) if len(str(date.day)) < 2 else str(date.day)
    month = ('0' + str(date.month)
             if len(str(date.month)) < 2 else str(date.month))

    # defining the 'yy' and 'yyyy' year formats
    year_2 = str(date.year)[-2:]
    year_4 = str(date.year)

    # formating according the desired br format
    # 010123
    if '6' in str(format_date):
        return f'{day}{month}{year_2}'
    # 01012023
    elif '8a' in str(format_date):
        return f'{day}{month}{year_4}'
    # 01/01/23
    elif '8b' in str(format_date):
        return f'{day}/{month}/{year_2}'
    # 01-01-23
    elif '8c' in str(format_date):
        return f'{day}-{month}-{year_2}'
    # 01/01/2023
    elif '10a' in str(format_date):
        return f'{day}/{month}/{year_4}'
    # 01-01-2023
    elif '10b' in str(format_date):
        return f'{day}-{month}-{year_4}'
    elif format_date is None:
        return date
    else:
        return date.strftime(format_date)


def add_left_zero(value: str | int, len_desired: int) -> str:
    zeros = ''.join(['0' for _ in range(len_desired - len(str(value)))])
    return zeros + str(value)


def get_day_greetings(message: str | None = None):
    current_hour = datetime.now().hour
    message = 'Como sr(a) está?' if message is None else message
    if current_hour < 12:
        greeting = f'Bom dia! {message}'
    elif 12 <= current_hour < 18:
        greeting = f'Boa tarde! {message}'
    else:
        greeting = f'Boa noite! {message}'
    return greeting


def match_numeric(value: str) -> str:
    return ''.join(re.findall('\d', str(value)))


def match_letter(value: str) -> str:
    return ''.join(re.findall(r'[a-zA-Z]+', value))


def change_letters_per_number(text: str) -> int:
    import string
    # Get the letters of alphabet
    alphabet = str(string.ascii_lowercase)

    # Convert the text to lower case to match with the lower letters in the
    # alphabet
    text = text.lower()

    # Replace letter by the position number of this letter in the alphabet and
    # keep char number
    number = int(
        ''.join(str(alphabet.find(c) + 1)
                if len(re.findall('\d', c)) == 0 else c for c in text)
                )

    return number


def convert_img_to_pdf(path_img: str | os.PathLike) -> None:
    from PIL import Image
    base_path = path_img.split('.')[0]
    pdf_path = f'{base_path}.pdf'

    # Open an image file
    with Image.open(path_img) as image:
        # Convert the image to RGB, even if it's already in RGB mode,
        # to ensure the image is saved as a color image
        imagelist = [image.convert('RGB')]
        image.save(pdf_path, save_all=True, append_images=imagelist)


def partial_comparation(
        list_1: list,
        list_2: list,
        confidence: int = 70
        ) -> list:
    result = []
    for word1 in list_1:
        found = False
        for word2 in list_2:
            if is_similar(word1, word2, confidence):
                found = True
                break
        result.append(found)
    return result


def is_similar(str1: str, str2: str, confidence: int = 70) -> bool:
    from fuzzywuzzy import fuzz
    return fuzz.ratio(str(str1), str(str2)) >= confidence


def get_similiar_string_in_list(
        word: str,
        list_to_search: list,
        confidence: int = 70
        ) -> str | None:
    for word2 in list_to_search:
        if is_similar(word, word2, confidence):
            return word2
    return None


def find_wrong_rows_csv_file(file_path: str | os.PathLike, sep: str = ';'):
    with open(file_path, 'r') as f:
        read_ = f.read()
        rows = read_.split('\n')
        len_columns = [len(row.split(sep)) for row in rows]
        filt_len_columns = [len_ for len_ in len_columns if len_ > 1]
        val_max = max(filt_len_columns)
        val_min = min(filt_len_columns)
        wrong_rows = [i for i, row in enumerate(rows) if len(row.split(sep)) != val_min]
        logging.info(f'{val_max}, {val_min}, {wrong_rows}')
        return wrong_rows


def color_table_line(
        file_path: str | os.PathLike,
        color_argb_background: str = 'FFFFFF00',
        color_argb_values: str | None = None,
        index_line: int | list = 1
        ) -> None:
    """
        Colorizes a specific line in a table in an Excel file.

        This function applies a background color and/or a font color to a
        specific line in a table in an Excel file.
        The file is modified in-place.

        :param file_path: The path to the Excel file.
        :type file_path: str
        :param color_argb_background: The ARGB value of the background color,
        defaults to 'FFFFFF00'.
        :type color_argb_background: str, optional
        :param color_argb_values: The ARGB value of the font color,
        defaults to None.
        :type color_argb_values: str, optional
        :param index_line: The index of the line to colorize, defaults to 1.
                            If multiple indices are provided, each line will
                            be colorized.
        :type index_line: int or list[int], optional

        :process:
            1. Load the existing workbook from the file.
            2. Select the active worksheet.
            3. If index_line is an integer, convert it to a list containing a
            single element.
            4. Iterate over each line in index_line.
            5. Select the row corresponding to the line.
            6. If color_argb_values is not None, create a font with the
            specified color.
                Apply this font to all cells in the row.
            7. If color_argb_background is not None, create a fill pattern
            with the specified background color.
                Apply this fill pattern to all cells in the row.
            8. Save the modified workbook back to the file.

        Example:
            >>> file_path = "table.xlsx"
            >>> color_table_line(file_path, color_argb_background='FFFFFF00',
            color_argb_values=None, index_line=1)
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Color, PatternFill

    # Load an existing workbook
    wb = load_workbook(file_path)
    ws = wb.active

    if isinstance(index_line, int):
        index_line = [index_line]

    for line in index_line:
        # Select the first row
        header_row = ws[line]

        # color de values
        if color_argb_values is not None:
            font = Font(color=Color(color_argb_values))

            # Apply this font to all cells in the header row
            for cell in header_row:
                cell.font = font

        # color de background
        if color_argb_background is not None:
            yellow_fill = PatternFill(
                start_color=color_argb_background,
                end_color=color_argb_background,
                fill_type="solid"
                )

            # Apply this font to all cells in the header row
            for cell in header_row:
                cell.fill = yellow_fill

    # Save the workbook
    wb.save(file_path)


def get_list_start_end_dates_by_range_days_on_year(
        start_date: str = '',
        end_date: str = '',
        step: int = 15,
        year: int = 2023,
        format_date='6'
        ) -> tuple[list[str | datetime], list[str | datetime]]:
    ini_ = f'01/01/{year}' if len(start_date) == 0 else start_date
    ini_datetime = convert_string_date_to_datetime(ini_)
    end_ = (get_date_back(0) if len(end_date) == 0
            else convert_string_date_to_datetime(end_date))
    if ini_datetime is not None and end_ is not None:
        diff = (end_ - ini_datetime).days
        range_ = int(diff / step)
        list_start_dates = [ini_datetime]
        list_end_dates = []
        next_date_ = ini_datetime
        for i in range(range_):
            last_date = next_date_ + timedelta(1)
            next_date_ = next_date_ + timedelta(step)
            if i == range_ - 1:
                next_date_ = end_
            list_end_dates.append(next_date_)
            if i > 0:
                list_start_dates.append(last_date)

        for i in range(range_):
            list_start_dates[i] = string_format_br_date(list_start_dates[i], format_date=format_date)
            list_end_dates[i] = string_format_br_date(list_end_dates[i], format_date=format_date)
        return list_start_dates, list_end_dates
    return list(), list()


def append_text_to_file(
        text: str,
        name_file: str,
        path_to_save: str | os.PathLike,
        create_if_not_exist: bool = False
        ) -> str:
    os.makedirs(path_to_save, exist_ok=True)
    full_path_file = os.path.abspath(f'{path_to_save}/{name_file}.txt')

    if os.path.exists(full_path_file) or create_if_not_exist:
        with open(full_path_file, 'a') as f:
            f.write(text)
    return text


def get_day_of_week(isoweekday: int) -> str | int:
    switcher = {
        1: "SEGUNDA",
        2: "TERCA",
        3: "QUARTA",
        4: "QUINTA",
        5: "SEXTA",
        6: "SABADO",
        7: "DOMINGO",
    }

    return switcher.get(isoweekday, isoweekday)


def get_month_name(number_month: int) -> str | int:
    switcher = {
        1: "JANEIRO",
        2: "FEVEREIRO",
        3: "MARÇO",
        4: "ABRIL",
        5: "MAIO",
        6: "JUNHO",
        7: "JULHO",
        8: "AGOSTO",
        9: "SETEMBRO",
        10: "OUTUBRO",
        11: "NOVEMBRO",
        12: "DEZEMBRO"
    }

    return switcher.get(number_month, number_month)


def thread_it(
        qtd_threads: int,
        list_params: list,
        function: Callable
        ) -> None:
    from tqdm import tqdm
    import threading as td
    threads = [td.Thread() for _ in range(qtd_threads)]
    counter = 0
    for param in tqdm(list_params, desc="Threading", unit="item"):
        threads[counter] = td.Thread(target=lambda: function(param))
        threads[counter].start()
        counter += 1
        if counter >= qtd_threads:
            [threads[i].join() for i in range(counter)]
            counter = 0
    if counter < qtd_threads:
        [threads[i].join() for i in range(counter)]


def scan_image(
        image_path: str | os.PathLike,
        output_path: str | os.PathLike
        ) -> None:
    import cv2
    import numpy as np
    # Read the image
    image = cv2.imread(image_path, cv2.IMREAD_COLOR)

    # Convert the image to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    height, width = gray.shape
    gray = cv2.resize(
        gray,
        (width * 4, height * 4),
        interpolation=cv2.INTER_LINEAR
        )

    processed = cv2.equalizeHist(gray)

    processed = cv2.GaussianBlur(processed, (3, 3), 0)

    # Dilation can be used to strengthen the edges
    kernel = np.ones((3, 3), np.uint8)
    dilated = cv2.dilate(processed, kernel, iterations=1)

    # Combine the dilated edges with the original image
    processed = cv2.addWeighted(processed, 0.7, dilated, 0.8, 0.2)

    # Save the preprocessed image
    cv2.imwrite(output_path, processed)


def remove_background(
        path_from: str | os.PathLike,
        path_to: str | os.PathLike
        ) -> None:
    from rembg import remove
    from PIL import Image
    # Processing the image
    input_img = Image.open(path_from)

    # Removing the background from the given Image
    output = remove(input_img)

    # Saving the image in the given path
    output.save(path_to)


def is_iterable(obj):
    if isinstance(obj, str) or isinstance(obj, dict):
        return False  # Treat strings as non-iterable
    try:
        iter(obj)
        return True
    except TypeError:
        return False


def _append_results(report: dict, result: list) -> None:
    if is_iterable(result):
        for k, r in enumerate(result):
            try:
                report[f'result_{k + 1}'] += [r]
            except Exception as e:
                logging.error(e)
                report[f'result_{k + 1}'] = [r]
    else:
        try:
            report['result'] += [result]
        except Exception as e:
            logging.error(e)
            report['result'] = [result]


def get_report_from_interaction(
        func: Callable,
        list_interaction: list
        ) -> pd.DataFrame:
    report = {'interactor': list_interaction}
    for param in list_interaction:
        try:
            result = func(param)
            _append_results(report, result)
        except Exception as e:
            logging.error(e)
            _append_results(
                report,
                [f'ERROR: {str(e)}' for _ in list(report.keys())[1:]])
    # logging.info(json.dumps(report,indent=6))
    logging.info(report)
    report = pd.DataFrame(report)
    return report


def map_network_drive(unc_path: str | os.PathLike):
    available_drive = next((d for d in string.ascii_uppercase
                            if not os.path.exists(f'{d}:')), None)
    if not available_drive:
        raise Exception("No available drive letter to map network drive.")

    drive_letter = f"{available_drive}:"
    logging.info(drive_letter)
    subprocess.run(['net', 'use', drive_letter, unc_path])
    return drive_letter


def unmap_network_drive(drive_letter) -> None:
    subprocess.run(['net', 'use', drive_letter, '/delete', '/yes'])


def set_environ_variables(variable_name, variable) -> None:
    # Set an environment variable permanently for the current user
    subprocess.run(['setx', variable_name, variable], check=True)


def rename_duplicated_columns(columns):
    key_col = {}
    renamed_columns = []
    for col in columns:
        key_col[col] = []
    for col in columns:
        key_col[col].append(1)
    for key in key_col:
        for i, _ in enumerate(key_col[key]):
            name = key if i == 0 else f'{key}.{i}'
            renamed_columns.append(name)
    return renamed_columns


def fill_empty_values_with_last_non_empty(
        column: str,
        table: pd.DataFrame
        ) -> pd.DataFrame:
    list_index_non_empty = [
        i for i, val in enumerate(table[column]) if len(val) > 0
        ]
    for i, index in enumerate(list_index_non_empty):
        next_pos = i + 1
        if next_pos >= len(list_index_non_empty):
            next_index = len(table) - 1
        else:
            next_index = list_index_non_empty[next_pos] - 1
        table.loc[index:next_index, column] = table.loc[index, column]
    return table


def read_text_from_pdf(path_to_pdf: str | os.PathLike) -> str:
    # pip install PyMuPDF
    import fitz
    doc = fitz.open(path_to_pdf)
    pdf_text = " ".join([page.get_text() for page in doc])
    return pdf_text


def find_files(directory: str | os.PathLike, partial_name: str) -> list:
    matches = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if partial_name in file:
                matches.append(os.path.join(root, file))
    return matches


def get_list_of_weekday_in_current_month(weekday: int = 1) -> list:
    import calendar
    from datetime import datetime
    # Get current year and month
    now = datetime.now()
    year = now.year
    month = now.month

    # Get the number of days in the current month
    _, num_days = calendar.monthrange(year, month)

    # Find all Mondays in the current month
    mondays = []
    for day in range(1, num_days + 1):
        date = datetime(year, month, day)
        if date.weekday() == weekday:  # Monday is represented by 0
            mondays.append(day)

    return mondays


def generate_id():
    import hashlib
    import uuid

    hash_value = hashlib.sha1(uuid.uuid4().bytes).hexdigest()[:10]
    timestamp = str(int(time.time()))
    return f"{timestamp}_{hash_value}"


def generate_unique_filename(original_filename: str) -> str:
    # Extrai a extensão do arquivo original
    ext = os.path.splitext(original_filename)[1]

    # Concatena os componentes para formar o novo nome
    new_filename = f"{generate_id()}{ext}"
    return new_filename


def shave_marks(text):
    import unicodedata
    norm_text = unicodedata.normalize('NFD', text)
    shaved = ''.join(c for c in norm_text if not unicodedata.combining(c))
    return unicodedata.normalize('NFC', shaved)


class ReadSSWTextReport:
    def __init__(self, file_path: str | os.PathLike, delimiter='+'):
        self.file_path = file_path
        self.delimiter = delimiter
        self.df: pd.DataFrame = pd.DataFrame()

    def find_dash(self):
        with open(self.file_path, 'r') as f:
            dashs = [
                row for i, row in enumerate(f.readlines())
                if self.delimiter in row
                ]
        if len(dashs) > 0:
            return dashs[0]

    def get_len_char_columns(self, dash):
        columns_length = [len(col) + 1 for col in dash.split(self.delimiter)]
        return columns_length

    def get_dataframe(self, encoding='latin-1', **kwargs) -> pd.DataFrame:
        dash = self.find_dash()
        if dash is not None:
            columns_length = self.get_len_char_columns(dash)
            self.df = pd.read_fwf(
                self.file_path, 
                widths=columns_length, 
                encoding=encoding, 
                **kwargs
                )
            logging.info(self.df)
            return self.df
        else:
            raise ValueError(f'No "{self.delimiter}" delimiter found ')


if __name__ == '__main__':
    ...