from typing import List, Any, Optional
from openpyxl.workbook import Workbook
from openpyxl.utils import get_column_letter

from ..base_generator import BaseGenerator, COLUMNS_TYPE


class BaseXLSXReportsGenerator:
    """Base class for generating XLSX reports."""

    def __init__(self, title: str, base_generator: BaseGenerator):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = title
        self.base_generator = base_generator

    def assign_cell_titles(
        self, columns: List[COLUMNS_TYPE], row_num: int = 1
    ):
        for col_num, (column_title, column_width) in enumerate(columns, 1):
            cell = self.worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            column_letter = get_column_letter(col_num)
            column_dimensions = self.worksheet.column_dimensions[column_letter]
            column_dimensions.width = column_width

    def assign_cell_data(self, row: list, row_num: int):
        for col_num, cell_value in enumerate(row, 1):
            cell = self.worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    def generate(
            self, items: List[Any], header_columns: Optional[COLUMNS_TYPE] = None,
            save_file_path: Optional[str] = None
    ) -> Workbook:
        row_num: int = 1

        # Add optional header columns first
        if header_columns is not None:
            self.assign_cell_titles(columns=header_columns)
            row_num += 2  # Leave a blank row after headers

        # Add main columns
        columns: List[COLUMNS_TYPE] = self.base_generator.get_columns()
        self.assign_cell_titles(columns=columns, row_num=row_num)
        row_num += 1

        # Add data rows
        for item in items:
            self.assign_cell_data(
                row=self.base_generator.generate_row_fn(item), row_num=row_num
            )
            row_num += 1

        if save_file_path is not None:
            self.workbook.save(save_file_path)

        return self.workbook
