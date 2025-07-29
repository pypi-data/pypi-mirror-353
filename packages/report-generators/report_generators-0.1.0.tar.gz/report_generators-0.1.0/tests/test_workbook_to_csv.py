"""Tests for workbook to CSV conversion utility."""
import csv
from io import StringIO
from openpyxl import Workbook
from src.report_generators.utils.workbook_to_csv import workbook_to_csv_buffer


def test_workbook_to_csv_conversion():
    """Test converting a workbook to CSV."""
    # Create a test workbook
    wb = Workbook()
    ws = wb.active
    
    # Add some test data
    ws['A1'] = 'Name'
    ws['B1'] = 'Age'
    ws['A2'] = 'John'
    ws['B2'] = 30
    
    # Convert to CSV
    csv_buffer = workbook_to_csv_buffer(wb)
    
    # Check the conversion
    assert isinstance(csv_buffer, StringIO)
    
    csv_buffer.seek(0)
    reader = csv.reader(csv_buffer)
    rows = list(reader)
    
    assert rows[0] == ['Name', 'Age']
    assert rows[1] == ['John', '30']
