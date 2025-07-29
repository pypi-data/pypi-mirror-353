import base64
import hashlib
import re
import subprocess
import traceback
import pyperclip
import psutil
from datetime import datetime
from subprocess import PIPE, Popen
import smtplib
from dateutil import parser
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.header import Header
import requests
import pandas as pd
import xlrd
from shutil import copyfile
import openpyxl
import csv
import yaml
import os
import shutil
import winreg
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.common.exceptions import SessionNotCreatedException
from webdriver_manager.chrome import ChromeDriverManager
from webdriver_manager.microsoft import EdgeChromiumDriverManager
r"""
出现问题时尝试：pip install --upgrade twine urllib3     pip install urllib3==1.26.5
cd pip_Mzhtools & del /q dist && del /q Mzhtools.egg-info && python setup.py bdist_wheel && twine upload dist/*
"""


def read_txt(filename):
    '''
    读取txt文档
    :param filename: txt文件名
    :return: 文件内容
    '''
    try:
        with open(filename, "r", encoding='utf8') as f:
            data = f.read()
        return data
    except:
        with open(filename, "r", encoding='gbk') as f:
            data = f.read()
        return data


def write_to_txt(content, filename, mode='a'):
    '''
    内容写入txt
    :param content: 写入内容，可字符串可列表。列表多行写入
    :param filename: 保存文件名
    :param mode: 写入模式，默认“a”
    :return: 空
    '''
    if type(content) == str:
        with open(filename, mode) as f:
            f.write(content)
    else:
        with open(filename, mode) as f:
            f.write('\n'.join(content))


def writelist_to_csv(ls, filename, mode='a'):
    '''
    将列表写入csv文件
    :param ls: 数据列表
    :param filename: 写入文件名.csv
    :param mode: 写入方式，默认"a"
    :return: 空
    '''
    if type(ls[0]) == str:
        with open(filename, mode, newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(ls)
    else:
        with open(filename, mode, newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerows(ls)


def read_csv(filename):
    '''
    读取csv文件
    :param filename: 文件名
    :return: 数据列表
    '''
    with open(filename, "r") as csvfile:
        data_list = [l for l in csv.reader(csvfile)]
        return data_list


def dabao(path):
    """
    打包路径：c:\pybuild
    :param path: 全路径
    :return: 无
    """
    os.makedirs(r'c:\pybuild', exist_ok=True)
    try:
        cmd_gbk(
            rf'rd /S /Q "c:\pybuild"&&mkdir "c:\pybuild"&&pyinstaller {path} --workpath c:\pybuild  --distpath c:\pybuild\dist')
        return
    except:
        pass
    cmd_utf(
        rf'rd /S /Q "c:\pybuild"&&mkdir "c:\pybuild"&&pyinstaller {path} --workpath c:\pybuild  --distpath c:\pybuild\dist')


def qywx_sendtxt(text, key, all=None, personal=None):
    """
    向指定企业微信群发送文本信息
    :param text: 发送的文本内容
    :param key: 群机器人key
    :param all: 默认None，不@全体人员。all的值为真时，@全体人员
    :return: 空
    """
    if personal:
        headers = {"Content-Type": "text/plain"}
        send_url = rf'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={key}'
        send_data = {
            "msgtype": "text",  # 消息类型
            "text": {
                "content": text,  # 文本内容，最长不超过2048个字节，必须是utf8编码
                "mentioned_list": ["@all"]
            }
        }
        requests.post(url=send_url, headers=headers, json=send_data)
    elif all:
        headers = {"Content-Type": "text/plain"}
        send_url = rf'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={key}'
        send_data = {
            "msgtype": "text",  # 消息类型
            "text": {
                "content": text,  # 文本内容，最长不超过2048个字节，必须是utf8编码
                "mentioned_list": ["@all"]
            }
        }
        requests.post(url=send_url, headers=headers, json=send_data)
    else:
        headers = {"Content-Type": "text/plain"}
        send_url = rf'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={key}'
        send_data = {
            "msgtype": "text",  # 消息类型
            "text": {
                "content": text,  # 文本内容，最长不超过2048个字节，必须是utf8编码
            }
        }
        requests.post(url=send_url, headers=headers, json=send_data)


def qywx_sendImg(imgPath, key):
    """
    向指定企业微信群发送图片
    :param imgPath: 图片路径
    :param key: 群机器人key
    :return: 空
    """
    url = rf'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={key}'
    with open(imgPath, "rb") as f:
        fd = f.read()
        base64Content = str(base64.b64encode(fd), "utf-8")
    with open(imgPath, "rb") as f:
        fd = f.read()
        md = hashlib.md5()
        md.update(fd)
        md5Content = md.hexdigest()
    headers = {"content-type": "application/json"}
    msg = {"msgtype": "image", "image": {"base64": base64Content, "md5": md5Content}}
    requests.post(url, headers=headers, json=msg)


def qywx_sendfile(file, key):
    """
    向指定企业微信群发送文件
    :param file: 文件路径
    :param key: 群机器人key
    :return: 空
    """
    # 获取media_id
    id_url = f'https://qyapi.weixin.qq.com/cgi-bin/webhook/upload_media?key={key}&type=file'
    files = {'file': open(file, 'rb')}
    res = requests.post(url=id_url, files=files)
    media_id = res.json()['media_id']
    # 发送文件
    webhook = rf'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={key}'
    header = {
        "Content-Type": "application/json",
        "Charset": "UTF-8"
    }
    data = {
        "msgtype": "file",
        "file": {
            "media_id": media_id
        }
    }
    requests.post(url=webhook, json=data, headers=header)


def qywx_sendmarkdown(title, name, data, key):
    """
    向指定企业微信群发送markdown
    :param title: markdown标题
    :param name: markdown的key值
    :param data: markdown的value值
    :param key: 群机器人key
    :return: 空
    """
    headers = {"Content-Type": "text/plain"}
    send_url = rf'https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key={key}'
    content = ''
    if title:
        content = title + '\n'
    for i in zip(name, data):
        content += f"> {i[0]}：<font color=\"info\">{i[1]}</font> \n"
    send_data = {
        "msgtype": "markdown",  # 消息类型，此时固定为markdown
        "markdown": {
            "content": content
        }
    }
    requests.post(url=send_url, headers=headers, json=send_data)


def delete_file(target_dir, days):
    """
    删除指定文件夹里超过指定天数的所有东西
    :param target_dir: 文件夹路径
    :param days: 超过的天数
    :return: 空
    """
    now_time = datetime.now().strftime('%Y-%m-%d')
    for (dirpath, dirnames, filenames) in os.walk(target_dir):
        for i in filenames:
            modify_time = datetime.fromtimestamp(os.path.getmtime(os.path.join(dirpath, i))).strftime('%Y-%m-%d')
            Days = int((parser.parse(now_time) - parser.parse(modify_time)).days)
            if Days > days:
                os.remove(os.path.join(dirpath, i))
        for dirname in dirnames:
            if not os.listdir(os.path.join(dirpath, dirname)):
                shutil.rmtree(os.path.join(dirpath, dirname), ignore_errors=True)


def cmd_utf(order):
    """
    utf-8编码
    :param order: 输入对应命令
    :return: outinfo, errinfo
    """
    proc = Popen(order, stdin=None, stdout=PIPE, stderr=PIPE, shell=True)
    outinfo, errinfo = proc.communicate()
    outinfo = outinfo.decode('utf-8')
    errinfo = errinfo.decode('utf-8')
    return outinfo, errinfo


def cmd_gbk(order):
    """
    gbk编码
    :param order: 输入对应命令
    :return: outinfo, errinfo
    """
    proc = Popen(order, stdin=None, stdout=PIPE, stderr=PIPE, shell=True)
    outinfo, errinfo = proc.communicate()
    outinfo = outinfo.decode('gbk')
    errinfo = errinfo.decode('gbk')
    return outinfo, errinfo


def py_run(path):
    """
    传入py文件路径，在本代码调用其他代码
    :param path:
    :return: 空
    """
    if os.path.dirname(path) == '':
        path = fr'.\{path}'
    try:
        print(cmd_gbk(rf"cd /d {os.path.dirname(path)} && python {os.path.basename(path)}")[0])
        return
    except:
        pass
    print(cmd_utf(rf"cd /d {os.path.dirname(path)} && python {os.path.basename(path)}")[0])


def pushplus(title, content, token="cf8736250267472e954737c221b33d23"):
    """
    推送到微信推送加
    :param title:标题
    :param content: 内容
    :param token: 没有默认自己的微信
    :return: 空
    """
    url = fr"https://www.pushplus.plus/send?token={token}&title={title}&content={content}&template=html"
    payload = {}
    files = {}
    headers = {
        'User-Agent': 'Apifox/1.0.0 (https://apifox.com)'
    }
    requests.request("GET", url, headers=headers, data=payload, files=files)


def send_email(name_list, fpath='', title='', text=''):
    """
    :param name_list: ['349****42@qq.com']，可以添加多个收件人
    :param fpath: 传入字符串，以空格连接，邮箱附件路径，可以添加多个，默认空
    :param title: 邮件标题，没有此参数时，默认空
    :param text: 邮件正文，没有此参数时，默认空
    :return: 空
    """
    sender_qqemail = '191891173@qq.com'  # 发件人邮箱
    secret = 'tvthacuvztgicaib'  # 发件人邮箱密钥
    # 生成连结对象,参数分别是邮件服务器和端口号
    con = smtplib.SMTP_SSL('smtp.qq.com', 465)
    # 使用用户名和密码登录,这里密码以星号隐藏了
    con.login(sender_qqemail, secret)
    # 生成一个邮件对象，由于邮件包含文本、图片、HTML、附件等内容，
    # 所以这里用MIMEMultipart()生成邮件对象，以支持多种数据格式
    mail_obj = MIMEMultipart()
    # 生成邮件表头的内容
    mail_header = Header(title, 'utf-8').encode()
    # 主题
    mail_obj['Subject'] = mail_header
    # 发送者邮箱
    mail_obj['From'] = f'{sender_qqemail} <{sender_qqemail}>'
    # 接收者邮箱
    mail_obj['To'] = '我'
    # 添加邮件正文
    mail_text = MIMEText(text, 'plain', 'utf-8')
    mail_obj.attach(mail_text)
    for path in fpath.strip().split(' '):
        if path.endswith('.txt'):
            # 添加txt附件
            with open(path, 'rb') as f:
                txt = f.read()
                txt = MIMEText(txt, 'base64', 'utf-8')
                txt["Content-Type"] = 'application/octet-stream'
                txt["Content-Disposition"] = 'attachment; filename="I.txt"'
                mail_obj.attach(txt)
        if path.endswith('.xlsx'):
            # 添加Excel附件
            with open(path, 'rb') as f:
                Excel = f.read()
                Excel = MIMEText(Excel, 'base64', 'utf-8')
                Excel["Content-Type"] = 'application/octet-stream'
                Excel["Content-Disposition"] = 'attachment; filename="ove.xlsx"'
                mail_obj.attach(Excel)
        if path.endswith('.zip'):
            # 添加Zip附件
            with open(path, 'rb') as f:
                Zip = f.read()
                Zip = MIMEText(Zip, 'base64', 'utf-8')
                Zip["Content-Type"] = 'application/octet-stream'
                Zip["Content-Disposition"] = 'attachment; filename="class.rar"'
                mail_obj.attach(Zip)
        if path.endswith('.png'):
            # 添加图片附件
            with open(path, 'rb') as f:
                img2 = f.read()
                img_2 = MIMEImage(img2)
                # 指定图片类型与文件名，以下语句设置图片文件以附件形式加到邮件中
                img_2['Content-Disposition'] = 'attachment;filename="flower.png"'
                # 加入到邮件中
                mail_obj.attach(img_2)
        if path.endswith('.docx'):
            # 添加word附件
            with open(path, 'rb') as f:
                doc = f.read()
                # 以数据流的形式读入文件
                doc = MIMEText(doc, 'base64', 'utf-8')
                # 以下语句设置文件以附件形式加到邮件中
                doc['Content-Disposition'] = 'attachment;filename="test.docx"'
                # 加入到邮件中
                mail_obj.attach(doc)

    # 发送邮件
    con.sendmail(sender_qqemail, name_list, mail_obj.as_string())
    # 断开连结
    con.quit()
    print('发送邮件成功...')


def readrow_excel(path, sheet_name=None, index=None):
    '''
    xlrd横向读取表格
    :param path: 文件路径
    :param sheet_name: 表单名，默认None
    :param index: 索引，默认None
    :return: 横向表格数据
    '''
    data_list = []
    if sheet_name:
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_name(sheet_name)
        data_list = [sheet.row_values(rowx=i) for i in range(sheet.nrows)]
        book.release_resources()
    if index != None:
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(index)
        data_list = [sheet.row_values(rowx=i) for i in range(sheet.nrows)]
        book.release_resources()
    return data_list


def readcol_excel(path, sheet_name=None, index=None):
    '''
    xlrd纵向读取表格
    :param path: 文件路径
    :param sheet_name: 表单名，默认None
    :param index: 索引，默认None
    :return: 纵向表格数据
    '''
    data_list = []
    if sheet_name:
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_name(sheet_name)
        data_list = [sheet.col_values(colx=i) for i in range(sheet.ncols)]
        book.release_resources()
    if index != None:
        book = xlrd.open_workbook(path)
        sheet = book.sheet_by_index(index)
        data_list = [sheet.col_values(colx=i) for i in range(sheet.ncols)]
        book.release_resources()
    return data_list


def writelist_toExcel(ls, savename, loadname=None, sheet_name=None):
    """
    按行写入表格
    :param ls: 数据列表
    :param savename: 保存的文件名
    :param loadname: 是否写入已存在表格，默认为None，即写入新表格
    :param sheet_name: 表单名，内容写进哪个表单
    :return:
    """
    if loadname == None:
        book = openpyxl.Workbook()
        sh = book.active
        for row in ls:
            sh.append(row)
        book.save(savename)
        book.close()
    else:
        wb = openpyxl.load_workbook(loadname)
        sheet = wb[sheet_name]
        for row in ls:
            sheet.append(row)
        wb.save(savename)
        wb.close()


def pd_read_excel(path, sheet_name):
    datalist = pd.read_excel(path, sheet_name, header=None, index_col=None)
    LS = []
    for j in range(len(datalist)):
        ls = []
        for i in datalist.columns:
            ls.append(datalist.iloc[j][i])
        LS.append(ls)
    return LS


def pd_writelist_toExcel(ls, savename, sheet_name):
    if not os.path.exists(savename):
        book = openpyxl.Workbook()
        sh = book.active
        sh.title = sheet_name
        book.save(savename)
    df = pd.DataFrame(ls)
    with pd.ExcelWriter(savename, datetime_format="YYYY-MM-DD") as writer:
        df.to_excel(writer, sheet_name=sheet_name, header=False, index=False)


def pd_writelist_appendExcel(ls, savename, sheet_name):
    data = pd_read_excel(savename, sheet_name)
    ls = data + ls
    df = pd.DataFrame(ls)
    with pd.ExcelWriter(savename, datetime_format="YYYY-MM-DD") as writer:
        df.to_excel(writer, sheet_name=sheet_name, header=False, index=False)


def CopyFiles_to_Clipboard(file_ls):
    """
    复制文件到剪切板
    :param file_ls:全路径的列表
    :return:拷贝失败的文件列表
    """
    pyperclip.copy('')
    pan = psutil.disk_partitions()[-1][0]
    os.makedirs(f'{pan}临时', exist_ok=True)
    for (dirpath, dirnames, filenames) in os.walk(f'{pan}临时'):
        for fn in filenames:
            os.remove(os.path.join(dirpath, fn))
    files = []
    fail = []
    for fpath in file_ls:
        copyfile(fpath, f'{pan}临时/{os.path.basename(fpath)}')
    for f in os.listdir(f'{pan}临时'):
        fpath = f'{pan}临时/{f}'
        try:
            for jj in [' ', '(', ')', '“', '”']:
                if jj in fpath:
                    os.rename(fpath, fpath.replace(jj, ''))
                    fpath = fpath.replace(jj, '')
            if os.path.getsize(fpath) < 100 * 1024 * 1024:
                files.append(fpath)
            else:
                fail.append(fpath)
        except:
            fail.append(fpath)
    args = ['powershell', fr'Get-Item {",".join(files)}| Set-Clipboard']
    proc = subprocess.Popen(args=args, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    outinfo, errinfo = proc.communicate()
    errinfo = errinfo.decode('gbk')
    win32clipboard.OpenClipboard()
    filenames = win32clipboard.GetClipboardData(win32clipboard.CF_HDROP)
    win32clipboard.CloseClipboard()
    if errinfo != "" or len(filenames) != len(file_ls):
        fail += files
    print("发送失败：", fail)
    print("报错信息：", errinfo)
    return list(set(fail))


def dirfiles_numlimit(dirfullname, limit_num):
    """
    限制文件夹文件数量
    :param dirfullname: 文件夹全路径
    :param limit_num: 文件夹限制文件数量
    :return:空
    """
    for (dirpath, dirnames, filenames) in os.walk(dirfullname):
        count = 0
        for fn in filenames:
            fpath = os.path.join(dirpath, fn)
            new_dir = fr'{dirpath}/{os.path.basename(dirpath)}_{int(count / limit_num)}'
            os.makedirs(new_dir, exist_ok=True)
            copyfile(fpath, fr'{new_dir}/{fn}')
            os.remove(fpath)
            count += 1


def read_yaml_file(file_path):
    '''
    返回 YAML 文件数据
    :param file_path:
    :return:
    '''
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            data = yaml.safe_load(file)
            return data
    except FileNotFoundError:
        print(f"错误：文件 {file_path} 未找到。")
    except yaml.YAMLError as e:
        print(f"错误：解析YAML文件时出现错误 - {e}")
    return None


def selenium_ensure_chromedriver():
    """
    确保谷歌浏览器驱动存在且版本匹配，不创建 WebDriver 实例。
    返回驱动全路径
    """
    new_path = 'c:/selenium_driver/chrome_driver/'
    driver_file = os.path.join(new_path, 'chromedriver.exe')
    if not os.path.exists(new_path):
        os.makedirs(new_path)
    # 如果驱动不存在，直接下载
    if not os.path.exists(driver_file):
        print("驱动不存在，正在下载...")
        driver_path = ChromeDriverManager().install()
        print("驱动下载地址：", driver_path)
        shutil.copy(driver_path, new_path)
    else:
        print("使用现有驱动文件...")

    # 尝试验证驱动是否可用（不创建完整会话）
    try:
        service = ChromeService(executable_path=driver_file)
        # 实际不会打开浏览器，但会触发驱动检查
        webdriver.Chrome(service=service).quit()
        print("驱动验证通过。")
    except SessionNotCreatedException as e:
        print("检测到驱动版本不匹配，正在重新下载最新驱动...")
        driver_path = ChromeDriverManager().install()
        print("驱动下载地址：", driver_path)
        shutil.copy(driver_path, new_path)
        print("驱动已更新。")
    return driver_file


def selenium_ensure_edgedriver():
    """
    确保 msedgedriver 存在且版本匹配，不创建完整的 WebDriver 实例。
    返回驱动文件的完整路径。
    """
    driver_dir = 'c:/selenium_driver/edge_driver/'
    driver_file = os.path.join(driver_dir, 'msedgedriver.exe')

    # 创建目录（如果不存在）
    if not os.path.exists(driver_dir):
        os.makedirs(driver_dir)

    # 如果驱动不存在，下载并复制到指定目录
    if not os.path.exists(driver_file):
        print("Edge驱动不存在，正在下载...")
        driver_path = EdgeChromiumDriverManager().install()
        print("Edge驱动下载地址：", driver_path)
        shutil.copy(driver_path, driver_dir)
    else:
        print("使用现有Edge驱动文件...")

    # 尝试验证驱动是否可用
    try:
        service = EdgeService(executable_path=driver_file)
        # 启动服务并立即关闭，无需创建完整会话
        driver = webdriver.Edge(service=service)
        driver.quit()
        print("Edge驱动验证通过。")
    except SessionNotCreatedException as e:
        print("检测到Edge驱动版本不匹配或无法使用，正在重新下载最新驱动...")
        driver_path = EdgeChromiumDriverManager().install()
        print("Edge驱动下载地址：", driver_path)
        shutil.copy(driver_path, driver_dir)
        print("Edge驱动已更新。")

    return driver_file


def selenium_detect_browsers_and_setup_drivers():
    def is_chrome_installed():
        try:
            with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\Wow6432Node\Google\Update\Clients") as key:
                for i in range(1024):
                    try:
                        subkey_name = winreg.EnumKey(key, i)
                        if "Google Chrome" in subkey_name:
                            with winreg.OpenKey(key, subkey_name) as subkey:
                                path, _ = winreg.QueryValueEx(subkey, "Path")
                                if os.path.isfile(path):
                                    return True, path
                    except OSError:
                        break
        except FileNotFoundError:
            pass

        default_paths = [
            os.path.expanduser(r"~\AppData\Local\Google\Chrome\Application\chrome.exe"),
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files\Google\Chrome\Application\chrome.exe"
        ]

        for path in default_paths:
            if os.path.exists(path):
                return True, path

        return False, None

    def is_edge_installed():
        try:
            with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, r"SOFTWARE\WOW6432Node\Microsoft\Edge\Application") as key:
                version, _ = winreg.QueryValueEx(key, "CurrentVersion")
                with winreg.OpenKey(key, f"{version}\\Shell\\Open\\Command") as subkey:
                    path, _ = winreg.QueryValueEx(subkey, "")
                    edge_path = path.split('"')[1]
                    if os.path.isfile(edge_path):
                        return True, edge_path
        except FileNotFoundError:
            pass

        default_paths = [
            r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
            r"C:\Program Files\Microsoft\Edge\Application\msedge.exe"
        ]

        for path in default_paths:
            if os.path.exists(path):
                return True, path

        return False, None

    def ensure_driver(browser_name):
        driver_dir = f"c:/selenium_driver/{browser_name}_driver/"
        if not os.path.exists(driver_dir):
            os.makedirs(driver_dir)

        if browser_name == 'chrome':
            driver_file = os.path.join(driver_dir, 'chromedriver.exe')
            manager = ChromeDriverManager()
        elif browser_name == 'edge':
            driver_file = os.path.join(driver_dir, 'msedgedriver.exe')
            manager = EdgeChromiumDriverManager()
        else:
            raise ValueError(f"Unsupported browser: {browser_name}")

        if not os.path.exists(driver_file):
            print(f"{browser_name.capitalize()}驱动不存在，正在下载...")
            driver_path = manager.install()
            print(f"{browser_name.capitalize()}驱动下载地址：", driver_path)
            shutil.copy(driver_path, driver_dir)
        else:
            print(f"使用现有{browser_name.capitalize()}驱动文件...")

        try:
            if browser_name == 'chrome':
                service = ChromeService(executable=driver_file)
                driver = webdriver.Chrome(service=service)
            else:
                service = EdgeService(executable=driver_file)
                driver = webdriver.Edge(service=service)
            driver.quit()
            print(f"{browser_name.capitalize()}驱动验证通过。")
        except SessionNotCreatedException:
            print(f"检测到{browser_name.capitalize()}驱动版本不匹配或无法使用，正在重新下载最新驱动...")
            driver_path = manager.install()
            print(f"{browser_name.capitalize()}驱动下载地址：", driver_path)
            shutil.copy(driver_path, driver_dir)
            print(f"{browser_name.capitalize()}驱动已更新。")
            driver_file = os.path.join(driver_dir, os.path.basename(driver_path))

        return driver_file

    result = {
        'chrome': {
            'installed': False,
            'path': '',
            'driver_path': ''
        },
        'edge': {
            'installed': False,
            'path': '',
            'driver_path': ''
        }
    }

    # Check and manage Chrome
    chrome_installed, chrome_path = is_chrome_installed()
    result['chrome']['installed'] = chrome_installed
    result['chrome']['path'] = chrome_path
    if chrome_installed:
        result['chrome']['driver_path'] = ensure_driver('chrome')

    # Check and manage Edge
    edge_installed, edge_path = is_edge_installed()
    result['edge']['installed'] = edge_installed
    result['edge']['path'] = edge_path
    if edge_installed:
        result['edge']['driver_path'] = ensure_driver('edge')

    return result


def is_workday(date_str, flag):
    """
    判断指定日期是否是工作日
    :param date_str: 指定日期的字符串形式
    :param flag: 节假日类型（0，1，2，3）中的一个值，分别表示 工作日、周末、节日、调休
    :return: True or False
    """
    years_months_days = re.findall(r'\d+', date_str)
    # 查看接口详情URL=http://timor.tech/api/holiday
    url = f"http://timor.tech/api/holiday/info/{'-'.join(years_months_days)}"
    headers = {
        'User-Agent': 'Mozilla/5.0'
    }
    while True:
        try:
            response = requests.get(url, headers=headers)
            # 打印调试信息
            print(f"URL: {url}")
            print(f"Response Text: {response.text}")
            data = response.json()
            code = data.get("code")
            if code == 0:
                day_type = data.get("type", {}).get("type")
                # 节假日类型（0，1，2，3），分别表示 工作日、周末、节日、调休
                return day_type == flag
        except Exception as e:
            print(f"网络异常: {e}")
            print(traceback.format_exc())
