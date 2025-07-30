#  ____       _____           _
# |  _ \ _   |_   _|__   ___ | |___ ____
# | |_) | | | || |/ _ \ / _ \| / __|_  /
# |  __/| |_| || | (_) | (_) | \__ \/ /
# |_|    \__, ||_|\___/ \___/|_|___/___|
#        |___/

# Copyright (c) 2024 Sidney Zhang <zly@lyzhang.me>
# PyToolsz is licensed under Mulan PSL v2.
# You can use this software according to the terms and conditions of the Mulan PSL v2.
# You may obtain a copy of Mulan PSL v2 at:
#          http://license.coscl.org.cn/MulanPSL2
# THIS SOFTWARE IS PROVIDED ON AN "AS IS" BASIS, WITHOUT WARRANTIES OF ANY KIND,
# EITHER EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO NON-INFRINGEMENT,
# MERCHANTABILITY OR FIT FOR A PARTICULAR PURPOSE.
# See the Mulan PSL v2 for more details.

# 说明
# 这是一个把pandas/polars的DataFrame数据保存成一个具有格式的excel文件。
# 例如把计算好的结算数据写到excel文件中，依照格式直接形成结算表。

from pathlib import Path
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl import Workbook
from openpyxl.worksheet import cell_range
import itertools
from openpyxl.utils.cell import get_column_letter, cols_from_range, coordinate_from_string
from collections.abc import Iterable
from typing import Self
from numbers import Number


import pandas as pd
import polars as pl

__all__ = ["transColname2Letter", "saveExcel"]

def cellBorder(types:dict = None) -> Border:
    """
    生成边框样式。
    """
    if types :
        nomorlBorder = Border(
            left = Side(**types["left"]) if "left" in types.keys() else Side(border_style="dotted"),
            right= Side(**types["right"]) if "right" in types.keys() else Side(border_style="dotted"),
            top = Side(**types["top"]) if "top" in types.keys() else Side(border_style="dotted"),
            bottom = Side(**types["bottom"]) if "bottom" in types.keys() else Side(border_style="dotted"))
    else:
        nomorlBorder = Border(left = Side(border_style="dotted"),
                              right= Side(border_style="dotted"),
                              top = Side(border_style="dotted"),
                              bottom = Side(border_style="dotted"))
    return nomorlBorder

def transColname2Letter(colnames:list[str],
                        xlsDataRange:str|None = None) -> dict[str]:
    """
    将列名转换为excel列字母。
    """
    colLetter = [i[0][0] for i in cols_from_range(xlsDataRange)]
    if len(colnames)!= len(xlsDataRange):
        raise ValueError("colnames must be equal to xlsDataRange")
    return dict(zip(colnames, colLetter))

class saveExcel(object):
    """
    保存DataFrame到excel文件。
    """
    def __init__(self, filename:str|Path, 
                 startRow:int = 1, startColumn:int = 1) -> None:
        self.__filename = filename
        self.__wb = Workbook()
        self.__ws = None
        self.__startRow = startRow
        self.__startColumn = startColumn
        self.__rowplace = None
        self.__nextPoint = None
        self.__data = None
        self.__data_length = None
        self.__data_width = None
        self.__columns_sort = None
    def __make_nextpoint(self, plus_n:str = 0) -> None:
        if plus_n < 0 :
            raise ValueError("plus_n must be greater than 0")
        if self.__rowplace is None:
            spoint = "{}{}".format(get_column_letter(self.__startColumn),self.__startRow)
            epoint = "{}{}".format(get_column_letter(self.__data_width+self.__startColumn-1),
                                   self.__startRow+plus_n)
        else:
            spoint = "{}{}".format(get_column_letter(self.__startColumn),self.__rowplace+1)
            epoint = "{}{}".format(get_column_letter(self.__data_width+self.__startColumn-1),
                                   self.__rowplace+plus_n+1)
        self.__nextPoint = (spoint, epoint)
    def actionNewSheet(self, sheetname:str|None = None, 
                       need_gridline:bool = False) -> None:
        self.__ws = self.__wb.active
        self.__ws.sheet_view.showGridLines = need_gridline
        self.__ws.title = sheetname if sheetname else "Sheet1"
    def save(self) -> None:
        self.__wb.save(self.__filename)
    def __enter__(self) -> Self:
        return self
    def __exit__(self, exc_type, exc_value, traceback) -> None:
        self.save()
    def usingData(self, data:pd.DataFrame|pl.DataFrame,
                  usingSortCols:list[str]|None = None) -> None:
        self.__data = data
        self.__data_length, self.__data_width = self.__data.shape
        if isinstance(self.__data, pl.DataFrame) :
            self.__data = self.__data.to_pandas()
        if usingSortCols :
            self.__columns_sort = usingSortCols
        else:
            self.__columns_sort = self.__data.columns.to_list()
        self.__data = self.__data[self.__columns_sort]
    def __setRowBorder(self, startpoint:int, endpoint:int, 
                     left = None, middle = None, right = None) -> None:
        """
        按照行来设置单元格边框样式。
        """
        for cell in itertools.chain(*self.__ws[startpoint:endpoint]):
            if cell.coordinate == startpoint :
                cell.border =  cellBorder(left)
            elif cell.coordinate == endpoint :
                cell.border =  cellBorder(right)
            else:
                cell.border =  cellBorder(middle)
    def __getColumnsRange(self) -> list:
        res = []
        sP = get_column_letter(self.__startColumn)
        eP = get_column_letter(self.__data_width+self.__startColumn-1)
        for i in itertools.chain(*cols_from_range("{}1:{}1".format(sP,eP))):
            tmp = coordinate_from_string(i)
            res.append(tmp[0])
        res = list(set(res))
        res.sort()
        return res
    def __columns_to_letter(self, colname:str|list[str]|None = None):
        if colname :
            allCNs = self.__data.columns.to_list()
            allLs = self.__getColumnsRange()
            if isinstance(colname, str):
                return allLs[allLs.index(colname)]
            else:
                return [allLs[allLs.index(i)] for i in colname]
        else :
            return self.__getColumnsRange()
    def __setNumberFormat(self, giveRange:str|None = None,
                          sformat:str|dict = 'General') -> None:
        """
        设定数字格式。
        """
        if giveRange :
            cellrange = giveRange
            colist = []
            for i in itertools.chain(*cols_from_range(giveRange)):
                tmp = coordinate_from_string(i)
                colist.append(tmp[0])
            colist = list(set(colist))
            colist.sort()
        else:
            cellrange = "{}:{}".format(*self.__nextPoint)
            colist = self.__getColumnsRange()
        crkey = cell_range.CellRange(range_string=cellrange).bounds 
        rolist = list(range(crkey[1],crkey[3]+1))
        if isinstance(sformat, dict):
            for icol in sformat.keys():
                if icol in colist :
                    cellist = ["{}{}".format(*x) for x in list(itertools.product([icol], rolist))]
                    for cellname in cellist:
                        self.__ws[cellname].number_format = sformat[icol]
        else:
            cellist = ["{}{}".format(*x) for x in list(itertools.product(rolist, colist))]
            for cellname in cellist:
                self.__ws[cellname].number_format = sformat
    def writeTitle(self, title:str, crossline:int = 0,
                   font_type:dict|None = None, 
                   border_type:dict|None = None, height:int = 75) -> None:
        """
        写入标题。
        """
        if font_type :
            fontype = Font(**font_type["font"]) if "font" in font_type.keys() else Font(name='微软雅黑', 
                                                                                        size=23, 
                                                                                        bold=True)
            aligtype = Alignment(**font_type["align"]) if "align" in font_type.keys() else Alignment(horizontal="center",
                                                                                                     vertical="center")
        else:
            fontype = Font(name='微软雅黑', size=23, bold=True)
            aligtype = Alignment(horizontal="center",vertical="center")
        self.__make_nextpoint(plus_n=crossline)
        self.__ws.merge_cells("{}:{}".format(*self.__nextPoint))
        self.__ws[self.__nextPoint[0]].value = title
        self.__ws[self.__nextPoint[0]].font = fontype
        self.__ws[self.__nextPoint[0]].alignment = aligtype
        if crossline > 0 :
            for i in range(self.__ws[self.__nextPoint[0]].row,(self.__ws[self.__nextPoint[1]].row + 1)):
                self.__ws.row_dimensions[i].height = height
        else:
            self.__ws.row_dimensions[self.__ws[self.__nextPoint[0]].row].height = height
        if border_type :
            self.__setRowBorder(*self.__nextPoint,
                                left = border_type["left"],
                                middle = border_type["middle"],
                                right = border_type["right"])
        self.__rowplace = self.__ws[self.__nextPoint[1]].row
    def __writeRowData(self, value:Iterable, font_type:dict|None = None, 
                       border_type:dict|None = None, crossline:int = 0,
                       numberformat:dict|str = "General",
                       height:int = 50) -> None:
        if font_type :
            fontype = Font(**font_type["font"]) if "font" in font_type.keys() else Font(name='微软雅黑', 
                                                                                        size=11.5, 
                                                                                        bold=True)
            aligtype = Alignment(**font_type["align"]) if "align" in font_type.keys() else Alignment(horizontal="center",
                                                                                                     vertical="center",
                                                                                                     wrapText = True)
            filltype = PatternFill(**font_type["fill"]) if "fill" in font_type.keys() else PatternFill(patternType=None)
        else:
            fontype = Font(name='微软雅黑', size=11.5, bold=True)
            aligtype = Alignment(horizontal="center",vertical="center")
            filltype = PatternFill(patternType=None)
        self.__make_nextpoint(plus_n=crossline)
        for irow in value:
            tmp = dict(zip(self.__getColumnsRange(),irow))
            self.__ws.append(tmp)
        for cell in itertools.chain(*self.__ws[self.__nextPoint[0]:self.__nextPoint[1]]) :
            cell.font = fontype
            cell.alignment = aligtype
            cell.fill = filltype
        s = self.__ws[self.__nextPoint[0]].row
        b = self.__ws[self.__nextPoint[1]].row
        for i in range(s,(b+1)):
            self.__ws.row_dimensions[i].height = height
        if border_type :
            self.__setRowBorder(*self.__nextPoint,
                                left = border_type["left"],
                                middle = border_type["middle"],
                                right = border_type["right"])
        if isinstance(numberformat, dict):
            self.__setNumberFormat(sformat=numberformat)
        self.__rowplace = b
    def writeData(self, height:int = 33, col_crossline:int = 0,
                  font_type:dict|None = None, col_font_type:dict|None = None,
                  borde_type:dict|None = None, col_border_type:dict|None = None,
                  numberformat:dict|str|None = None) -> None :
        """
        把数据按照格式要求写入Sheet表格。
        """
        colmname = [self.__data.columns.to_numpy().tolist()]
        thisdata = self.__data.to_numpy().tolist()
        self.__writeRowData(colmname, font_type = col_font_type,
                            border_type=col_border_type, crossline=col_crossline,
                            height=height)
        for rri in thisdata :
            self.__writeRowData([rri], font_type = font_type,
                            border_type=borde_type, crossline=0,
                            numberformat=numberformat,
                            height=height)
    def writeSummaryData(self, height:int = 33, agg_fun:str = "sum",
                         agg_cols:list[str]|None = None,
                         pass_cols:list[str]|None = None, pass_seq:str = "--",
                         row_name:str = "总计",
                         name_merge_cols:list[str]|None = None,
                         font_type:dict|None = None, 
                         borde_type:dict|None = None,
                         numberformat:dict|str|None = None) -> None :
        if font_type :
            fontype = Font(**font_type["font"]) if "font" in font_type.keys() else Font(name='微软雅黑', 
                                                                                        size=11.5)
            sfontype = Font(**font_type["font"]) if "font" in font_type.keys() else Font(name='微软雅黑', 
                                                                                         size=11.5, 
                                                                                         bold=True)
            aligtype = Alignment(**font_type["align"]) if "align" in font_type.keys() else Alignment(horizontal="center",
                                                                                                     vertical="center")
            filltype = PatternFill(**font_type["fill"]) if "fill" in font_type.keys() else PatternFill(patternType=None)
        else:
            fontype = Font(name='微软雅黑', size=11.5)
            sfontype = Font(name='微软雅黑', size=11.5, bold=True)
            aligtype = Alignment(horizontal="center",vertical="center")
            filltype = PatternFill(patternType=None)
        self.__make_nextpoint()
        if name_merge_cols:
            stGML = len(name_merge_cols)
            if name_merge_cols != self.__data.columns.to_list()[:stGML] :
                raise ValueError("name_merge_cols must be the first {} columns".format(stGML))
        else :
            stGML = 0
        res_data = [row_name] * stGML
        for i in self.__data.columns.to_list()[stGML:]:
            if agg_cols:
                if i in agg_cols:
                    res_data.append(self.__data[i].agg(agg_fun))
                    continue
            if pass_cols:
                if i in pass_cols:
                    res_data.append(pass_seq)
                    continue
            if pd.api.types.is_numeric_dtype(self.__data[i].dtype) :
                res_data.append(self.__data[i].agg(agg_fun))
            else:
                res_data.append(self.__data[i].head().iloc[0])
        if stGML > 0 :
            self.__ws.merge_cells("{}:{}{}".format(self.__nextPoint[0],
                                                   get_column_letter(
                                                       stGML+self.__ws[self.__nextPoint[0]].column-1
                                                    ),
                                                   self.__ws[self.__nextPoint[1]].row))
            self.__ws[self.__nextPoint[0]].value = row_name
            self.__ws[self.__nextPoint[0]].font = sfontype
            self.__ws[self.__nextPoint[0]].alignment = aligtype
            self.__ws[self.__nextPoint[0]].fill = filltype
        middlepoint = "{}{}".format(get_column_letter(stGML+self.__ws[self.__nextPoint[0]].column),
                                    self.__ws[self.__nextPoint[0]].row)
        p = 0
        for cell in itertools.chain(*self.__ws[middlepoint:self.__nextPoint[1]]):
            cell.value = res_data[stGML+p]
            p += 1
            cell.font = fontype
            cell.alignment = aligtype
            cell.fill = filltype
            if numberformat:
                if isinstance(numberformat, dict) :
                    cLLs = self.__getColumnsRange()
                    cell.number_format = numberformat[cLLs[stGML+p]]
                else:
                    cell.number_format = numberformat
        if borde_type :
            self.__setRowBorder(*self.__nextPoint,
                                left = borde_type["left"],
                                middle = borde_type["middle"],
                                right = borde_type["right"])
        self.__rowplace = self.__ws[self.__nextPoint[1]].row
    def writeSpecialThings(self, giveRange:str, 
                           value:str|Number|None = None, 
                           height:int = 33, 
                           font_type:dict|None = None,
                           border_type:dict|None = None,
                           numberformat:dict|None = None) -> None :
        if font_type :
            fontype = Font(**font_type["font"]) if "font" in font_type.keys() else Font(name='微软雅黑', size=11.5)
            aligtype = Alignment(**font_type["align"]) if "align" in font_type.keys() else Alignment(horizontal="center",
                                                                                                     vertical="center")
            filltype = PatternFill(**font_type["fill"]) if "fill" in font_type.keys() else PatternFill(patternType=None)
        else:
            fontype = Font(name='微软雅黑', size=11.5, bold=True)
            aligtype = Alignment(horizontal="center",vertical="center")
            filltype = PatternFill(patternType=None)
        bordertype = {"left":None,"middle":None,"right":None}
        if border_type :
            bordertype = {**bordertype, **border_type}
        if ":" in giveRange :
            sP = giveRange.split(":")[0]
            self.__ws.merge_cells(giveRange)
            self.__ws[sP].value = value
            self.__ws[sP].font = fontype
            self.__ws[sP].alignment = aligtype
            self.__ws[sP].fill = filltype
            GRBs = cell_range.CellRange(range_string=giveRange).bounds
            for i in range(GRBs[1],GRBs[3]+1):
                self.__ws.row_dimensions[i].height = height
                if border_type :
                    aP = "{}{}".format(get_column_letter(GRBs[0]),i)
                    eP = "{}{}".format(get_column_letter(GRBs[2]),i)
                    self.__setRowBorder(aP, eP,
                                    left = bordertype["left"],
                                    middle = bordertype["middle"],
                                    right = bordertype["right"])
        else:
            self.__ws[giveRange].value = value
            self.__ws[giveRange].font = fontype
            self.__ws[giveRange].alignment = aligtype
            self.__ws[giveRange].fill = filltype
            if border_type :
                self.__setRowBorder(giveRange, giveRange,
                                    left = bordertype["left"],
                                    middle = bordertype["middle"],
                                    right = bordertype["right"])
            self.__ws.row_dimensions[
                cell_range.CellRange(range_string=giveRange).bounds[1]
            ].height = height
    def setColumnsWidth(self, width:int = 17.5, colname:str|None = None,
                        rangeName:str|None = None) -> None :
        if colname is not None and rangeName is not None :
            raise ValueError("colname and rangeName can not be set at the same time.")
        if self.__data is None :
            raise ValueError("You must add data before you can set the column width.")
        if colname :
            tmp = self.__columns_to_letter(colname)
            self.__ws.column_dimensions[tmp].width = width
        elif rangeName :
            self.__ws.column_dimensions[rangeName].width = width
        else :
            for icol in self.__getColumnsRange():
                self.__ws.column_dimensions[icol].width = width


if __name__ == "__main__":
    """
    这是一个使用示例。
    """
    data = pd.DataFrame({"a": [1,2,3], 
                         "b": [4,5,6],
                         "c": [7,8,9],
                         "d": [14,15,16],
                         "e": [24,25,26],
                         "f": [34,35,36]})
    with saveExcel("D:\\test.xlsx",startRow=2, startColumn=2) as wEmodel :
        wEmodel.usingData(data)
        wEmodel.actionNewSheet(sheetname="Sheet测试")
        wEmodel.writeTitle("测试标题")
        wEmodel.writeData(col_font_type={"fill":{"patternType":"solid","fgColor":"000066CC"},
                                         "font":{"name":'微软雅黑', "size":11.5, "bold":True,"color":"00FFFFFF"}},
                          col_border_type={"left":{"left":{"border_style":"thick"},
                                                   "top":{"border_style":"thick"},
                                                   "bottom":{"border_style":"double"}},
                                            "middle":{"top":{"border_style":"thick"},
                                                      "bottom":{"border_style":"double"}},
                                            "right":{"right":{"border_style":"thick"},
                                                     "top":{"border_style":"thick"},
                                                     "bottom":{"border_style":"double"}}},
                          font_type={"font":{"name":'微软雅黑', "size":11.5}},
                          borde_type={"left":{"left":{"border_style":"thick"},
                                              "top":{"border_style":None}},
                                      "middle":{"top":{"border_style":None}},
                                      "right":{"right":{"border_style":"thick"},
                                               "top":{"border_style":None}}},
                          numberformat={"B":"0.00%","C":"0.00%","D":"0.00%","E":"0.00%"})
        wEmodel.writeSummaryData(
            name_merge_cols=["a","b"],pass_cols=["e"], numberformat="#,##0.00;-#,##0.00;-;@",
            font_type={"fill":{"patternType":"solid","fgColor":"00CCFFCC"}},
            borde_type={
                "left":{"left":{"border_style":"thick"},
                        "top":{"border_style":"thin"},
                        "bottom":{"border_style":"thick"}},
                "middle":{"top":{"border_style":"thin"},
                          "bottom":{"border_style":"thick"}},
                "right":{"right":{"border_style":"thick"},
                         "top":{"border_style":"thin"},
                         "bottom":{"border_style":"thick"}}})