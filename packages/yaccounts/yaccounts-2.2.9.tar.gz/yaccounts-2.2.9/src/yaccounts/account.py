from datetime import datetime
import enum
import pathlib

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import SheetFormatProperties
import pandas as pd

from .budget_categories import AccountCategory, BudgetType
from .payroll import PayrollData
from .config import CACHE_WORKDAY_DATA
from .run_report import run_activity_report, run_gift_report, run_grant_report
from .yutils import TermColors, print_color, warn, workday_str_amount_to_float, error


class _ColumnIndices(enum.Enum):
    """Enum for column indices in the DataFrame."""

    CATEGORY = 1
    BUDGET = 2
    PREV_ACTUALS = 3
    JAN = 4
    FEB = 5
    MAR = 6
    APR = 7
    MAY = 8
    JUN = 9
    JUL = 10
    AUG = 11
    SEP = 12
    OCT = 13
    NOV = 14
    DEC = 15
    TOTAL_THIS_YEAR = 16
    BALANCE = 17


_CATEGORY_ORDER = [
    AccountCategory.WAGES,
    AccountCategory.BENEFITS,
    AccountCategory.STUDENT_AID,
    AccountCategory.SUPPLIES,
    AccountCategory.CAPITAL,
    AccountCategory.TRAVEL,
    AccountCategory.CONTRACT_SERVICES,
    AccountCategory.UNALLOCATED,
    AccountCategory.INDIRECT_COSTS,
]


class Account:
    def __init__(self, account_code, year):

        self.account_code = account_code
        if not (
            self.account_code.startswith("GR")
            or self.account_code.startswith("AC")
            or self.account_code.startswith("GF")
        ):
            error(
                f"Account code {self.account_code} does not start with 'GR', 'AC' or 'GF'. "
                "Other account codes are not supported yet."
            )
        self.year = year

        # Define columns: 'category', then months 1-12
        columns = (
            ["Category", "Budget", "Prev Actuals"]
            + [datetime(1900, month, 1).strftime("%b") for month in range(1, 13)]
            + ["YTD", "Balance"]
        )
        self._df = pd.DataFrame(columns=columns)

    def get_workday_data(self, dr):
        actuals_cached_path = pathlib.Path("actuals.pkl")

        # Get actuals
        if actuals_cached_path.exists() and CACHE_WORKDAY_DATA:
            print("Loading actuals from cache...")
            self._df = pd.read_pickle(actuals_cached_path)
        else:
            self._get_actuals_data(dr)
            if CACHE_WORKDAY_DATA:
                self._df.to_pickle("actuals.pkl")

        self._finalize_workday_data()

    def _finalize_workday_data(self):

        # Convert all columns except 'Category' to float
        for col in self._df.columns:
            if col != "Category":
                self._df[col] = self._df[col].apply(
                    lambda x: float(x) if x != "" else ""
                )

        # Replace NaN values with 0
        self._df = self._df.fillna(0.0)

        if self.account_code.startswith("GR"):
            # Map each category to an order number
            order_map = {category: idx for idx, category in enumerate(_CATEGORY_ORDER)}
            self._df["Order"] = self._df["Category"].map(order_map)

            # Sort by the order
            self._df = self._df.sort_values(by="Order").drop(columns=["Order"])
            self._df = self._df.reset_index(drop=True)
        else:
            # Sort by Category
            self._df = self._df.sort_values(by="Category").reset_index(drop=True)

        # YTD column is the sum of all months
        self._df["YTD"] = self._df[
            [datetime(1900, month, 1).strftime("%b") for month in range(1, 13)]
        ].sum(axis=1)

        # Balance is Budget - Prev Actuals - YTD
        self._df["Balance"] = (
            self._df["Budget"] - self._df["Prev Actuals"] - self._df["YTD"]
        )

        # Add a totals row
        total_row = self._df.sum(numeric_only=True)
        total_row["Category"] = "Total"
        total_row["Budget"] = total_row["Budget"]
        total_row["Prev Actuals"] = total_row["Prev Actuals"]
        total_row["Balance"] = total_row["Balance"]
        total_row["YTD"] = total_row["YTD"]
        self._df = pd.concat([self._df, pd.DataFrame([total_row])], ignore_index=True)

    def add_payroll_data(self, payroll_data):
        """Add payroll data to the account DataFrame."""
        print_color(TermColors.BLUE, "Adding payroll data to the actuals report")

        if not isinstance(payroll_data, PayrollData):
            error(
                "payroll_data must be an instance of PayrollData. "
                f"Got {type(payroll_data)} instead."
            )

        payroll_df = payroll_data.df

        ##################### Process payroll_df #####################
        print(f"There are {len(payroll_df)} payroll entries.")

        # Filter down to relevant grant, where 'Grant' starts with self.account_code
        if self.account_code.startswith("GR"):
            payroll_df = payroll_df[
                payroll_df["Grant"].str.startswith(f"{self.account_code} ")
            ].copy()
        elif self.account_code.startswith("AC"):
            payroll_df = payroll_df[
                payroll_df["Activity"].str.startswith(f"{self.account_code} ")
            ].copy()
        elif self.account_code.startswith("GF"):
            payroll_df = payroll_df[
                payroll_df["Gift"].str.startswith(f"{self.account_code} ")
            ].copy()
        else:
            raise ValueError("Unsupported account code for payroll data.")
        print(f"There are {len(payroll_df)} payroll entries for {self.account_code}.")

        # Add a 'Month' column based on 'Budget Date'
        payroll_df["Month"] = pd.to_datetime(payroll_df["Budget Date"]).dt.month

        # Clean the 'Amount' column
        payroll_df["Amount"] = payroll_df["Amount"].apply(workday_str_amount_to_float)

        # Step 2: Group by Employee and Month
        payroll_summary = (
            payroll_df.groupby(["Employee", "Month"])["Amount"].sum().reset_index()
        )

        # Step 3: Pivot: one row per Employee, one column per month
        # and format the month names
        payroll_pivot = payroll_summary.pivot(
            index="Employee", columns="Month", values="Amount"
        ).fillna(0)
        month_cols = [datetime(1900, m, 1).strftime("%b") for m in range(1, 13)]
        payroll_pivot.columns = [
            datetime(1900, int(m), 1).strftime("%b") for m in payroll_pivot.columns
        ]

        # Convert the data to float
        payroll_pivot = payroll_pivot.astype(float)

        # Make sure all months are there
        for m in month_cols:
            if m not in payroll_pivot.columns:
                payroll_pivot[m] = 0.0

        # Prepare the df for merging with actuals_df
        payroll_pivot["Budget"] = ""
        payroll_pivot["Prev Actuals"] = ""
        payroll_pivot["Balance"] = ""
        payroll_pivot["YTD"] = payroll_pivot[month_cols].sum(axis=1)
        # Intent employee names a bit
        payroll_pivot["Category"] = payroll_pivot.index.map(lambda x: f"    {x}")
        final_columns = (
            ["Category", "Budget", "Prev Actuals"] + month_cols + ["YTD", "Balance"]
        )
        payroll_pivot = payroll_pivot[final_columns]
        print(
            f"There is data for {len(payroll_pivot)} employees within these payroll entries."
        )
        if payroll_pivot.empty:
            warn(
                "No payroll data found for this account. "
                "Skipping payroll data addition."
            )
            return

        # Merge with the actuals_df
        wages_category = (
            AccountCategory.WAGES
            if self.account_code.startswith("GR")
            else "5000:Salaries and Wages"
        )
        wages_indices = self._df.index[self._df["Category"] == wages_category]
        if wages_indices.empty:
            error(
                f"Could not find the wages category ('{wages_category}') in the DataFrame."
            )
        wages_index = wages_indices[0]
        before = self._df.iloc[: wages_index + 1]
        after = self._df.iloc[wages_index + 1 :]
        self._df = pd.concat([before, payroll_pivot, after], ignore_index=True)

    def to_excel(self, output_path=None):
        if output_path is None:
            output_path = f"{self.account_code}_{self.year}.xlsx"
        print_color(TermColors.GREEN, f"Creating {output_path}")

        formatted_df = self._df.copy()

        # Convert category enum to string for Excel export
        formatted_df["Category"] = formatted_df["Category"].apply(
            lambda x: x.value if isinstance(x, AccountCategory) else str(x)
        )

        formatted_df.to_excel(output_path, index=False, engine="openpyxl")
        self._format_excel(output_path)

    def _format_excel(self, excel_filename):
        """Format the Excel file with specific styles."""
        wb = load_workbook(excel_filename)
        ws = wb.active
        currency_format = "#,##0.00"

        # Insert Title Row
        ws.insert_rows(1)
        title_text = f"{self.account_code} - Year {self.year} (Generated {datetime.now().strftime('%b %d, %Y')})"
        ws.cell(row=1, column=1).value = title_text

        num_columns = ws.max_column
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)

        title_cell = ws.cell(row=1, column=1)
        title_cell.font = Font(size=18, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Define a thin black border
        thin_border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

        # === Insert new header row ===
        ws.insert_rows(2)  # Insert a new row 2

        # Reassign header values manually after insert
        headers = [
            "Category",
            "Budget",
            "Prev Actuals",
            "Jan",
            "Feb",
            "Mar",
            "Apr",
            "May",
            "Jun",
            "Jul",
            "Aug",
            "Sep",
            "Oct",
            "Nov",
            "Dec",
            "YTD",
            "Balance",
        ]

        for col, header_text in enumerate(headers, start=1):
            ws.cell(row=3, column=col).value = header_text

        # Add "Actuals" merged across Jan to YTD
        first_actuals_col = _ColumnIndices.JAN.value  # 4
        last_actuals_col = _ColumnIndices.TOTAL_THIS_YEAR.value  # 16
        ws.merge_cells(
            start_row=2,
            start_column=first_actuals_col,
            end_row=2,
            end_column=last_actuals_col,
        )
        actuals_cell = ws.cell(row=2, column=first_actuals_col)
        actuals_cell.value = "Actuals"
        actuals_cell.alignment = Alignment(horizontal="center", vertical="center")
        actuals_cell.font = Font(bold=True)

        # Apply borders to the merged "Actuals" cells
        for col in range(first_actuals_col, last_actuals_col + 1):
            ws.cell(row=2, column=col).border = thin_border

        # Merge vertically: Category, Budget, Prev Actuals, Balance
        for col in [
            _ColumnIndices.CATEGORY.value,
            _ColumnIndices.BUDGET.value,
            _ColumnIndices.PREV_ACTUALS.value,
            _ColumnIndices.BALANCE.value,
        ]:
            ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
            merged_cell = ws.cell(row=2, column=col)
            merged_cell.value = headers[col - 1]  # Set correct text!
            merged_cell.alignment = Alignment(horizontal="center", vertical="center")
            merged_cell.font = Font(bold=True)

            # Apply borders to the merged cells
            ws.cell(row=2, column=col).border = thin_border
            ws.cell(row=3, column=col).border = (
                thin_border  # Also border the lower merged part
            )

        # Apply borders to all header cells in row 3
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=3, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.border = thin_border

        # Merge vertically: Category, Budget, Prev Actuals, Balance
        for col in [
            _ColumnIndices.CATEGORY.value,
            _ColumnIndices.BUDGET.value,
            _ColumnIndices.PREV_ACTUALS.value,
            _ColumnIndices.BALANCE.value,
        ]:
            ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
            merged_cell = ws.cell(row=2, column=col)
            merged_cell.value = headers[col - 1]  # Set correct text!
            merged_cell.alignment = Alignment(horizontal="center", vertical="center")
            merged_cell.font = Font(bold=True)

        # Style the second header row (row 3)
        for col in range(1, num_columns + 1):
            header_cell = ws.cell(row=3, column=col)
            header_cell.alignment = Alignment(horizontal="center", vertical="center")
            header_cell.font = Font(bold=True)

        # Freeze top 3 rows
        ws.freeze_panes = ws["A4"]

        # === Format the data rows ===
        fill_5 = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )
        fill_15 = PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
        )
        fill_total = PatternFill(
            start_color="595959", end_color="595959", fill_type="solid"
        )
        yellow_fill = PatternFill(
            start_color="FFFF00", end_color="FFFF00", fill_type="solid"
        )

        rows_to_group = []
        shade_toggle = True

        for idx, row in enumerate(
            ws.iter_rows(min_row=4, min_col=1), start=4
        ):  # Data now starts at row 4
            category_cell = row[0]
            is_employee = isinstance(
                category_cell.value, str
            ) and category_cell.value.startswith("    ")

            if not is_employee:
                fill = fill_5 if shade_toggle else fill_15
                for cell in row:
                    cell.fill = fill
                shade_toggle = not shade_toggle
            else:
                rows_to_group.append(idx)

            for cell in row[1:]:
                cell.number_format = currency_format

        if rows_to_group:
            start_row = min(rows_to_group)
            end_row = max(rows_to_group)
            ws.row_dimensions.group(start_row, end_row, outline_level=1, hidden=False)

        ws.sheet_format = SheetFormatProperties(outlineLevelRow=1)

        # Set column widths
        ws.column_dimensions[get_column_letter(_ColumnIndices.CATEGORY.value)].width = (
            30
        )
        ws.column_dimensions[get_column_letter(_ColumnIndices.BUDGET.value)].width = 13
        ws.column_dimensions[
            get_column_letter(_ColumnIndices.PREV_ACTUALS.value)
        ].width = 13
        ws.column_dimensions[
            get_column_letter(_ColumnIndices.TOTAL_THIS_YEAR.value)
        ].width = 13
        ws.column_dimensions[get_column_letter(_ColumnIndices.BALANCE.value)].width = 13
        for month in range(_ColumnIndices.JAN.value, _ColumnIndices.DEC.value + 1):
            ws.column_dimensions[get_column_letter(month)].width = 12

        # Highlight the total row
        total_row = len(ws["A"])
        for cell in ws[total_row]:
            cell.fill = fill_total
            cell.font = Font(bold=True, color="FFFFFF")

        balance_cell = ws.cell(row=total_row, column=_ColumnIndices.BALANCE.value)
        balance_cell.fill = yellow_fill
        balance_cell.font = Font(bold=True, color="000000")

        wb.save(excel_filename)

    def _get_actuals_data(self, dr):
        """Use workday to get actuals"""

        end_month = 12 if self.year != datetime.now().year else datetime.now().month
        month_year_list = [
            datetime(self.year, month, 1) for month in range(1, end_month + 1)
        ]

        for month in month_year_list:
            if self.account_code.startswith("GR"):
                data = run_grant_report(dr, self.account_code, month)
            elif self.account_code.startswith("AC"):
                data = run_activity_report(dr, self.account_code, month)
            elif self.account_code.startswith("GF"):
                data = run_gift_report(dr, self.account_code, month)

            for category in data:
                # If the actuals_df does not have a row for this category, add it
                if category not in self._df["Category"].values:
                    new_row = {"Category": category}
                    self._df = pd.concat(
                        [self._df, pd.DataFrame([new_row])], ignore_index=True
                    )

                # Update the actuals_df for this category and month
                self._df.loc[self._df["Category"] == category, month.strftime("%b")] = (
                    data[category][BudgetType.ACTUALS]
                )

                # If January, update previous actuals
                if month.month == 1:
                    self._df.loc[self._df["Category"] == category, "Prev Actuals"] = (
                        data[category][BudgetType.ACTUALS_PREV]
                    )

                # If month is last month of loop, update budget
                if month.month == end_month:
                    self._df.loc[self._df["Category"] == category, "Budget"] = data[
                        category
                    ][BudgetType.BUDGET]
