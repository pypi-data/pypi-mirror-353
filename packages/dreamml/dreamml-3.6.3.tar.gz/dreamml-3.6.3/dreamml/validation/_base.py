import string
from typing import List
import numpy as np
import pandas as pd
from tqdm.auto import tqdm
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter

from dreamml.logging import get_logger
from dreamml.reports.pdf_maker import convert_excel_to_pdf, pdf_maker_available

_logger = get_logger(__name__)


class ValidationStyles:
    """A class to apply various styles and formats to Excel validation reports.

    Attributes:
        dir (str): Path to save the validation report.
        writer (pd.ExcelWriter): Pandas Excel writer object.
        letters (str): Uppercase ASCII letters used for column identification.
        sheets (dict): Dictionary of sheet names to worksheet objects.
        wb (Workbook): openpyxl workbook object.
        _create_pdf (bool): Flag to determine whether to create a PDF version of the report.
    """

    def __init__(
        self,
        path_to_save: str = "./validation_report.xlsx",
        create_file: bool = True,
        create_pdf: bool = False,
    ):
        """Initializes the ValidationStyles instance.

        Args:
            path_to_save (str, optional): Path to save the validation report. Defaults to "./validation_report.xlsx".
            create_file (bool, optional): Whether to create an Excel file. Defaults to True.
            create_pdf (bool, optional): Whether to create a PDF version of the report. Defaults to False.
        """
        self.dir = path_to_save
        if create_file:
            # Intended for use with development reports to prevent creation of validation report files.
            self.writer = pd.ExcelWriter(
                path=path_to_save,
                engine="openpyxl",
                datetime_format="dd-mm-yy hh:mm:ss",
            )
            self.letters = string.ascii_uppercase
            self.sheets = self.writer.sheets
            self.wb = self.writer.book
        self._create_pdf = create_pdf

    def create_pdf(self):
        """Converts the Excel validation report to PDF if creation is enabled.

        Raises:
            Warning: If PDF conversion is unavailable due to missing fonts.
        """
        if not self._create_pdf:
            return

        pdf_file = self.dir[:-5] if self.dir.endswith(".xlsx") else self.dir
        pdf_file += ".pdf"

        if not pdf_maker_available:
            _logger.warning("Can't make pdf as fonts aren't installed.")
            return
        convert_excel_to_pdf(self.dir, pdf_file)

    def add_cell_width(self, sheet_name: str):
        """Sets the width of each column in the specified Excel sheet based on content.

        The width is determined by the maximum length of the headers and the data within each column,
        plus an additional buffer of 2 units.

        Args:
            sheet_name (str): The name of the sheet in the Excel workbook to adjust.
        """
        ws = self.sheets[sheet_name]

        dim_holder = DimensionHolder(worksheet=ws)

        for col in ws.columns:
            max_col_width = 0
            for cell in col:
                col_width = len(str(cell.value))
                if max_col_width < col_width:
                    max_col_width = col_width

            max_col_width += 2

            dim_holder[get_column_letter(col[0].column)] = ColumnDimension(
                ws, min=col[0].column, max=col[0].column, width=max_col_width
            )

        ws.column_dimensions = dim_holder

    def add_header_color(
        self,
        data: pd.DataFrame,
        sheet_name: str,
        startrow: int = 0,
        startcol: int = 0,
        color: str = "77d496",
    ):
        """Applies background color and styling to the header row of the Excel sheet.

        Args:
            data (pd.DataFrame): The DataFrame being written to Excel.
            sheet_name (str): The name of the sheet in the Excel workbook.
            startrow (int, optional): The starting row index for data insertion. Defaults to 0.
            startcol (int, optional): The starting column index for data insertion. Defaults to 0.
            color (str, optional): Hex color code for the header background. Defaults to "77d496".
        """
        ws = self.sheets[sheet_name]
        startrow += 1  # openPyXl iterates from 1, not 0
        startcol += 1

        drm_header_format = NamedStyle(name="drm_header_format")
        drm_header_format.font = Font(bold=True)
        drm_header_format.fill = PatternFill(
            start_color=color, end_color=color, fill_type="solid"
        )
        drm_header_format.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
        brd = Side(border_style="thin", color="000000")
        drm_header_format.border = Border(left=brd, right=brd, top=brd, bottom=brd)

        style_list = [x.name for x in self.wb._named_styles]
        if drm_header_format.name not in style_list:
            self.wb.add_named_style(drm_header_format)

        for col in ws.iter_cols(
            min_row=startrow,
            max_row=startrow,
            min_col=startcol,
            max_col=startcol + data.shape[1] - 1,
        ):
            for cell in col:
                cell.style = "drm_header_format"

    def add_numeric_format(
        self,
        data: pd.DataFrame,
        sheet_name: str,
        startrow: int = 0,
        startcol: int = 0,
        min_value: int = 10,
        to_eventrate_format: List[int] = [],
        to_integer_format: List[int] = [],
    ):
        """Applies numeric formatting to specified columns in the Excel sheet.

        Formats columns as integers, percentages, or decimals based on provided parameters.

        Args:
            data (pd.DataFrame): The DataFrame being written to Excel.
            sheet_name (str): The name of the sheet in the Excel workbook.
            startrow (int, optional): The starting row index for data insertion. Defaults to 0.
            startcol (int, optional): The starting column index for data insertion. Defaults to 0.
            min_value (int, optional): Minimum value to determine decimal formatting. Defaults to 10.
            to_eventrate_format (List[int], optional): List of column indices (1-based) to format as percentages. Defaults to [].
            to_integer_format (List[int], optional): List of column indices (1-based) to format as integers. Defaults to [].
        """
        ws = self.sheets[sheet_name]
        startrow += 1  # openPyXl iterates from 1, not 0
        startcol += 1

        for col in ws.iter_cols(
            min_row=startrow,
            max_row=startrow + data.shape[0],
            min_col=startcol,
            max_col=startcol + data.shape[1] - 1,
        ):
            for cell in col:
                if isinstance(cell.value, (float, int)):
                    if cell.column in to_integer_format:
                        cell.number_format = "0"
                    elif cell.column in to_eventrate_format:
                        cell.number_format = FORMAT_PERCENTAGE_00
                    else:
                        cell.number_format = "0.00" if cell.value < min_value else "0"

    def add_table_borders(
        self, data: pd.DataFrame, sheet_name: str, startrow: int = 0, startcol: int = 0
    ):
        """Adds borders around the specified table in the Excel sheet.

        Args:
            data (pd.DataFrame): The DataFrame representing the table.
            sheet_name (str): The name of the sheet in the Excel workbook.
            startrow (int, optional): The starting row index for the table. Defaults to 0.
            startcol (int, optional): The starting column index for the table. Defaults to 0.
        """
        ws = self.sheets[sheet_name]
        startrow += 1  # openPyXl iterates from 1, not 0
        startcol += 1

        # Left border
        for col in ws.iter_cols(
            min_row=startrow,
            max_row=startrow + data.shape[0],
            min_col=startcol,
            max_col=startcol,
        ):
            for cell in col:
                cell.border = Border(
                    left=Side(border_style="thin", color="000000"),
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                )

        # Right border
        for col in ws.iter_cols(
            min_row=startrow,
            max_row=startrow + data.shape[0],
            min_col=startcol + data.shape[1] - 1,
            max_col=startcol + data.shape[1] - 1,
        ):
            for cell in col:
                cell.border = Border(
                    left=cell.border.left,
                    right=Side(border_style="thin", color="000000"),
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                )

        # Bottom border
        for row in ws.iter_rows(
            min_row=startrow + data.shape[0],
            max_row=startrow + data.shape[0],
            min_col=startcol,
            max_col=startcol + data.shape[1] - 1,
        ):
            for cell in row:
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=Side(border_style="thin", color="000000"),
                )

    def add_bottom_table_borders(
        self, data: pd.DataFrame, sheet_name: str, startrow: int = 0, startcol: int = 0
    ):
        """Adds top and bottom borders to the table in the Excel sheet.

        Args:
            data (pd.DataFrame): The DataFrame representing the table.
            sheet_name (str): The name of the sheet in the Excel workbook.
            startrow (int, optional): The starting row index for the table. Defaults to 0.
            startcol (int, optional): The starting column index for the table. Defaults to 0.
        """
        ws = self.sheets[sheet_name]
        startrow += 1  # openPyXl iterates from 1, not 0

        drm_result_format = NamedStyle(name="drm_result_format")
        drm_result_format.font = Font(bold=True)
        brd = Side(border_style="thin", color="000000")
        drm_result_format.border = Border(left=brd, right=brd, top=brd, bottom=brd)

        style_list = [x.name for x in self.wb._named_styles]
        if drm_result_format.name not in style_list:
            self.wb.add_named_style(drm_result_format)

        for col in ws.iter_rows(
            min_row=startrow + data.shape[0],
            max_row=startrow + data.shape[0],
            min_col=startcol,
            max_col=startcol + data.shape[1],
        ):
            for cell in col:
                cell.style = "drm_result_format"

    def add_traffic_light_color(self, sheet_name: str):
        """Applies traffic light colors (green, yellow, red) to cells based on their values.

        Args:
            sheet_name (str): The name of the sheet in the Excel workbook.
        """
        ws = self.sheets[sheet_name]

        drm_traffic_light_format = NamedStyle(name="drm_traffic_light_format")
        drm_traffic_light_format.alignment = Alignment(
            wrap_text=True, horizontal="center"
        )
        drm_traffic_light_format.font = Font(bold=True)
        brd = Side(border_style="thin", color="000000")
        drm_traffic_light_format.border = Border(
            left=brd, right=brd, top=brd, bottom=brd
        )

        style_list = [x.name for x in self.wb._named_styles]
        if drm_traffic_light_format.name not in style_list:
            self.wb.add_named_style(drm_traffic_light_format)

        for col in ws.columns:
            for cell in col:
                if cell.value == "green":
                    cell.style = "drm_traffic_light_format"
                    cell.fill = PatternFill(
                        start_color="66CC99", end_color="66CC99", fill_type="solid"
                    )
                if cell.value == "yellow":
                    cell.style = "drm_traffic_light_format"
                    cell.fill = PatternFill(
                        start_color="FFFF33", end_color="FFFF33", fill_type="solid"
                    )
                if cell.value == "red":
                    cell.style = "drm_traffic_light_format"
                    cell.fill = PatternFill(
                        start_color="CC0000", end_color="CC0000", fill_type="solid"
                    )


class BaseTest:
    """A base class for testing machine learning models with validation capabilities."""

    def _predict(self, X) -> np.ndarray:
        """Applies the model to the input data to generate predictions.

        If a vectorizer is present, it transforms the input features before prediction.
        Supports models with `predict_proba`, `transform`, or `predict` methods.

        Args:
            X (pd.DataFrame): Feature matrix for prediction.

        Returns:
            np.ndarray: Array of predicted values.

        Raises:
            AttributeError: If the estimator does not have a supported prediction method.
        """
        if hasattr(self, "vectorizer") and self.vectorizer is not None:
            X = self.vectorizer.transform(X)

        if hasattr(self.estimator, "predict_proba"):
            return self.estimator.predict_proba(X[self.used_features])[:, 1]
        if hasattr(self.estimator, "transform"):
            return self.estimator.transform(X)

        return self.estimator.predict(X)

    def label_binarizing(self, task: str, label_binarizer, labels: List[str], target_name: str, **eval_set):
        """Binarizes labels for multiclass classification tasks.

        Transforms labels into a binary format suitable for evaluation metrics that require binary inputs.

        Args:
            task (str): The type of task, e.g., "multiclass".
            label_binarizer: An instance of a label binarizer.
            labels (List[str]): List of label names.
            target_name (str): Name of the target variable.
            **eval_set: Evaluation datasets to be binarized.

        Returns:
            dict: A dictionary of evaluation sets with binarized labels.
        """
        if task == "multiclass":
            for sample_name, (X_sample, y_sample) in tqdm(eval_set.items()):
                if len(y_sample.shape) == 1 and label_binarizer is not None:
                    data = label_binarizer.transform(y_sample)
                    y_sample = pd.DataFrame(
                        data=data, columns=labels, index=y_sample.index
                    )
                eval_set[sample_name] = (X_sample, y_sample)

        return eval_set

    def _calculate_macro_score(
        self, metric_class, y_true: np.ndarray, y_pred: np.ndarray
    ) -> float:
        """Calculates the macro-average of a specified metric across all classes.

        Handles cases with NaN values by excluding them from the calculation.

        Args:
            metric_class: A metric function/class that takes true and predicted values.
            y_true (np.ndarray): True labels matrix (n_samples, n_classes).
            y_pred (np.ndarray): Predicted probabilities matrix (n_samples, n_classes).

        Returns:
            float: The macro-averaged metric across classes.

        Raises:
            ValueError: If no valid classes are found for metric calculation.
        """
        y_true = y_true.values if isinstance(y_true, pd.DataFrame) else y_true
        y_pred = y_pred.values if isinstance(y_pred, pd.DataFrame) else y_pred

        metric_list = []
        num_classes = y_true.shape[1]
        mask = ~np.isnan(y_true)

        for class_idx in range(num_classes):
            mask_by_class_idx = mask[:, class_idx]
            y_true_idx_class = y_true[:, class_idx][mask_by_class_idx]
            y_pred_idx_class = y_pred[:, class_idx][mask_by_class_idx]

            if len(np.unique(y_true_idx_class)) == 1:
                continue
            if len(np.unique(y_pred_idx_class)) == 1:
                metric_list.append(0)
                continue

            metric_list.append(metric_class(y_true_idx_class, y_pred_idx_class))
        
        if not metric_list:
            raise ValueError("No valid classes found for metric calculation.")
        
        return np.mean(metric_list)