import sys
import os
import json
import time
import uuid
from pathlib import Path
from copy import deepcopy
from typing import Optional, List, Dict, Tuple, Any

import shap
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from dreamml.validation._base import ValidationStyles

import dreamml as dml
from dreamml.modeling.metrics import metrics_mapping
from dreamml.modeling.models.estimators import BaseModel
from dreamml.features.categorical import CategoricalFeaturesTransformer

from dreamml.validation.classification._gini_test import GiniTest
from dreamml.validation.classification._ks_test import KSTest
from dreamml.validation.classification._f_test import FTest
from dreamml.validation.classification._ciec_test import CIECTest

from dreamml.validation.classification._segment_performance_test import SegmentPerformanceTest
from dreamml.validation.classification._simple_model_comparison_test import SimpleModelComparisonTest
from dreamml.validation.classification._model_inference_time_test import ModelInferenceTimeTest
from dreamml.validation.classification._single_dataset_performance_test import SingleDatasetPerformanceTest
from dreamml.validation.classification._train_test_performance_test import TrainTestPerformanceTest
from dreamml.validation.classification._roc_report_test import RocReportTest
from dreamml.validation.classification._confusion_matrix_report_test import ConfusionMatrixReportTest
from dreamml.validation.classification._boosting_overfit_test import BoostingOverfitTest
from dreamml.validation.classification._prediction_drift_test import PredictionDriftTest
from dreamml.validation.classification._calibration_score_test import CalibrationScoreTest




from dreamml.validation.classification._data_statistics import DataStatisticsTest
from dreamml.validation.classification._permutation_importance import PermutationImportanceChecker
from dreamml.validation._task_description import TaskDescriptionTest
from dreamml.validation._task_description import BusinessCaseDescription
from dreamml.validation.wrappers._estimator import Estimator

from dreamml.configs.config_storage import ConfigStorage
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from dreamml.utils.get_last_experiment_directory import get_experiment_dir_path
from dreamml.logging import get_logger
from dreamml.logging.monitoring import ReportStartedLogData, ReportFinishedLogData
from dreamml.utils.prepare_artifacts_config import ClassificationArtifactsConfig

_logger = get_logger(__name__)


def prepare_validation_test_config(config, raise_exception: bool = True):
    """
    Prepare the validation test configuration based on the provided config.

    Args:
        config (dict): Configuration dictionary containing validation parameters.
        raise_exception (bool, optional): Whether to raise an exception if the experiment directory is not found. Defaults to True.

    Returns:
        dict: A dictionary containing the prepared validation test configuration.

    Raises:
        FileNotFoundError: If the experiment directory is not found and raise_exception is set to True.
    """
    experiment_dir_path = get_experiment_dir_path(
        config.get("results_path", ""),
        experiment_dir_name=config.get("dir_name"),
        use_last_experiment_directory=config.get(
            "use_last_experiment_directory", False
        ),
        raise_exception=raise_exception,
    )
    experiment_config = ConfigStorage.from_pretrained(f"{experiment_dir_path}/config")
    return {
        "n_folds": config.get("n_folds", 4),
        "psi_threshold": config.get("psi_threshold", 0.2),
        "bootstrap_samples": config.get("bootstrap_samples", 50),
        "fairness_threshold": config.get("fairness_threshold", 0.5),
        "fairness_features": config.get("fairness_features", []),
        "validation_method_for_model_training": experiment_config.pipeline.validation_type,
        "traffic_lights": experiment_config.validation.traffic_lights,
    }


def prepare_artifacts_config(
    config: dict,
    experiment_config: ConfigStorage = None,
    model: BaseModel = None,
    encoder: CategoricalFeaturesTransformer = None,
) -> Tuple[Any, Any]:
    """
    Prepare the artifacts configuration for the validation report.

    This function generates the necessary configuration to pass to the ValidationReport object.

    Args:
        config (dict): Configuration dictionary specifying the final model name and the directory containing experiment artifacts.
        experiment_config (ConfigStorage, optional): Pretrained experiment configuration. Defaults to None.
        model (BaseModel, optional): The trained model instance. Defaults to None.
        encoder (CategoricalFeaturesTransformer, optional): The categorical features transformer. Defaults to None.

    Returns:
        Tuple[Any, Any]: A tuple containing the artifacts configuration dictionary and the evaluation set.

    Raises:
        None
    """
    prepare_artifacts = ClassificationArtifactsConfig(
        config=config,
        experiment_config=experiment_config,
        model=model,
        encoder=encoder,
    )
    artifacts_config, eval_set = prepare_artifacts.prepare_artifacts_config()
    return artifacts_config, eval_set


class ValidationReport(ValidationStyles):
    """
    ValidationReport generates a comprehensive report for the final model to be submitted for validation or to business stakeholders.

    The report includes various statistical tests, feature importances, model parameters, and other relevant information.
    """

    def __init__(
        self,
        config,
        estimator,
        metric_name,
        vectorizer: callable = None,
        path_to_save: str = "./val_report.xlsx",
        used_features: Optional[List[str]] = None,
        categorical_features: Optional[List[str]] = None,
        create_file: bool = True,
        metric_col_name: str = "gini",
        metric_params: dict = None,
        images_dir_path: str = None,
        task: str = "binary",
        subtask: str = "tabular",  # [tabular, nlp]
        multiclass_artifacts: Optional[Dict] = None,
        custom_model: bool = False,
        user_config=None,
        text_column: str = None,
        text_preprocessed_column: str = None,
        group_column: str = None,
        time_column: str = None,
        create_pdf: bool = False,
        number_of_simulations_1_1: int = 200,
        number_of_simulations_3_2: int = 100
    ):
        """
        Initialize the ValidationReport with necessary configurations and estimator.

        Args:
            config (dict): Configuration dictionary for the validation report.
            estimator: The model estimator used for validation.
            metric_name (str): Name of the metric used for evaluation.
            vectorizer (callable, optional): Function to vectorize input data. Defaults to None.
            path_to_save (str, optional): Path where the validation report will be saved. Defaults to "./val_report.xlsx".
            used_features (Optional[List[str]], optional): List of features used in the model. Defaults to None.
            categorical_features (Optional[List[str]], optional): List of categorical features. Defaults to None.
            create_file (bool, optional): Whether to create the report file. Defaults to True.
            metric_col_name (str, optional): Column name for the metric. Defaults to "gini".
            metric_params (dict, optional): Parameters for the metric. Defaults to None.
            images_dir_path (str, optional): Directory path to save images. Defaults to None.
            task (str, optional): Type of task (e.g., "binary"). Defaults to "binary".
            subtask (str, optional): Subtask type (e.g., "tabular", "nlp"). Defaults to "tabular".
            multiclass_artifacts (Optional[Dict], optional): Artifacts for multiclass tasks. Defaults to None.
            custom_model (bool, optional): Indicates if a custom model is used. Defaults to False.
            user_config (optional): User-specific configuration. Defaults to None.
            text_column (str, optional): Name of the text column. Defaults to None.
            text_preprocessed_column (str, optional): Name of the preprocessed text column. Defaults to None.
            group_column (str, optional): Name of the group column. Defaults to None.
            time_column (str, optional): Name of the time column. Defaults to None.
            create_pdf (bool, optional): Whether to create a PDF version of the report. Defaults to False.
            number_of_simulations_1_1 (int, optional): Number of simulations for test 1.1. Defaults to 200.
            number_of_simulations_3_2 (int, optional): Number of simulations for test 3.2. Defaults to 100.

        Raises:
            None
        """
        self.artifacts_config = {
            "estimator": estimator,
            "used_features": used_features,
            "vectorizer": vectorizer,
            "categorical_features": categorical_features,
            "task": task,
            "subtask": subtask,
            "metric_name": metric_name,
            "images_dir_path": images_dir_path,
            "multiclass_artifacts": multiclass_artifacts,
            "metric_params": metric_params,
            "text_column": text_column,
            "text_preprocessed_column": text_preprocessed_column,
            "log_target_transformer": None,  # Placeholder indicating that log_target will not be applied
            "group_column": group_column,
            "time_column": time_column,
            "number_of_simulations_1_1": number_of_simulations_1_1,
            "number_of_simulations_3_2": number_of_simulations_3_2,
            "custom_model": custom_model,
        }
        self.custom_model = custom_model
        self.user_config = user_config
        super().__init__(path_to_save, create_file, create_pdf)
        self.validation_test_config = prepare_validation_test_config(
            config, raise_exception=(not self.custom_model)
        )

        artifacts_config = deepcopy(self.artifacts_config)
        estimator_artifacts_config = {
            k: v
            for k, v in artifacts_config.items()
            if k
            not in [
                "text_column",
                "text_preprocessed_column",
                "subtask",
                "images_dir_path",
                "multiclass_artifacts",
                "group_column",
                "metric_params",
                "time_column",
            ]
        }

        if hasattr(estimator, "model"):
            artifacts_config = deepcopy(estimator_artifacts_config)
            artifacts_config["estimator"] = estimator.model

        elif hasattr(estimator, "estimator"):
            artifacts_config["estimator"] = estimator.estimator

        self.estimator = Estimator(
            estimator=artifacts_config["estimator"],
            vectorizer=artifacts_config["vectorizer"],
            log_target_transformer=artifacts_config["log_target_transformer"],
            used_features=artifacts_config["used_features"],
            categorical_features=artifacts_config["categorical_features"],
            task=artifacts_config["task"],
            metric_name=artifacts_config["metric_name"],
        )

        self.config = config
        self.test_results = {}
        self.test_5_results = []
        self.test_5_1_result = None
        self.images_dir_path = images_dir_path
        self.sheet_descriptions = {}
        # self.hyperlinks_for_total_results_page = {}
        self.metric_name = metric_name
        self.metric_col_name = metric_col_name
        self.metric_params = metric_params
        self.multiclass_artifacts = multiclass_artifacts

    @property
    def stability_block_light(self):
        """
        Get the final traffic light status for the "Model Stability" block.

        This property evaluates the results of stability tests and determines the overall status as red, yellow, or green.

        Returns:
            pd.DataFrame: A DataFrame containing the expected results of the stability tests and the overall status.
        """
        alternative = np.where(
            "red" in self.test_5_results,
            "red",
            np.where("yellow" in self.test_5_results, "yellow", "green"),
        )
        result = pd.DataFrame(
            {
                "Expected Result of Test 5.2": [self.test_5_1_result],
                "Worst Result of Other Tests": [alternative.tolist()],
            }
        )
        result["Expected Result for 'Model Stability' Block"] = "-"
        if self.test_5_1_result == "green":
            result["Expected Result for 'Model Stability' Block"] = np.where(
                alternative.tolist() == "red", "yellow", "green"
            )
        elif self.test_5_1_result == "yellow":
            result["Expected Result for 'Model Stability' Block"] = np.where(
                alternative.tolist() == "red", "red", "yellow"
            )
        else:
            result["Expected Result for 'Model Stability' Block"] = "red"

        self.test_results["Model Stability"] = result[
            "Expected Result for 'Model Stability' Block"
        ].values.tolist()[0]
        return result

    def _create_business_task_description_page(self, **data):
        """
        Create the business task description page in the validation report.

        This page includes details about the business objective, data collection dates, data selection criteria, target variable definition, and the ML algorithm used.

        Args:
            **data: Arbitrary keyword arguments specific to the business task description.

        Returns:
            None

        Raises:
            None
        """
        _logger.info("0. Collecting statistics about the business task.")
        bcd = BusinessCaseDescription(
            self.artifacts_config, self.validation_test_config
        )
        stats = bcd._create_description(**data)
        stats.to_excel(self.writer, sheet_name="Business Task Description", index=False)
        self.sheet_descriptions["Business Task Description"] = ""

        ws = self.writer.sheets["Business Task Description"]
        b_task_cell_format = NamedStyle(name="b_task_cell_format")
        b_task_cell_format.font = Font(color="808080")
        b_task_cell_format.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
        brd = Side(border_style="thin", color="000000")
        b_task_cell_format.border = Border(left=brd, right=brd, top=brd, bottom=brd)

        self.wb.add_named_style(b_task_cell_format)

        b_task_row_title_format = NamedStyle(name="b_task_row_title_format")
        b_task_row_title_format.font = Font(bold=True)
        b_task_row_title_format.alignment = Alignment(
            horizontal="center", vertical="center"
        )
        brd = Side(border_style="thin", color="000000")
        b_task_row_title_format.border = Border(
            left=brd, right=brd, top=brd, bottom=brd
        )

        self.wb.add_named_style(b_task_row_title_format)

        b_task_col_title_format = NamedStyle(name="b_task_col_title_format")
        b_task_col_title_format.font = Font(bold=True)
        b_task_col_title_format.fill = PatternFill(
            start_color="00CC99", end_color="00CC99", fill_type="solid"
        )
        b_task_col_title_format.alignment = Alignment(
            horizontal="center", vertical="center"
        )
        brd = Side(border_style="thin", color="000000")
        b_task_col_title_format.border = Border(
            left=brd, right=brd, top=brd, bottom=brd
        )

        self.wb.add_named_style(b_task_col_title_format)

        business_task_text = (
            "<Enter the model name>\n"
            "e.g., Customer churn model for Premier service package"
        )
        task_description_text = (
            "<Enter the model description: modeling purpose, "
            "which business process the model is built for, what data is used>\n"
            "e.g., Model to identify bank customers likely to close the Premier Service Package. \n"
            "The model is built using data from the data center on customer profile closures/openings. \n"
            "Customer profile tribe data Mass Personalization is used as factors."
        )
        report_dt_description_text = (
            "<Specify the reporting dates> \n" "e.g., 31.01.2024, 28.02.2024, 31.03.2024"
        )
        target_description_text = "<Enter the definition of the target event/variable>"
        data_selection_text = (
            "<Enter population selection criteria: filters, exclusions>\n"
            "e.g., 1. Customers with information in the customer profile view on the reporting date.\n"
        )
        used_algo_text = "<dreamml_base XGBClassifier>"
        ws.merge_cells("B2:E2")
        ws.merge_cells("B3:E3")
        ws.merge_cells("B4:D4")
        ws.merge_cells("B5:D5")
        ws.merge_cells("B6:E6")
        ws.merge_cells("B7:E7")

        ws["B2"] = business_task_text
        ws["B3"] = task_description_text
        ws["B4"] = report_dt_description_text
        ws["B5"] = data_selection_text
        ws["B6"] = target_description_text
        ws["B7"] = used_algo_text
        ws["E4"] = report_dt_description_text
        ws["E5"] = data_selection_text
        ws["A1"] = "Parameter"
        ws["A2"] = "Business Task"
        ws["A3"] = "Task Description"
        ws["A4"] = "Data Collection Dates"
        ws["A5"] = "Observation Selection"
        ws["A6"] = "Target Variable Description"
        ws["A7"] = "ML Algorithm Used"
        ws["B1"] = "Train Sample"
        ws["C1"] = "Valid Sample"
        ws["D1"] = "Test Sample"
        ws["E1"] = "Out-Of-Time Sample"

        for row in ws["B2":"E7"]:
            for cell in row:
                cell.style = "b_task_cell_format"
        for row in ws["A1":"E1"]:
            for cell in row:
                cell.style = "b_task_col_title_format"
        for row in ws["A2":"A7"]:
            for cell in row:
                cell.style = "b_task_row_title_format"

        for i in range(1, 6):
            ws.column_dimensions[get_column_letter(i)].width = 30
        for i in range(1, 8):
            ws.row_dimensions[i].height = 75

        # Hyperlink to the table of contents
        ws["A9"] = "<<< Return to Table of Contents"
        ws["A9"].hyperlink = f"#'Table of Contents'!A1"
        ws["A9"].style = "Hyperlink"

    def _create_pipeline_description_page(self, **data):
        """
        Create the pipeline description page in the validation report.

        This page includes descriptions of all pipeline stages used in the model.

        Args:
            **data: Arbitrary keyword arguments specific to the pipeline description.

        Returns:
            None

        Raises:
            None
        """
        _logger.info("1. Collecting statistics about the used pipeline.")
        pst = TaskDescriptionTest(self.artifacts_config, self.validation_test_config)
        stats = pst._create_description(**data)
        stats.to_excel(self.writer, sheet_name="Pipeline Description", index=False)
        self.add_table_borders(stats, sheet_name="Pipeline Description")
        self.add_cell_width(sheet_name="Pipeline Description")
        self.add_header_color(stats, sheet_name="Pipeline Description", color="00CC99")
        self.add_numeric_format(stats, sheet_name="Pipeline Description", startcol=1)

        ws = self.sheets["Pipeline Description"]
        # Hyperlink to the table of contents
        ws[f"A{stats.shape[0] + 3}"] = "<<< Return to Table of Contents"
        ws[f"A{stats.shape[0] + 3}"].hyperlink = f"#'Table of Contents'!A1"
        ws[f"A{stats.shape[0] + 3}"].style = "Hyperlink"

        self.sheet_descriptions["Pipeline Description"] = (
            "Page describing all pipeline stages"
        )

    def _create_dml_info_page(self):
        """
        Create the dreamml_base information page in the validation report.

        This page provides information about the dreamml_base module and the kernel version used.

        Args:
            None

        Returns:
            None

        Raises:
            None
        """
        kernel_folder_path = Path(sys.executable).parent.parent
        kernel_name = os.path.basename(kernel_folder_path)
        try:
            kernel_path = os.path.join(kernel_folder_path, "jh_kernel", "kernel.json")
            with open(kernel_path, "r") as file:
                kernel = json.load(file)
            kernel_version = kernel["display_name"]
        except FileNotFoundError:
            kernel_version = kernel_name

        df = pd.DataFrame(
            {
                "Module": ["dreamml_base", "Kernel"],
                "Version": [dml.__version__, kernel_version],
            }
        )
        df.to_excel(self.writer, sheet_name="V", startrow=0, startcol=0, index=False)
        self.add_header_color(df, sheet_name="V", color="00CC99", startcol=0)
        self.add_cell_width(
            sheet_name="V",
        )

        ws = self.sheets["V"]
        msg = "Made in dreamml_base"
        ws[f"A{df.shape[0] + 4}"] = msg
        ws[f"A{df.shape[0] + 4}"].font = Font(italic=True)

    def _create_data_description_page(self, **data):
        """
        Create the data description page in the validation report.

        This page includes statistical information about the data used in the model.

        Args:
            **data: Arbitrary keyword arguments specific to the data description.

        Returns:
            None

        Raises:
            None
        """
        _logger.info("2. Collecting statistics about the data.")
        transformer = DataStatisticsTest(
            self.artifacts_config,
            self.validation_test_config,
        )
        result = transformer.transform(**data)

        startrows = [
            0,
            2 + len(result[0]),
            4 + len(result[0]) + len(result[1]),
        ]

        if self.artifacts_config["task"] not in ["multiclass", "multilabel"]:
            eventrate_columns = [[4], [], []]
            integer_columns = [[2, 3], [2, 3], [2]]
        else:
            idx_events_columns, idx_eventrate_columns = [], []
            for col_idx, column in enumerate(result[0].columns):
                if "events" in column:
                    idx_events_columns.append(col_idx + 1)
                if "eventrate" in column:
                    idx_eventrate_columns.append(col_idx + 1)

            eventrate_columns = [idx_eventrate_columns, [], []]
            integer_columns = [idx_events_columns, [2, 3], [2]]

        for result_df, start_index, e_col, int_col in zip(
            result, startrows, eventrate_columns, integer_columns
        ):
            result_df.to_excel(
                self.writer,
                startrow=start_index,
                sheet_name="Data Description",
                index=False,
            )
            self.add_table_borders(
                result_df, sheet_name="Data Description", startrow=start_index
            )
            self.add_header_color(
                result_df,
                sheet_name="Data Description",
                startrow=start_index,
                color="00CC99",
            )
            self.add_numeric_format(
                result_df,
                sheet_name="Data Description",
                startrow=start_index,
                to_eventrate_format=e_col,
                to_integer_format=int_col,
            )

        self.add_cell_width(sheet_name="Data Description")

        ws = self.sheets["Data Description"]
        # Hyperlink to the table of contents
        ws["E7"] = "<<< Return to Table of Contents"
        ws["E7"].hyperlink = f"#'Table of Contents'!A1"
        ws["E7"].style = "Hyperlink"

        self.sheet_descriptions["Data Description"] = "Page with data statistics"

    def _create_features_importance_page(self, **data):
        """
        Create the feature importance page in the validation report.

        This page displays the importance of features based on SHAP values.

        Args:
            **data: Arbitrary keyword arguments specific to feature importance.

        Returns:
            None

        Raises:
            None
        """
        msg = "7. Evaluating feature importance using SHAP values."
        _logger.info(msg)
        if "valid" in data:
            x, _ = data["valid"]
        elif "train" in data:
            x, _ = data["train"]
        elif "test" in data:
            x, _ = data["test"]
        elif "OOT" in data:
            x, _ = data["OOT"]

        shap_values, imp = self.estimator.get_shap_importance(x)
        imp.to_excel(
            self.writer, sheet_name="Feature Importance", startrow=0, index=False
        )
        self.add_numeric_format(imp, sheet_name="Feature Importance", min_value=100)
        self.add_table_borders(imp, sheet_name="Feature Importance")
        self.add_header_color(imp, sheet_name="Feature Importance", color="00CC99")
        self.add_cell_width(sheet_name="Feature Importance")
        ws = self.sheets["Feature Importance"]

        if isinstance(shap_values, np.ndarray):
            plt.clf()
            shap.initjs()
            shap.summary_plot(
                shap_values,
                features=x[self.estimator.used_features],
                feature_names=self.estimator.used_features,
                show=False,
            )
            plt.savefig(
                os.path.join(self.images_dir_path, "shap_summary_plot.png"),
                bbox_inches="tight",
                pad_inches=0.1,
            )
            plt.close()

            img = Image(os.path.join(self.images_dir_path, "shap_summary_plot.png"))
            img.anchor = f"A{imp.shape[0] + 2}"
            ws.add_image(img)

        # Hyperlink to the table of contents
        ws["E2"] = "<<< Return to Table of Contents"
        ws["E2"].hyperlink = f"#'Table of Contents'!A1"
        ws["E2"].style = "Hyperlink"

        msg = "Page with feature importance based on SHAP values."
        self.sheet_descriptions["Feature Importance"] = msg

    def _create_permutation_importance_page(self, **data):
        """
        Create the permutation importance page in the validation report.

        This page assesses feature importance based on permutation tests.

        Args:
            **data: Arbitrary keyword arguments specific to permutation importance.

        Returns:
            None

        Raises:
            None
        """
        """
        Permutation Importance
        """
        msg = "8. Assessing feature importance based on permutations."
        _logger.info(msg)

        perm_test = PermutationImportanceChecker(
            writer=self.writer,
            model=self.artifacts_config["estimator"],
            features_list=self.artifacts_config["used_features"],
            cat_features=self.artifacts_config["categorical_features"],
            images_dir_path=self.images_dir_path,
            metric_name=self.metric_name,
            metric_col_name=self.metric_col_name,
            metric_params=self.metric_params,
            task=self.artifacts_config["task"],
        )
        perm_df = perm_test.validate(**data)

        perm_df.to_excel(
            self.writer, sheet_name="Permutation Importance", startrow=0, index=False
        )
        self.add_table_borders(perm_df, sheet_name="Permutation Importance")
        self.add_header_color(
            perm_df, sheet_name="Permutation Importance", color="00CC99"
        )
        self.add_cell_width(sheet_name="Permutation Importance")
        self.add_bottom_table_borders(perm_df, sheet_name="Permutation Importance")
        self.add_numeric_format(
            perm_df, sheet_name="Permutation Importance", min_value=100
        )

        ws = self.sheets["Permutation Importance"]
        ws["E2"] = "Permutation importance - metric for feature importance in the built model"
        ws["E3"] = (
            "Calculated as the relative change in model quality metric when feature values are shuffled"
        )
        ws["E5"] = (
            "Factors relevancy - the proportion of factors with importance of 20% or more of the factor with maximum importance"
        )
        ws["E7"] = "* - this test is informative"

        # Hyperlink to the table of contents
        ws["E9"] = "<<< Return to Table of Contents"
        ws["E9"].hyperlink = f"#'Table of Contents'!A1"
        ws["E9"].style = "Hyperlink"

        msg = "Page assessing feature importance based on permutations."
        self.sheet_descriptions["Permutation Importance"] = msg
        plt.close()

        img = Image(f"{self.images_dir_path}/Permutation importance.png")
        img.anchor = f"A{perm_df.shape[0] + 4}"
        ws.add_image(img)

    def _create_model_params_page(self, **data):
        """
        Create the model parameters page in the validation report.

        This page lists the hyperparameters of the model.

        Args:
            **data: Arbitrary keyword arguments specific to model parameters.

        Returns:
            None

        Raises:
            None
        """
        msg = "9. Collecting model hyperparameters."
        _logger.info(msg)
        try:
            params = self.estimator.get_estimator_params
        except KeyError:
            params = pd.DataFrame(columns=["Hyperparameter", "Value"])

        params.to_excel(
            self.writer,
            sheet_name="Model Hyperparameters",
            index=False,
            startrow=0,
        )
        self.add_table_borders(params, sheet_name="Model Hyperparameters")
        self.add_header_color(
            params, sheet_name="Model Hyperparameters", color="00CC99"
        )
        self.add_cell_width(
            sheet_name="Model Hyperparameters",
        )

        ws = self.sheets["Model Hyperparameters"]
        if params.empty:
            ws["A2"] = "LAMA does not have an interface for hyperparameters."
        # Hyperlink to the table of contents
        ws["E2"] = "<<< Return to Table of Contents"
        ws["E2"].hyperlink = f"#'Table of Contents'!A1"
        ws["E2"].style = "Hyperlink"

        self.sheet_descriptions["Model Hyperparameters"] = ""

    def _create_gini_coef(self, **data):
        """
        Create the Gini Coefficient page in the validation report.

        This page presents the Gini coefficient of the model across all samples.

        Args:
            **data: Arbitrary keyword arguments specific to the Gini coefficient.

        Returns:
            None

        Raises:
            None
        """
        _sheet_name = "Gini Coefficient"
        msg = "3. Model Gini Coefficient"
        _logger.info(msg)
        metric = metrics_mapping["gini"](
            model_name=self.config["model_name"],
            task="binary"
        )

        gini_test = GiniTest(self.estimator, metric)
        result, traffic_light_gini_degradation = gini_test.transform(**data)

        self.test_results[_sheet_name] = traffic_light_gini_degradation

        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=3,
            index=False,
        )

        self.add_table_borders(result, sheet_name=_sheet_name)
        # self.add_header_color(result, sheet_name="Gini Coefficient", color="00CC99")
        self.add_numeric_format(result, sheet_name=_sheet_name, min_value=100)
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws = self.sheets[_sheet_name]
        msg = "Test 1 - Model Gini Coefficient"

        ws[f"A1"] = msg
        ws["A9"] = "<<< Return to Table of Contents"
        ws["A9"].hyperlink = f"#'Table of Contents'!A1"
        ws["A9"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg

    def _create_ks_coef(self, **data):
        """
        Create the Kolmogorov-Smirnov statistic page in the validation report.

        This page presents the Kolmogorov-Smirnov statistic of the model across all samples.

        Args:
            **data: Arbitrary keyword arguments specific to the KS statistic.

        Returns:
            None

        Raises:
            None
        """
        _sheet_name = "Kolmogorov-Smirnov Statistic"
        msg = "4. Kolmogorov-Smirnov Statistic"
        _logger.info(msg)

        ks_test = KSTest(self.estimator)
        result, result_traffic_light = ks_test.transform(**data)

        self.test_results[_sheet_name] = result_traffic_light

        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=3,
            index=False,
        )

        self.add_table_borders(result, sheet_name=_sheet_name)
        self.add_numeric_format(result, sheet_name=_sheet_name, min_value=100)
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws = self.sheets[_sheet_name]
        msg = "Test 2 - Kolmogorov-Smirnov Statistic"

        ws[f"A1"] = msg
        ws["A9"] = "<<< Return to Table of Contents"
        ws["A9"].hyperlink = f"#'Table of Contents'!A1"
        ws["A9"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg
        # self.hyperlinks_for_total_results_page["2"] = _sheet_name

    def _create_f1_coef(self, **data):
        """
        Create the F1 Score page in the validation report.

        This page presents the F1 score of the model across all samples.

        Args:
            **data: Arbitrary keyword arguments specific to the F1 score.

        Returns:
            None

        Raises:
            None
        """
        _sheet_name = "F1 Score"
        msg = "5. F1 Score of the Model"
        _logger.info(msg)
        metric = metrics_mapping["f1_score"](
            model_name=self.config["model_name"],
            task="binary"
        )

        f1_test = FTest(self.estimator, metric)
        result, result_traffic_light = f1_test.transform(**data)

        self.test_results[_sheet_name] = result_traffic_light

        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=3,
            index=False,
        )

        self.add_table_borders(result, sheet_name=_sheet_name)
        self.add_numeric_format(result, sheet_name=_sheet_name, min_value=100)
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws = self.sheets[_sheet_name]
        msg = "Test 3 - F1 Score of the Model"

        ws[f"A1"] = msg
        ws["A9"] = "<<< Return to Table of Contents"
        ws["A9"].hyperlink = f"#'Table of Contents'!A1"
        ws["A9"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg
        # self.hyperlinks_for_total_results_page["1"] = _sheet_name

    def _create_ciec_coef(self, **data):
        """
        Create the Conditional Information Entropy Coefficient (CIEC) page in the validation report.

        This page presents the CIEC of the model across all samples.

        Args:
            **data: Arbitrary keyword arguments specific to the CIEC.

        Returns:
            None

        Raises:
            None
        """
        _sheet_name = "Conditional Information Entropy Coefficient"
        msg = "6. Conditional Information Entropy Coefficient"
        _logger.info(msg)

        ciec_test = CIECTest(self.estimator)
        result, result_traffic_light = ciec_test.transform(**data)

        self.test_results[_sheet_name] = result_traffic_light

        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=3,
            index=False,
        )

        self.add_table_borders(result, sheet_name=_sheet_name)
        self.add_numeric_format(result, sheet_name=_sheet_name, min_value=100)
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws = self.sheets[_sheet_name]
        msg = "Test 4 - Conditional Information Entropy Coefficient"

        ws[f"A1"] = msg
        ws["A9"] = "<<< Return to Table of Contents"
        ws["A9"].hyperlink = f"#'Table of Contents'!A1"
        ws["A9"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg

    def _create_segment_performance_coef(self, **data):
        _sheet_name = "Segment Performance"
        msg = "7. Segment Performance"
        _logger.info(msg)

        segment_performance_test = SegmentPerformanceTest(self.estimator, self.images_dir_path)
        result, traffic_light_segment_performance = segment_performance_test.transform(**data)

        self.test_results[_sheet_name] = traffic_light_segment_performance

        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=5,
            index=False,
        )

        self.add_table_borders(result, sheet_name=_sheet_name)
        self.add_numeric_format(result, sheet_name=_sheet_name, min_value=100)
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws = self.sheets[_sheet_name]
        msg = "Test 5 - Segment Performance"

        i = 12
        for sample in result.keys():
            img = Image(os.path.join(self.images_dir_path, f"{sample}_segment_performance_heatmap.png"))
            img.anchor = f"A{i}"
            ws.add_image(img)
            i += 27

        ws[f"A1"] = msg
        ws["A3"] = "Display performance score segmented by 2 top (or given) features in a heatmap"
        ws["A9"] = "<<< Return to Table of Contents"
        ws["A9"].hyperlink = f"#'Table of Contents'!A1"
        ws["A9"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg

    def _create_simple_model_comparison_coef(self, **data):
        _sheet_name = "Simple Model Comparison"
        msg = "8. Simple Model Comparison"
        _logger.info(msg)

        simple_model_comparison_test = SimpleModelComparisonTest(self.estimator)
        result, traffic_light_simple_model_comparison = simple_model_comparison_test.transform(**data)

        self.test_results[_sheet_name] = traffic_light_simple_model_comparison

        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=5,
            index=False,
        )

        self.add_table_borders(result, sheet_name=_sheet_name)
        self.add_numeric_format(result, sheet_name=_sheet_name, min_value=100)
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws = self.sheets[_sheet_name]
        msg = "Test 6 - Simple Model Comparison"

        ws[f"A1"] = msg
        ws["A3"] = "Compare given model score to simple model score (according to given model type)"
        ws["A9"] = "<<< Return to Table of Contents"
        ws["A9"].hyperlink = f"#'Table of Contents'!A1"
        ws["A9"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg

    def _create_model_inference_time_coef(self, **data):
        _sheet_name = "Model Inference Time"
        msg = "9. Model Inference Time"
        _logger.info(msg)

        model_inference_time_test = ModelInferenceTimeTest(self.estimator)
        result, traffic_light_model_inference_time = model_inference_time_test.transform(**data)

        self.test_results[_sheet_name] = traffic_light_model_inference_time

        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=3,
            index=False,
        )

        self.add_table_borders(result, sheet_name=_sheet_name)
        self.add_numeric_format(result, sheet_name=_sheet_name, min_value=100)
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws = self.sheets[_sheet_name]
        msg = "Test 7 - Model Inference Time"

        ws[f"A1"] = msg
        ws["A9"] = "<<< Return to Table of Contents"
        ws["A9"].hyperlink = f"#'Table of Contents'!A1"
        ws["A9"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg

    def _create_single_dataset_performance_coef(self, **data):
        _sheet_name = "Single Dataset Performance"
        msg = "10. Single Dataset Performance"
        _logger.info(msg)

        single_dataset_performance_test = SingleDatasetPerformanceTest(self.estimator)
        result_dict, traffic_light_single_dataset_performance = single_dataset_performance_test.transform(**data)

        self.test_results[_sheet_name] = traffic_light_single_dataset_performance

        i = 3
        for key, data in result_dict.items():
            data.to_excel(
                self.writer,
                sheet_name=_sheet_name,
                startrow=i,
                index=False,
            )
            self.add_table_borders(data, sheet_name=_sheet_name)
            self.add_numeric_format(data, sheet_name=_sheet_name, min_value=100)
            self.add_traffic_light_color(sheet_name=_sheet_name)
            self.add_cell_width(sheet_name=_sheet_name)

            i += (len(data) + 3)

        ws = self.sheets[_sheet_name]
        msg = "Test 8 - Single Dataset Performance"

        ws[f"A1"] = msg
        ws["E2"] = "<<< Return to Table of Contents"
        ws["E2"].hyperlink = f"#'Table of Contents'!A1"
        ws["E2"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg

    def _create_train_test_performance_coef(self, **data):
        _sheet_name = "Train Test Performance"
        msg = "11. Train Test Performance"
        _logger.info(msg)

        train_test_performance_test = TrainTestPerformanceTest(self.estimator)
        result, traffic_light_train_test_performance = train_test_performance_test.transform(**data)

        self.test_results[_sheet_name] = traffic_light_train_test_performance

        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=3,
            index=False,
        )

        self.add_table_borders(result, sheet_name=_sheet_name)
        self.add_numeric_format(result, sheet_name=_sheet_name, min_value=100)
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws = self.sheets[_sheet_name]
        msg = "Test 9 - Train Test Performance"

        ws[f"A1"] = msg
        ws["G2"] = "<<< Return to Table of Contents"
        ws["G2"].hyperlink = f"#'Table of Contents'!A1"
        ws["G2"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg

    def _create_roc_report_coef(self, **data):
        _sheet_name = "Roc Report"
        msg = "12. Roc Report"
        _logger.info(msg)

        roc_report_test = RocReportTest(self.estimator)
        result, traffic_light_roc_report = roc_report_test.transform(**data)

        self.test_results[_sheet_name] = traffic_light_roc_report

        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=3,
            index=False,
        )

        self.add_table_borders(result, sheet_name=_sheet_name)
        self.add_numeric_format(result, sheet_name=_sheet_name, min_value=100)
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws = self.sheets[_sheet_name]
        msg = "Test 10 - Roc Report"

        ws[f"A1"] = msg
        ws["A9"] = "<<< Return to Table of Contents"
        ws["A9"].hyperlink = f"#'Table of Contents'!A1"
        ws["A9"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg

    def _create_confusion_matrix_report_coef(self, **data):
        _sheet_name = "Confusion Matrix Report"
        msg = "13. Confusion Matrix Report"
        _logger.info(msg)

        confusion_matrix_report_test = ConfusionMatrixReportTest(self.estimator)
        result_dict, traffic_light_confusion_matrix_report = confusion_matrix_report_test.transform(**data)

        self.test_results[_sheet_name] = traffic_light_confusion_matrix_report

        i = 3
        for key, data in result_dict.items():
            data.to_excel(
                self.writer,
                sheet_name=_sheet_name,
                startrow=i,
                index=False,
            )
            self.add_table_borders(data, sheet_name=_sheet_name)
            self.add_numeric_format(data, sheet_name=_sheet_name, min_value=100)
            self.add_traffic_light_color(sheet_name=_sheet_name)
            self.add_cell_width(sheet_name=_sheet_name)
            i += (len(data) + 3)

        ws = self.sheets[_sheet_name]
        msg = "Test 11 - Confusion Matrix Report"

        ws[f"A1"] = msg
        ws["E3"] = "<<< Return to Table of Contents"
        ws["E3"].hyperlink = f"#'Table of Contents'!A1"
        ws["E3"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg

    def _create_boosting_overfit_coef(self, **data):
        _sheet_name = "Boosting Overfit"
        msg = "14. Boosting Overfit"
        _logger.info(msg)

        boosting_overfit_test = BoostingOverfitTest(self.estimator)
        result, traffic_light_boosting_overfit = boosting_overfit_test.transform(**data)

        self.test_results[_sheet_name] = traffic_light_boosting_overfit

        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=3,
            index=False,
        )

        self.add_table_borders(result, sheet_name=_sheet_name)
        self.add_numeric_format(result, sheet_name=_sheet_name, min_value=100)
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws = self.sheets[_sheet_name]
        msg = "Test 12 - Boosting Overfit"

        ws[f"A1"] = msg
        ws["A9"] = "<<< Return to Table of Contents"
        ws["A9"].hyperlink = f"#'Table of Contents'!A1"
        ws["A9"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg

    def _create_prediction_drift_coef(self, **data):
        _sheet_name = "Prediction Drift"
        msg = "15. Prediction Drift"
        _logger.info(msg)

        prediction_drift_test = PredictionDriftTest(self.estimator)
        result, traffic_light_prediction_drift = prediction_drift_test.transform(**data)

        self.test_results[_sheet_name] = traffic_light_prediction_drift

        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=3,
            index=False,
        )

        self.add_table_borders(result, sheet_name=_sheet_name)
        self.add_numeric_format(result, sheet_name=_sheet_name, min_value=100)
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws = self.sheets[_sheet_name]
        msg = "Test 13 - Prediction Drift"

        ws[f"A1"] = msg
        ws["A9"] = "<<< Return to Table of Contents"
        ws["A9"].hyperlink = f"#'Table of Contents'!A1"
        ws["A9"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg

    def _create_calibration_score_coef(self, **data):
        _sheet_name = "Calibration Score"
        msg = "16. Calibration Score"
        _logger.info(msg)

        calibration_score_test = CalibrationScoreTest(self.estimator)
        result_dict, traffic_light_calibration_score = calibration_score_test.transform(**data)

        self.test_results[_sheet_name] = traffic_light_calibration_score

        i = 3
        for key, data in result_dict.items():
            data.to_excel(
                self.writer,
                sheet_name=_sheet_name,
                startrow=i,
                index=False,
            )
            self.add_table_borders(data, sheet_name=_sheet_name)
            self.add_numeric_format(data, sheet_name=_sheet_name, min_value=100)
            self.add_traffic_light_color(sheet_name=_sheet_name)
            self.add_cell_width(sheet_name=_sheet_name)
            i += (len(data) + 3)

        ws = self.sheets[_sheet_name]
        msg = "Test 14 - Calibration Score"

        ws[f"A1"] = msg
        ws["E2"] = "<<< Return to Table of Contents"
        ws["E2"].hyperlink = f"#'Table of Contents'!A1"
        ws["E2"].style = "Hyperlink"

        self.sheet_descriptions[_sheet_name] = msg

    def _create_total_results(self):
        """
        Create the total results page in the validation report.

        This page summarizes the results of all tests conducted.

        Args:
            None

        Returns:
            None

        Raises:
            None
        """
        _sheet_name = "Total Results"

        ws = self.sheets[_sheet_name]
        msg = "Table with Total Results"
        ws["A1"] = msg

        result = pd.DataFrame(self.test_results.items())

        result.to_excel(
            self.writer,
            sheet_name=_sheet_name,
            startrow=3,
            index=False,
        )

        self.add_table_borders(result, startrow=3, sheet_name=_sheet_name)
        # self.add_numeric_format(result, sheet_name="Pipeline Description", min_value=100)
        # self.add_table_borders(result, sheet_name="Pipeline Description")
        self.add_traffic_light_color(sheet_name=_sheet_name)
        self.add_cell_width(sheet_name=_sheet_name)

        ws["A2"].value = None
        ws["B1"].value = None
        ws["A4"].value = None
        ws["B4"].value = None

        # self.add_header_color(result, sheet_name="Gini Coefficient", color="00CC99")

    def f(self, data, features):
        new_data = dict()
        for sample in data:
            X, y = data[sample]
            new_data[sample] = (X[features].copy(), y.copy())
        return new_data
    
    def create_report(self, **data):
        """
        Generate the complete validation report.

        This method orchestrates the creation of all necessary pages and sections in the validation report.

        Args:
            **data: Arbitrary keyword arguments specific to the report creation.

        Returns:
            None

        Raises:
            None
        """
        report_id = uuid.uuid4().hex

        task = self.artifacts_config["task"]
        custom_model = self.custom_model
        user_config = self.user_config
        used_features = self.artifacts_config["used_features"]
        features_data = self.f(data, used_features)

        if custom_model:
            experiment_name = None
        else:
            experiment_name = Path(self.dir).parent.parent.name

        start_time = time.time()
        _logger.monitor(
            f"Creating validation report for {task} task.",
            extra={
                "log_data": ReportStartedLogData(
                    task=task,
                    development=False,
                    custom_model=custom_model,
                    experiment_name=experiment_name,
                    report_id=report_id,
                    user_config=user_config,
                )
            },
        )

        self._create_dml_info_page()

        self._create_business_task_description_page(**data)
        self._create_pipeline_description_page(**data)

        pd.DataFrame([" "]).to_excel(self.writer, sheet_name="Total Results")
        self._create_data_description_page(**data)

        self._create_gini_coef(**data)
        self._create_ks_coef(**data)
        self._create_f1_coef(**data)
        self._create_ciec_coef(**data)
        self._create_segment_performance_coef(**features_data)
        self._create_simple_model_comparison_coef(**features_data)
        self._create_model_inference_time_coef(**features_data)
        self._create_single_dataset_performance_coef(**features_data)
        self._create_train_test_performance_coef(**features_data)
        self._create_roc_report_coef(**features_data)
        self._create_confusion_matrix_report_coef(**features_data)
        self._create_boosting_overfit_coef(**features_data)
        self._create_prediction_drift_coef(**features_data)
        self._create_calibration_score_coef(**features_data)









        try:
            self._create_features_importance_page(**data)
        except:
            pass

        if self.artifacts_config["subtask"] != "nlp":
            self._create_permutation_importance_page(**data)

        try:
            self._create_model_params_page(**data)
        except:
            pass
        self._create_total_results()

        self.writer.save()

        elapsed_time = time.time() - start_time
        _logger.monitor(
            f"Validation report for {task} task is created in {elapsed_time:.1f} seconds.",
            extra={
                "log_data": ReportFinishedLogData(
                    report_id=report_id,
                    elapsed_time=elapsed_time,
                )
            },
        )

    def create_traffic_light(self, **data):
        """
        Generate the final traffic light summary by conducting necessary tests.

        This method runs essential validation tests and determines the overall status as red, yellow, or green based on test outcomes.

        Args:
            **data: Arbitrary keyword arguments specific to the traffic light creation.

        Returns:
            List[str]: A list of results from the conducted tests, including the final status.

        Raises:
            None
        """
        results = []

        if self.config["task"] != "binary":
            return ["None" for _ in range(5)]

        metric = metrics_mapping["gini"](
            model_name=self.config["model_name"],
            task="binary"
        )
        gini_test = GiniTest(self.estimator, metric)
        result, traffic_light_gini_degradation = gini_test.transform(**data)
        results.append(traffic_light_gini_degradation)

        ks_test = KSTest(self.estimator)
        result, result_traffic_light = ks_test.transform(**data)
        results.append(result_traffic_light)

        metric = metrics_mapping["f1_score"](
            model_name=self.config["model_name"],
            task="binary"
        )
        f1_test = FTest(self.estimator, metric)
        result, result_traffic_light = f1_test.transform(**data)
        results.append(result_traffic_light)

        ciec_test = CIECTest(self.estimator)
        result, result_traffic_light = ciec_test.transform(**data)
        results.append(result_traffic_light)




        if "red" in results:
            self.test_results["Overall"] = "red"
        elif "yellow" in results:
            self.test_results["Overall"] = "yellow"
        else:
            self.test_results["Overall"] = "green"
        results.append(str(self.test_results["Overall"]))

        return results