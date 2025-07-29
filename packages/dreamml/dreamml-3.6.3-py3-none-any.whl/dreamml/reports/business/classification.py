import os
from typing import Optional, List, Dict, Tuple, Any

import numpy as np
import shap
from matplotlib import pyplot as plt
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from dreamml.reports._classification_metrics import (
    CalculateDataStatistics,
    CalculateDetailedMetrics,
)
from dreamml.reports.reports.style_utils import set_row_auto_height
from dreamml.validation._task_description import TaskDescriptionTest
from dreamml.validation.classification import ValidationReport

from dreamml.logging import get_logger
from dreamml.utils.prepare_artifacts_config import ClassificationArtifactsConfig
from dreamml.validation.classification._detailed_model_statistics import (
    ModelDetailedStatistics,
)
from dreamml.visualization.plots import (
    plot_quality_dynamics_per_segment_graph,
    plot_data_decile_statistics_graph,
)

_logger = get_logger(__name__)


def prepare_artifacts_config(config: dict) -> Tuple[Any, Any]:
    """
    Prepares the artifacts configuration and evaluation set based on the provided configuration dictionary.

    Args:
        config (dict): Configuration dictionary for preparing artifacts.

    Returns:
        Tuple[Any, Any]: A tuple containing the artifacts configuration and the evaluation set.

    Raises:
        Any relevant exception that may occur during the preparation of artifacts.
    """
    prepare_artifacts = ClassificationArtifactsConfig(config=config)
    artifacts_config, eval_set = prepare_artifacts.prepare_artifacts_config()

    artifacts_config["path_to_save"] = os.path.join(
        prepare_artifacts.experiment_dir_path,
        "docs",
        f"business_report_{prepare_artifacts.model_name}.xlsx",
    )

    return artifacts_config, eval_set


class ClassificationBusinessReport(ValidationReport):
    """
    A report generator for classification models tailored for business validation and stakeholders.

    Inherits from:
        ValidationReport
    """

    def __init__(
        self,
        config,
        estimator,
        metric_name,
        vectorizer: callable = None,
        path_to_save: str = "./business_report.xlsx",
        used_features: Optional[List[str]] = None,
        categorical_features: Optional[List[str]] = None,
        create_file: bool = True,
        metric_col_name: str = "gini",
        metric_params: dict = None,
        images_dir_path: str = None,
        task: str = "binary",
        subtask: str = "tabular",  # [tabular, nlp]
        multiclass_artifacts: Optional[Dict] = None,
        custom_model: bool = False,
        user_config=None,
        text_column: str = None,
        text_preprocessed_column: str = None,
        group_column: str = None,
        time_column: str = None,
        create_pdf: bool = False,
    ):
        """
        Initializes the ClassificationBusinessReport with the provided configuration and parameters.

        Args:
            config: Configuration settings for the report.
            estimator: The machine learning estimator/model.
            metric_name: Name of the metric used for evaluation.
            vectorizer (callable, optional): Function to vectorize input data. Defaults to None.
            path_to_save (str, optional): Path where the report will be saved. Defaults to "./business_report.xlsx".
            used_features (Optional[List[str]], optional): List of features used in the model. Defaults to None.
            categorical_features (Optional[List[str]], optional): List of categorical features. Defaults to None.
            create_file (bool, optional): Flag to create the report file. Defaults to True.
            metric_col_name (str, optional): Name of the metric column. Defaults to "gini".
            metric_params (dict, optional): Parameters for the metric. Defaults to None.
            images_dir_path (str, optional): Directory path to save images. Defaults to None.
            task (str, optional): Type of task (e.g., "binary"). Defaults to "binary".
            subtask (str, optional): Subtask type (e.g., "tabular", "nlp"). Defaults to "tabular".
            multiclass_artifacts (Optional[Dict], optional): Artifacts for multiclass tasks. Defaults to None.
            custom_model (bool, optional): Flag indicating if a custom model is used. Defaults to False.
            user_config (optional): User-specific configuration. Defaults to None.
            text_column (str, optional): Name of the text column. Defaults to None.
            text_preprocessed_column (str, optional): Name of the preprocessed text column. Defaults to None.
            group_column (str, optional): Name of the group column. Defaults to None.
            time_column (str, optional): Name of the time column. Defaults to None.
            create_pdf (bool, optional): Flag to create a PDF version of the report. Defaults to False.

        Raises:
            Any relevant exception that may occur during initialization.
        """
        super().__init__(
            config=config,
            estimator=estimator,
            metric_name=metric_name,
            vectorizer=vectorizer,
            path_to_save=path_to_save,
            used_features=used_features,
            categorical_features=categorical_features,
            create_file=create_file,
            metric_col_name=metric_col_name,
            metric_params=metric_params,
            images_dir_path=images_dir_path,
            task=task,
            subtask=subtask,
            multiclass_artifacts=multiclass_artifacts,
            custom_model=custom_model,
            user_config=user_config,
            text_column=text_column,
            text_preprocessed_column=text_preprocessed_column,
            group_column=group_column,
            time_column=time_column,
            create_pdf=create_pdf,
        )
        self._init_styles()

    def create_report(self, **data):
        """
        Generates the complete business report by creating all necessary pages and sections.

        Args:
            **data: Arbitrary keyword arguments containing data required for report generation.

        Raises:
            Any relevant exception that may occur during report creation.
        """
        self._create_business_task_description_page(**data)
        self._create_data_samples_statistics_page(**data)
        self._create_pipeline_description_page(**data)
        self._create_model_params_page(**data)
        self._create_detailed_stats_page(**data)
        self._create_quality_dynamics_page(**data)
        self._create_data_distribution_page(**data)
        self._create_features_importance_page(**data)

        _logger.info("Calculating test results.")
        self._create_data_quality_page(**data)
        self._create_model_quality_page(**data)
        self._create_model_calibration_page(**data)

        if "OOT" in data:
            self._create_model_stability_page(**data)
        else:
            self._create_model_stability_page_without_oot(**data)

        drop_sheets = [
            "Data Quality (PSI Analysis)",
            "Model Quality (Accuracy)",
            "Model Calibration",
            "Model Stability",
        ]

        for sheet in drop_sheets:
            del self.wb[sheet]
            del self.sheets[sheet]

        self._count_total_result()  # Fill in the summary results table
        self._create_total_result_page()

        for sheet_name, ws in self.sheets.items():
            set_row_auto_height(ws)

        self.writer.save()
        self.create_pdf()

    def _init_styles(self):
        """
        Initializes and defines the styles used in the Excel report, such as cell formats, fonts, and borders.

        Raises:
            Any relevant exception that may occur during style initialization.
        """
        cell_format = NamedStyle(name="cell_format")
        cell_format.font = Font(color="808080")
        cell_format.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
        brd = Side(border_style="thin", color="000000")
        cell_format.border = Border(left=brd, right=brd, top=brd, bottom=brd)

        row_title_format = NamedStyle(name="row_title_format")
        row_title_format.font = Font(bold=True)
        row_title_format.alignment = Alignment(
            wrap_text=True, horizontal="center", vertical="center"
        )
        brd = Side(border_style="thin", color="000000")
        row_title_format.border = Border(left=brd, right=brd, top=brd, bottom=brd)

        col_title_format = NamedStyle(name="col_title_format")
        col_title_format.font = Font(bold=True)
        col_title_format.fill = PatternFill(
            start_color="00CC99", end_color="00CC99", fill_type="solid"
        )
        col_title_format.alignment = Alignment(horizontal="center", vertical="center")
        brd = Side(border_style="thin", color="000000")
        col_title_format.border = Border(left=brd, right=brd, top=brd, bottom=brd)

        self.wb.add_named_style(cell_format)
        self.wb.add_named_style(row_title_format)
        self.wb.add_named_style(col_title_format)

    def _create_business_task_description_page(self, **data):
        """
        Creates the Business Task Description page in the report, outlining the model's purpose and related details.

        Args:
            **data: Arbitrary keyword arguments containing data required for this page.

        Raises:
            Any relevant exception that may occur during page creation.
        """
        _logger.info("1. Gathering statistics about the business task.")
        sheet_name = "1. Task & Sample"

        ws = self.wb.create_sheet(title=sheet_name)
        self.sheets[sheet_name] = ws
        self.sheet_descriptions[sheet_name] = ""

        business_task_text = (
            "<Enter the model name.>\n"
        )
        task_description_text = (
            "<Enter the model description: modeling goal.>"
        )
        task_type_text = self.artifacts_config["task"]
        data_selection_date_text = (
            "<Specify the data collection date>\n" "for example: 31.01.2024"
        )
        target_description_text = "<Enter the definition of the target event/variable.>"
        data_selection_text = (
            "<Enter the population selection criteria: filters, exclusions.>\n"
        )
        used_algo_text = f"<DreamML {self.artifacts_config['estimator'].__class__.__name__}>"

        ws.merge_cells("B2:E2")
        ws.merge_cells("B3:E3")
        ws.merge_cells("B4:E4")
        ws.merge_cells("B5:E5")
        ws.merge_cells("B6:E6")
        ws.merge_cells("B7:E7")
        ws.merge_cells("B8:E8")
        ws.merge_cells("B9:E9")

        ws["B2"] = business_task_text
        ws["B3"] = task_description_text
        ws["B4"] = task_type_text
        ws["B5"] = target_description_text
        ws["B6"] = data_selection_date_text
        ws["B7"] = data_selection_text
        ws["B8"] = used_algo_text

        metrics = {self.metric_name, "gini", "precision_recall_auc"}
        ws["B9"] = ", ".join(metrics)

        ws["A1"] = "1. Task Description"
        ws["A1"].font = Font(size=22, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40

        ws["A2"] = "Model Name"
        ws["A3"] = "Task Description"
        ws["A4"] = "Task Type"
        ws["A5"] = "Target Variable"
        ws["A6"] = "Data Collection Date"
        ws["A7"] = "Observation Selection"
        ws["A8"] = "ML Algorithm"
        ws["A9"] = "Quality Metrics"

        for row in ws["B2":"E9"]:
            for cell in row:
                cell.style = "cell_format"
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
        for row in ws["A2":"A9"]:
            for cell in row:
                cell.style = "row_title_format"

        ws.column_dimensions["A"].width = 22
        for i in range(2, 6):
            ws.column_dimensions[get_column_letter(i)].width = 18

    def _create_data_samples_statistics_page(self, **data):
        """
        Creates the Data Samples Statistics page, presenting statistical information about the dataset.

        Args:
            **data: Arbitrary keyword arguments containing data required for this page.

        Raises:
            Any relevant exception that may occur during page creation.
        """
        _logger.info("2. Gathering statistics about the sample.")
        sheet_name = "1. Task & Sample"

        features = data["train"][0].columns.to_series()
        transformer = CalculateDataStatistics(
            None,
            features,
            self.config,
            task=self.artifacts_config["task"],
            business=True,
        )

        result = transformer._calculate_samples_stats(**data)

        result.to_excel(
            self.writer,
            startrow=11,
            sheet_name=sheet_name,
            index=False,
            float_format="%.2f",
        )

        ws = self.sheets[sheet_name]

        ws["A11"] = "2. Sample Statistics"
        ws["A11"].font = Font(size=22, bold=True)
        ws["A11"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[11].height = 40

        for row in ws["A13" :f"G{12+len(result)}"]:
            for cell in row:
                cell.style = "cell_format"
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
        for row in ws["A12":"G12"]:
            for cell in row:
                cell.style = "row_title_format"

        ws.column_dimensions["A"].width = 22
        for i in range(2, 8):
            ws.column_dimensions[get_column_letter(i)].width = 18

    def _create_pipeline_description_page(self, **data):
        """
        Creates the Pipeline Description page, detailing all stages of the modeling pipeline.

        Args:
            **data: Arbitrary keyword arguments containing data required for this page.

        Raises:
            Any relevant exception that may occur during page creation.
        """
        _logger.info("3. Gathering statistics about the used pipeline.")
        sheet_name = "3. Pipeline"
        self.sheet_descriptions[sheet_name] = (
            "Page with the description of all pipeline stages"
        )
        pst = TaskDescriptionTest(self.artifacts_config, self.validation_test_config)
        stats = pst._create_description(**data)
        stats.to_excel(self.writer, sheet_name=sheet_name, index=False, startrow=1)

        ws = self.sheets[sheet_name]

        ws["A1"] = "3. Modeling Pipeline"
        ws["A1"].font = Font(size=22, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 130

        for row in ws["A3":"B9"]:
            for cell in row:
                cell.style = "cell_format"
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
        for row in ws["A2":"B2"]:
            for cell in row:
                cell.style = "row_title_format"

    def _create_model_params_page(self, **data):
        """
        Creates the Model Parameters page, listing all hyperparameters of the model.

        Args:
            **data: Arbitrary keyword arguments containing data required for this page.

        Raises:
            Any relevant exception that may occur during page creation.
        """
        _logger.info("4. Gathering model hyperparameters.")
        sheet_name = "4. Hyperparameters & 5. Quality"
        self.sheet_descriptions[sheet_name] = ""

        params = self.estimator.get_estimator_params
        params["Setup"] = ""
        params.index = np.arange(1, len(params) + 1)

        params.to_excel(
            self.writer,
            sheet_name=sheet_name,
            index=True,
            startrow=1,
        )

        ws = self.sheets[sheet_name]

        ws["A1"] = "4. Model Hyperparameters"
        ws["A1"].font = Font(size=22, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40

        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 10

        for row in ws["B3" :f"D{2+len(params)}"]:
            for cell in row:
                cell.style = "cell_format"
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
        for row in ws["B2":"D2"]:
            for cell in row:
                cell.style = "row_title_format"

        for i in range(3, 3 + len(params)):
            ws.row_dimensions[i].height = 14

    def _create_detailed_stats_page(self, **data):
        """
        Creates the Detailed Statistics page, providing an in-depth analysis of the model's quality.

        Args:
            **data: Arbitrary keyword arguments containing data required for this page.

        Raises:
            Any relevant exception that may occur during page creation.
        """
        _logger.info("5. Detailed calculation of model quality.")
        sheet_name = "4. Hyperparameters & 5. Quality"

        stats_test = ModelDetailedStatistics(
            self.artifacts_config,
            self.validation_test_config,
            self.metric_name,
            self.metric_col_name,
            self.metric_params,
            business=True,
        )
        stats = stats_test.create_report(**data)

        ws = self.sheets[sheet_name]

        ws["F1"] = "5. Model Quality (Gini)"
        ws["F1"].font = Font(size=22, bold=True)
        ws["F1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40

        stats.to_excel(
            self.writer,
            sheet_name=sheet_name,
            startrow=1,
            startcol=5,
            index=False,
        )

        for row in ws["F3" :f"{get_column_letter(5+stats.shape[1])}{2+stats.shape[0]}"]:
            for cell in row:
                cell.style = "cell_format"
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

        for row in ws[
            "G3" :f"{get_column_letter(5 + stats.shape[1])}{2 + stats.shape[0]}"
        ]:
            for cell in row:
                cell.style = "cell_format"
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
                cell.number_format = "0.0%"
        for row in ws["F2" :f"{get_column_letter(5+stats.shape[1])}2"]:
            for cell in row:
                cell.style = "row_title_format"

        picture_name = f"business_model_quality_per_segment.png"
        path_to_picture = os.path.join(self.images_dir_path, picture_name)
        if os.path.exists(path_to_picture):
            img = Image(path_to_picture)
            img.anchor = f"F{2+stats.shape[0]+2}"
            ws.add_image(img)

    def _create_quality_dynamics_page(self, **data):
        """
        Creates the Quality Dynamics page, visualizing the model's performance over time.

        Args:
            **data: Arbitrary keyword arguments containing data required for this page.

        Raises:
            Any relevant exception that may occur during page creation.
        """
        sheet_name = "6. Quality Dynamics"

        time_column = self.artifacts_config["time_column"]
        if time_column is None:
            _logger.warning(
                f'The page "{sheet_name}" will not be included in the report due to the absence of the time_column parameter.'
            )
            return

        _logger.info("6. Calculating model quality dynamics.")

        ws = self.wb.create_sheet(title=sheet_name)
        self.sheets[sheet_name] = ws
        self.sheet_descriptions[sheet_name] = ""

        ws["A1"] = "6. Model Quality Metric Dynamics"
        ws["A1"].font = Font(size=22, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40

        group_column = self.artifacts_config["group_column"]

        if group_column is None:
            unique_groups = [""]
        else:
            unique_groups = set()
            for sample_name in data:
                unique_groups.update(data[sample_name][0][group_column].unique())

            unique_groups = sorted(list(unique_groups))

        target_per_group = {}
        time_per_group = {}
        for group in unique_groups:
            target_per_group[group] = {}
            time_per_group[group] = {}

            for sample_name in data:
                x, y_true = data[sample_name]

                if group_column is None:
                    x_group = x
                    y_true_group = y_true
                else:
                    x_group = x[x[group_column] == group]
                    y_true_group = y_true[x[group_column] == group]

                if len(x_group) == 0:
                    target_per_group[group][sample_name] = ([], [])
                    time_per_group[group][sample_name] = []
                    continue

                time = x_group[time_column]

                y_pred_group = self.estimator.transform(x_group)

                target_per_group[group][sample_name] = (y_true_group, y_pred_group)
                time_per_group[group][sample_name] = time

        picture_name = f"business_quality_dynamics_segment.png"

        save_path = os.path.join(self.images_dir_path, picture_name)
        plot_quality_dynamics_per_segment_graph(
            time_per_group, target_per_group, save_path=save_path
        )

        path_to_picture = os.path.join(self.images_dir_path, picture_name)
        if os.path.exists(path_to_picture):
            img = Image(path_to_picture)
            img.anchor = f"A2"
            ws.add_image(img)

    def _create_data_distribution_page(self, **data):
        """
        Creates the Data Distribution page, showing the distribution of data across deciles.

        Args:
            **data: Arbitrary keyword arguments containing data required for this page.

        Raises:
            Any relevant exception that may occur during page creation.
        """
        _logger.info("7. Calculating data distribution.")
        sheet_name = "7. Distribution"

        ws = self.wb.create_sheet(title=sheet_name)
        self.sheets[sheet_name] = ws
        self.sheet_descriptions[sheet_name] = ""

        ws["A1"] = "7. Distribution by Deciles"
        ws["A1"].font = Font(size=22, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40

        group_column = self.artifacts_config["group_column"]

        if group_column is None:
            unique_groups = [""]
        else:
            unique_groups = set()
            for sample_name in data:
                unique_groups.update(data[sample_name][0][group_column].unique())

            unique_groups = sorted(list(unique_groups))

        n_bins = 10

        decile_index = [
            f"[{int(100 * i/n_bins)}-{int(100 * (i+1)/n_bins)}]" for i in range(n_bins)
        ]

        x, y_true = data["test"]
        start_row = 1
        for idx, group in enumerate(unique_groups):
            ws[f"A{start_row+1}"] = f"Segment {group}"
            ws[f"A{start_row+1}"].font = Font(size=14, bold=True)
            ws[f"A{start_row+1}"].alignment = Alignment(
                horizontal="left", vertical="bottom"
            )
            ws.row_dimensions[start_row + 1].height = 30

            start_row += 1

            if group_column is None:
                x_group = x
                y_true_group = y_true
            else:
                x_group = x[x[group_column] == group]
                y_true_group = y_true[x[group_column] == group]

            if len(x_group) == 0:
                continue

            y_pred_group = self.estimator.transform(x_group)

            transformer = CalculateDetailedMetrics(
                n_bins, "gini", {}, task=self.artifacts_config["task"]
            )
            try:
                data_dist = transformer.transform(y_true_group, y_pred_group)
            except ValueError as e:
                _logger.warning(
                    f"Can't calculate data distribution for group {group}: {e}"
                )
                continue

            data_dist = data_dist.iloc[:n_bins]

            data_dist = data_dist[["cum # obs", "cum_eventrate"]]
            data_dist = data_dist.reset_index()

            data_dist["index"] = decile_index
            data_dist.columns = [
                "Decile (Model Predictions)",
                "Number of Observations",
                "Event-rate (Actual)",
            ]

            data_dist.to_excel(
                self.writer, sheet_name=sheet_name, startrow=start_row, index=False
            )
            for row in ws[
                f"A{start_row+2}" :f"{get_column_letter(data_dist.shape[1])}{start_row + 1 + data_dist.shape[0]}"
            ]:
                for cell in row:
                    cell.style = "cell_format"
            for row in ws[
                f"A{start_row+1}" :f"{get_column_letter(data_dist.shape[1])}{start_row+1}"
            ]:
                for cell in row:
                    cell.style = "row_title_format"
            for row in ws[f"A{start_row+2}" :f"A{start_row + 1 + data_dist.shape[0]}"]:
                for cell in row:
                    cell.style = "row_title_format"

            picture_name = f"data_decile_statistics_segment_{idx}.png"
            path_to_picture = os.path.join(self.images_dir_path, picture_name)
            plot_data_decile_statistics_graph(data_dist, save_path=path_to_picture)

            if os.path.exists(path_to_picture):
                img = Image(path_to_picture)
                img.anchor = f"{get_column_letter(data_dist.shape[1]+1)}{start_row+1}"
                ws.add_image(img)

            start_row += len(data_dist) + 1

    def _create_features_importance_page(self, **data):
        """
        Creates the Features Importance page, evaluating the importance of model features using SHAP values.

        Args:
            **data: Arbitrary keyword arguments containing data required for this page.

        Raises:
            Any relevant exception that may occur during page creation.
        """
        _logger.info("8. Evaluating feature importance based on SHAP values.")
        sheet_name = "8. Features"

        if "valid" in data:
            x, _ = data["valid"]
        elif "train" in data:
            x, _ = data["train"]
        elif "test" in data:
            x, _ = data["test"]
        elif "OOT" in data:
            x, _ = data["OOT"]

        shap_values, imp = self.estimator.get_shap_importance(x)
        imp.to_excel(self.writer, sheet_name=sheet_name, startrow=1, index=True)
        ws = self.sheets[sheet_name]

        ws["A1"] = "8. Model Feature Importance"
        ws["A1"].font = Font(size=22, bold=True)
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 40

        for row in ws["B3" :f"C{2+len(imp)}"]:
            for cell in row:
                cell.style = "cell_format"
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
        for row in ws["A2":"C2"]:
            for cell in row:
                cell.style = "row_title_format"
        for row in ws["A3" :f"A{2+len(imp)}"]:
            for cell in row:
                cell.style = "row_title_format"

        for i in range(3, 3 + len(imp)):
            ws.row_dimensions[i].height = 20

        ws.column_dimensions["A"].width = 4
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 12

        if isinstance(shap_values, np.ndarray):
            plt.clf()
            shap.initjs()
            shap.summary_plot(
                shap_values,
                features=x[self.estimator.used_features],
                feature_names=self.estimator.used_features,
                show=False,
            )
            plt.savefig(
                os.path.join(self.images_dir_path, "shap_summary_plot.png"),
                bbox_inches="tight",
                pad_inches=0.1,
            )
            plt.close()

            img = Image(os.path.join(self.images_dir_path, "shap_summary_plot.png"))
            img.anchor = f"D2"
            ws.add_image(img)

        self.sheet_descriptions[sheet_name] = "Sheet with feature importance evaluation based on SHAP values."

    def _create_total_result_page(self):
        """
        Creates the Total Result page, summarizing the outcomes of all validation tests.

        Raises:
            Any relevant exception that may occur during page creation.
        """
        _logger.info("9. Summarizing test results.")
        sheet_name = "9. Validation DML"

        self.total_result_df.to_excel(
            self.writer,
            sheet_name=sheet_name,
            startrow=0,
            startcol=0,
            index=False,
        )

        self.add_table_borders(
            self.total_result_df[
                [
                    "Test Block",
                    "N",
                    "Test Short Description",
                    "Test Details",
                    "OOS Usage",
                    "OOT Usage",
                    "Result",
                ]
            ],
            sheet_name=sheet_name,
            startcol=0,
        )
        self.add_header_color(
            self.total_result_df,
            sheet_name=sheet_name,
            color="00CC99",
            startcol=0,
        )
        self.add_traffic_light_color(sheet_name=sheet_name)

        ws = self.sheets[sheet_name]

        # Set column widths
        ws.column_dimensions[get_column_letter(1)].width = 33
        ws.column_dimensions[get_column_letter(2)].width = 5
        ws.column_dimensions[get_column_letter(3)].width = 110
        ws.column_dimensions[get_column_letter(4)].width = 10
        ws.column_dimensions[get_column_letter(5)].width = 15
        ws.column_dimensions[get_column_letter(6)].width = 15
        ws.column_dimensions[get_column_letter(7)].width = 10
        ws.column_dimensions[get_column_letter(8)].width = 10
        ws.column_dimensions[get_column_letter(9)].width = 15

        # Align text in cells
        for row in ws[2 : 2 + self.total_result_df.shape[0]]:
            for col in [1, 3, 4, 5, 6]:
                cell = row[col]
                cell.alignment = Alignment(horizontal="center")
        for row in ws[2 : 2 + self.total_result_df.shape[0]]:
            for col in [0]:
                cell = row[col]
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Set cell borders
        for row in [ws[5:5], ws[11:11], ws[14:14], ws[17:17], ws[25:25], ws[27:27]]:
            for col in range(7):
                cell = row[col]
                cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=Side(border_style="thin", color="000000"),
                )

        # Merge cells
        ws.merge_cells("A2:A5")
        ws.merge_cells("A6:A11")
        ws.merge_cells("A12:A14")
        ws.merge_cells("A15:A17")
        ws.merge_cells("A18:A25")