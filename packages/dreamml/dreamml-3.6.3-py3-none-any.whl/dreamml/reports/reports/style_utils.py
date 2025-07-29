from math import ceil

import openpyxl
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


def set_row_auto_height(ws: openpyxl.worksheet):
    """
    Automatically adjusts the height of each row in the given worksheet based on the content and font size.

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet for which to set the row heights.

    """
    for row_idx in range(0, ws.max_row):
        max_size = 0
        for col_idx in range(ws.max_column):
            size = ws[row_idx + 1][col_idx].font.sz or 11
            max_size = max(max_size, size)

        new_height = _get_height_for_row(ws, row_idx, font_size=max_size)
        height = ws.row_dimensions[row_idx + 1].height

        ws.row_dimensions[row_idx + 1].height = new_height if height is None else height


def _get_column_width_by_cell(sheet: Worksheet, cell: Cell):
    """
    Calculates and returns the width of the column for the specified cell, accounting for merged cells.

    Args:
        sheet (Worksheet): The worksheet containing the cell.
        cell (Cell): The cell for which to determine the column width.

    Returns:
        float | None: The width of the column. If the cell is part of a merged range, the combined width of all merged columns is returned. If the column width is not set, returns `None`.
    """
    if cell.coordinate not in sheet.merged_cells:
        return sheet.column_dimensions[get_column_letter(cell.column)].width
    ranges = [
        range_ for range_ in sheet.merged_cells.ranges if cell.coordinate in range_
    ]
    if not ranges:
        return sheet.column_dimensions[get_column_letter(cell.column)].width
    range_ = ranges[0]
    return sum(
        sheet.column_dimensions[get_column_letter(col[1])].width for col in range_.top
    )


def _get_height_for_row(sheet: Worksheet, row_number: int, font_size: int = 12):
    """
    Determines the appropriate height for a specific row, considering merged columns and multiline cell content.

    Args:
        sheet (Worksheet): The worksheet containing the row.
        row_number (int): The zero-based index of the row for which to calculate the height.
        font_size (int, optional): The font size to use for height calculations. Defaults to 12.

    Returns:
        float: The calculated height for the row.
    """
    font_params = {"factor": 1.5, "height": font_size * 1.3}

    row = list(sheet.rows)[row_number]
    height = font_params["height"]
    max_height = 0
    for cell in row:
        words_count_at_one_row = (
            _get_column_width_by_cell(sheet, cell) / font_params["factor"]
        )
        # Calculating amount of lines
        lines = ceil(len(str(cell.value)) / words_count_at_one_row) if words_count_at_one_row else 1
        new_lines = len(str(cell.value).split("\n")) if cell.value else 1
        # If the number of characters is low but the cell contains multiple lines, count both
        lines = max(lines, new_lines)
        height = max(height, lines * font_params["height"])
        max_height = max(max_height, height)

    return max_height