import os
import pandas as pd
import openpyxl
import portalocker

from portalocker.exceptions import LockException
from typing import Optional, Union, Dict, Set

from ..utility_base import UtilityBase
from ..logger import Logger, LogWrapper


class ExcelWriter(UtilityBase):
    def __init__(
        self,
        path: str,
        verbose: bool = False,
        logger: Optional[Union[Logger, LogWrapper]] = None,
        log_level: Optional[int] = None
    ) -> None:
        """
        Initializes an ExcelWriter instance.

        :param path: Path to the Excel file.
        :param verbose: Flag to enable verbose output.
        :param logger: Optional logger instance.
        :param log_level: Optional logging level.
        """
        super().__init__(verbose, logger, log_level)

        self.path: str = path
        self.df_cache: Dict[str, pd.DataFrame] = {}
        self.modified_rows: Dict[str, Set[int]] = {}
        self.file_handle: Optional[object] = None

    def load_file(self, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        Loads the Excel file, locks it for exclusive access, and returns the specified sheet as a DataFrame.

        :param sheet_name: Optional sheet name to load. Defaults to the first sheet if not provided.
        :return: A deep copy of the loaded DataFrame.
        :raises FileNotFoundError: If the Excel file does not exist.
        :raises ValueError: If the specified sheet does not exist in the workbook.
        """
        if not os.path.exists(self.path):
            raise FileNotFoundError(f"Excel file not found: {self.path}")

        # Lock file
        try:
            self.file_handle = open(self.path, "rb+")
            portalocker.lock(self.file_handle, portalocker.LOCK_EX)
        except (OSError, LockException) as e:
            if self.file_handle:
                self.file_handle.close()
                self.file_handle = None
            self.logger.critical(f"Failed to lock file `{self.path}`: {e}", exc_info=True)
            raise 

        excel_file = pd.ExcelFile(self.path, engine="openpyxl")
        target_sheet = sheet_name or excel_file.sheet_names[0]

        if target_sheet not in excel_file.sheet_names:
            raise ValueError(f"Sheet '{target_sheet}' not found in Excel file.")

        df = excel_file.parse(sheet_name=target_sheet, header=1)
        self.df_cache[target_sheet] = df
        self.modified_rows[target_sheet] = set()

        self._log(f"Opened and locked file `{self.path}`, selected sheet `{target_sheet}`")

        return df.copy(deep=True)

    def modify_data(
        self,
        row_index: int,
        column_name: str,
        value: object,
        sheet_name: Optional[str] = None
    ) -> None:
        """
        Modifies a specific cell by row index and column name in a given sheet.

        :param row_index: Index of the row to modify.
        :param column_name: Name of the column to modify.
        :param value: New value to set.
        :param sheet_name: Optional name of the sheet. Uses the first loaded sheet if not provided.
        :raises ValueError: If the sheet is not loaded or the column does not exist.
        :raises IndexError: If the row index is out of range.
        """
        sheet = sheet_name or next(iter(self.df_cache.keys()), None)

        if not sheet or sheet not in self.df_cache:
            raise ValueError(f"Sheet '{sheet}' not loaded. Use load_file() first.")

        df = self.df_cache[sheet]

        if row_index < 0 or row_index >= len(df):
            raise IndexError(f"Row index {row_index} out of range for sheet '{sheet}'.")

        if column_name not in df.columns:
            raise ValueError(f"Column '{column_name}' not found in sheet '{sheet}'.")
        
        original = df.at[row_index, column_name]
        df.at[row_index, column_name] = value

        self._log(f"Modified data at Column: `{column_name}`, Row: `{row_index}`")
        self._log(f"Original value `{original}`")
        self._log(f"New value: `{value}`")

        self.modified_rows[sheet].add(row_index)

    def save_file(self) -> None:
        """
        Saves all modified cells back to their respective sheets in the Excel file.

        :raises ValueError: If there are no modifications to save or if any sheet is missing in the file.
        :raises RuntimeError: If no file is locked for writing.
        """
        if not self.modified_rows or not any(self.modified_rows.values()):
            raise ValueError("No modifications to save.")

        if self.file_handle is None:
            raise RuntimeError("Excel file not locked. Use load_file() before saving.")

        workbook = openpyxl.load_workbook(self.path)

        for sheet, rows in self.modified_rows.items():
            if sheet not in workbook.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found in Excel file.")

            ws = workbook[sheet]
            df = self.df_cache[sheet]

            for row_index in rows:
                for col_index, col_name in enumerate(df.columns):
                    new_value = df.at[row_index, col_name]
                    excel_row = row_index + 3  # Account for header row (header=1) + 1-based index
                    excel_col = col_index + 1  # openpyxl uses 1-based indexing
                    ws.cell(row=excel_row, column=excel_col, value=new_value)

        workbook.save(self.path)

        portalocker.unlock(self.file_handle)
        self.file_handle.close()
        self.file_handle = None

        self._log(f"Closed file `{self.path}`and released lock")

        self.modified_rows.clear()
        self.df_cache.clear()
